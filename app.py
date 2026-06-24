# -*- coding: utf-8 -*-
"""
TAFIROHA en ligne — application web (MVP)

Application WSGI pure Python (stdlib + Jinja2) :
- Authentification par session (cookie + table sessions en base)
- Rôle "admin" (cabinet comptable) : voit tous les clients
- Rôle "client" : voit uniquement ses propres exercices / historique
- Saisie de balance (import CSV) par exercice, pour les périodes N et N1
- Calcul du BILAN et du RESULTAT à partir des formules du fichier Excel
  d'origine (voir calc_engine.py), exactement comme la feuille
  "TABLE DE CONVERSION" du classeur.
"""
import csv
import io
import os
import re
import zipfile
from datetime import datetime
from urllib.parse import parse_qs, urlencode
from wsgiref.simple_server import make_server
from jinja2 import Environment, FileSystemLoader
import openpyxl

import db
import calc_engine as ce

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
jinja_env = Environment(loader=FileSystemLoader(os.path.join(BASE_DIR, "templates")))
STATIC_DIR = os.path.join(BASE_DIR, "static_files")


def render(name, **ctx):
    tpl = jinja_env.get_template(name)
    return tpl.render(**ctx)


def fmt(v):
    if v is None or v == "":
        return ""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return v
    return "{:,.0f}".format(v).replace(",", " ")


def slug(v):
    s = re.sub(r"[^a-z0-9]+", "-", str(v).lower()).strip("-")
    return s or "section"


def fmt_pct(v):
    """Formate une variation en pourcentage. Les cellules de variation en %
    du classeur source (formules du type =(F-G)/G) renvoient une fraction
    (0,23 pour 23%) — on la multiplie par 100 pour l'affichage."""
    if v is None or v == "":
        return ""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return v
    return "{:,.1f} %".format(v * 100).replace(",", " ").replace(".", ",")


# Feuilles imprimées en paysage dans le classeur Excel d'origine (orientation
# relevée via ws.page_setup.orientation sur chaque feuille du classeur) :
# on reproduit la même mise en page lors de l'impression/export PDF, une
# feuille = une page, dans l'orientation d'origine.
SHEET_LANDSCAPE = {
    "BILAN",
    "NOTE 3A", "NOTE 3B", "NOTE 3C", "NOTE 3C BIS", "NOTE 3D",
    "NOTE 4", "NOTE 6", "NOTE 7", "NOTE 8",
    "NOTE 13",
    "NOTE 15A", "NOTE 15B",
    "NOTE 16A", "NOTE 16B BIS",
    "NOTE 17", "NOTE 18", "NOTE 19",
    "NOTE 27B", "NOTE 28",
    "NOTE 31", "NOTE 32", "NOTE 33",
    "COMP-TVA (2)",
    "SUPPL2", "SUPPL3", "SUPPL4", "SUPPL5",
}


def is_landscape(sheet):
    return str(sheet).strip() in SHEET_LANDSCAPE


jinja_env.filters["fmt"] = fmt
jinja_env.filters["slug"] = slug
jinja_env.filters["pct"] = fmt_pct
jinja_env.filters["landscape"] = is_landscape


# ----------------------------------------------------------------------------
# SOMMAIRE — liste canonique des feuilles imprimables, dans l'ordre exact du
# classeur Excel (releve via wb.sheetnames) : GARDE, RECEVABILITE, NOTE36,
# NOTE36 Suite, FICHE R1-3, BILAN/RESULTAT/TFT, FICHE R4, NOTE 1-39, module
# DGI-INS. Chaque entree = (cle_feuille, libelle, id_carte_html). La feuille
# "Sommaire" elle-meme n'est pas une page imprimee (c'est le panneau de
# selection) et n'apparait donc pas dans cette liste.
# id_carte_html doit correspondre exactement a l'id="..." de la <div class=
# "card"> de la feuille dans exercice_view.html (cf. filtre slug pour les
# feuilles dont l'id est derive automatiquement du nom de feuille).
_CARD_ID_OVERRIDES = {
    "BILAN ACTIF": "bilan-actif",
    "BILAN PASSIF": "bilan-passif",
    "RESULTAT": "resultat",
    "TFT": "tft",
}


def _card_id(sheet_key):
    return _CARD_ID_OVERRIDES.get(sheet_key, slug(sheet_key))


SOMMAIRE_SHEETS_RAW = [
    ("GARDE", "GARDE — Page de garde"),
    ("RECEVABILITE", "RECEVABILITE"),
    ("NOTE36 (TABLE DES CODES)", "NOTE 36 — Table des codes"),
    ("NOTE36 Suite (Nomenclature)", "NOTE 36 suite — Nomenclature"),
    ("FICHE R1", "FICHE R1 — Identification 1"),
    ("FICHE R2", "FICHE R2 — Identification 2"),
    ("FICHE R3", "FICHE R3 — Identification 3"),
    ("BILAN ACTIF", "BILAN — Actif"),
    ("BILAN PASSIF", "BILAN — Passif"),
    ("RESULTAT", "Compte de résultat"),
    ("TFT", "Tableau des flux de trésorerie"),
    ("FICHE R4", "FICHE R4 — Récapitulatif des notes"),
    ("NOTE 1", "Note 1"), ("NOTE 2", "Note 2"),
    ("NOTE 3A", "Note 3A"), ("NOTE 3B", "Note 3B"), ("NOTE 3C", "Note 3C"),
    ("NOTE 3C BIS", "Note 3C bis"), ("NOTE 3D", "Note 3D"), ("NOTE 3E", "Note 3E"),
    ("NOTE 4", "Note 4"), ("NOTE 5", "Note 5"), ("NOTE 6", "Note 6"),
    ("NOTE 7", "Note 7"), ("NOTE 8", "Note 8"), ("NOTE 8A", "Note 8A"),
    ("NOTE 8B", "Note 8B"), ("NOTE 8C", "Note 8C"), ("NOTE 9", "Note 9"),
    ("NOTE 10", "Note 10"), ("NOTE 11", "Note 11"), ("NOTE 12", "Note 12"),
    ("NOTE 13", "Note 13"), ("NOTE 14", "Note 14"), ("NOTE 15A", "Note 15A"),
    ("NOTE 15B", "Note 15B"), ("NOTE 16A", "Note 16A"), ("NOTE 16B", "Note 16B"),
    ("NOTE 16B BIS", "Note 16B bis"), ("NOTE 16C", "Note 16C"),
    ("NOTE 17", "Note 17"), ("NOTE 18", "Note 18"), ("NOTE 19", "Note 19"),
    ("NOTE 20", "Note 20"), ("NOTE 21", "Note 21"), ("NOTE 22", "Note 22"),
    ("NOTE 23", "Note 23"), ("NOTE 24", "Note 24"), ("NOTE 25", "Note 25"),
    ("NOTE 26", "Note 26"), ("NOTE 27A", "Note 27A"), ("NOTE 27B", "Note 27B"),
    ("NOTE 28", "Note 28"), ("NOTE 29", "Note 29"), ("NOTE 30", "Note 30"),
    ("NOTE 31", "Note 31"), ("NOTE 32", "Note 32"), ("NOTE 33", "Note 33"),
    ("NOTE 34", "Note 34"), ("NOTE 35", "Note 35"), ("NOTE 37", "Note 37"),
    ("NOTE 38", "Note 38"), ("NOTE 39", "Note 39"),
    ("GARDE (DGI-INS)", "GARDE (DGI-INS)"),
    ("NOTES DGI - INS", "NOTES DGI - INS (récapitulatif)"),
    ("COMP-CHARGES", "COMP-CHARGES"),
    ("COMP-TVA", "COMP-TVA"),
    ("COMP-TVA (2)", "COMP-TVA (2)"),
    ("SUPPL1", "SUPPL1"), ("SUPPL2", "SUPPL2"), ("SUPPL3", "SUPPL3"),
    ("SUPPL4", "SUPPL4"), ("SUPPL5", "SUPPL5"), ("SUPPL6", "SUPPL6"),
    ("SUPPL7", "SUPPL7"),
    ("COMMENTAIRE", "Commentaires"),
]

SOMMAIRE_SHEETS = [(key, label, _card_id(key)) for key, label in SOMMAIRE_SHEETS_RAW]


def load_sommaire_selection(conn, exercice_id):
    """Renvoie {sheet: bool} — feuille cochee comme applicable (A) ou non
    (N/A) sur la page Sommaire. Par defaut (pas encore enregistre), toutes
    les feuilles sont consideres applicables, comme dans le classeur d'origine."""
    rows = conn.execute(
        "SELECT sheet, applicable FROM sommaire_selection WHERE exercice_id=?",
        (exercice_id,),
    ).fetchall()
    sel = {r["sheet"]: bool(r["applicable"]) for r in rows}
    return {key: sel.get(key, True) for key, _label, _cid in SOMMAIRE_SHEETS}


def save_sommaire_selection(conn, exercice_id, checked_sheets):
    checked = set(checked_sheets)
    for key, _label, _cid in SOMMAIRE_SHEETS:
        applicable = 1 if key in checked else 0
        conn.execute(
            "INSERT INTO sommaire_selection (exercice_id, sheet, applicable) VALUES (?,?,?) "
            "ON CONFLICT(exercice_id, sheet) DO UPDATE SET applicable=excluded.applicable",
            (exercice_id, key, applicable),
        )
    conn.commit()


class Request:
    def __init__(self, environ):
        self.environ = environ
        self.method = environ.get("REQUEST_METHOD", "GET")
        self.path = environ.get("PATH_INFO", "/")
        self.query = parse_qs(environ.get("QUERY_STRING", ""))
        self.cookies = {}
        cookie_header = environ.get("HTTP_COOKIE", "")
        for part in cookie_header.split(";"):
            part = part.strip()
            if "=" in part:
                k, v = part.split("=", 1)
                self.cookies[k] = v
        self.form = {}
        self._raw_body = b""
        if self.method == "POST":
            length = int(environ.get("CONTENT_LENGTH") or 0)
            self._raw_body = environ["wsgi.input"].read(length) if length else b""
            ctype = environ.get("CONTENT_TYPE", "")
            if "multipart/form-data" not in ctype:
                self.form = {
                    k: v[0] for k, v in parse_qs(self._raw_body.decode("utf-8", "replace")).items()
                }

    def qget(self, key, default=None):
        v = self.query.get(key)
        return v[0] if v else default


class Response:
    def __init__(self, body="", status="200 OK", headers=None, content_type="text/html; charset=utf-8"):
        self.body = body.encode("utf-8") if isinstance(body, str) else body
        self.status = status
        self.headers = headers or []
        self.headers.append(("Content-Type", content_type))


def redirect(location, set_cookie=None, delete_cookie=False):
    headers = [("Location", location)]
    if set_cookie:
        headers.append(("Set-Cookie", "session=%s; Path=/; HttpOnly" % set_cookie))
    if delete_cookie:
        headers.append(("Set-Cookie", "session=deleted; Path=/; HttpOnly; Max-Age=0"))
    return Response("", status="302 Found", headers=headers)


# ---------------------------------------------------------------- helpers --

def current_user(req, conn):
    token = req.cookies.get("session")
    return db.get_user_by_session(conn, token)


def require_login(req, conn):
    user = current_user(req, conn)
    return user


def parse_csv_balance(file_bytes):
    """CSV attendu : compte;designation;be_debit;be_credit;mvt_debit;mvt_credit;bs_debit;bs_credit"""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    # Normalise les fins de ligne (CRLF / CR seul / LF) avant de passer au
    # lecteur CSV : sans cela, csv.reader peut lever "new-line character
    # seen in unquoted field" sur des fichiers exportés depuis Excel/Mac
    # avec des retours chariot isolés (\r) ou mélangés à des \n.
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    sample = text[:1024]
    delim = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delim)
    rows = list(reader)
    if not rows:
        return []
    header = [h.strip().lower() for h in rows[0]]
    has_header = "compte" in header
    data_rows = rows[1:] if has_header else rows
    out = []
    for r in data_rows:
        if not r or not r[0].strip():
            continue
        r = r + [""] * (8 - len(r))
        out.append({
            "compte": r[0].strip(),
            "designation": r[1].strip(),
            "be_debit": r[2], "be_credit": r[3],
            "mvt_debit": r[4], "mvt_credit": r[5],
            "bs_debit": r[6], "bs_credit": r[7],
        })
    return out


# ------------------------------------------------- import xlsx officiel --
# Le fichier "TAFIROHA-DGI_Import.xlsx" (téléchargeable via /static/...) est
# le modèle officiel de balance à 6 colonnes : il contient les mêmes tables
# Excel BalanceN/BalanceN1/TFTN/TFTN1 que le classeur TAFIROHA d'origine,
# plus des formules de contrôle (équilibre, anomalies Vérif/SI/SF, cohérence
# comptes à éclater). On relit directement ces tables avec openpyxl et on
# recalcule les mêmes contrôles en Python (plus robuste que de dépendre des
# valeurs mises en cache par Excel).
XLSX_BALANCE_TABLES = {"N": "BalanceN", "N1": "BalanceN1"}
XLSX_TFT_TABLES = {"N": "TFTN", "N1": "TFTN1"}
XLSX_TABLE_COLS = ("Compte", "Désignation", "BE_Debit", "BE_Credit",
                   "Mvt_Debit", "Mvt_Credit", "BS_Debit", "BS_Credit")


def _read_table_rows(wb, table_name):
    """Lit les lignes d'une table Excel nommée (BalanceN/BalanceN1/TFTN/TFTN1),
    quelle que soit la feuille qui la contient, et ne retient que les lignes
    où le compte est renseigné."""
    for ws in wb.worksheets:
        for tbl in getattr(ws, "tables", {}).values():
            if tbl.name != table_name:
                continue
            cells = ws[tbl.ref]
            header = [c.value for c in cells[0]]
            col_idx = {name: header.index(name) for name in XLSX_TABLE_COLS if name in header}
            if "Compte" not in col_idx:
                return []
            out = []
            for row in cells[1:]:
                compte = row[col_idx["Compte"]].value
                if compte in (None, ""):
                    continue

                def val(key):
                    i = col_idx.get(key)
                    v = row[i].value if i is not None else None
                    return v if v is not None else 0

                out.append({
                    "compte": str(compte).strip(),
                    "designation": str(val("Désignation") or ""),
                    "be_debit": val("BE_Debit"), "be_credit": val("BE_Credit"),
                    "mvt_debit": val("Mvt_Debit"), "mvt_credit": val("Mvt_Credit"),
                    "bs_debit": val("BS_Debit"), "bs_credit": val("BS_Credit"),
                })
            return out
    return []


def parse_xlsx_import(file_bytes):
    """Lit le fichier officiel d'import (tables BalanceN/BalanceN1/TFTN/TFTN1)
    et retourne {"balance": {"N":[...], "N1":[...]}, "tft": {"N":[...], "N1":[...]}}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    return {
        "balance": {p: _read_table_rows(wb, t) for p, t in XLSX_BALANCE_TABLES.items()},
        "tft": {p: _read_table_rows(wb, t) for p, t in XLSX_TFT_TABLES.items()},
    }


def _num0(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def compute_controles_xlsx(parsed):
    """Reproduit en Python les contrôles intégrés au fichier d'import officiel :
    équilibre de la balance (BE/Mvt/BS) et cohérence du détail des comptes à
    éclater (106, 154, 1984, 4781, 4791, 4816, 4817, 4818) entre la Balance et
    la table TFT. Retourne une liste de messages d'anomalie (vide si tout est
    conforme)."""
    alerts = []
    for periode, rows in parsed["balance"].items():
        be_d = sum(_num0(r["be_debit"]) for r in rows)
        be_c = sum(_num0(r["be_credit"]) for r in rows)
        mv_d = sum(_num0(r["mvt_debit"]) for r in rows)
        mv_c = sum(_num0(r["mvt_credit"]) for r in rows)
        bs_d = sum(_num0(r["bs_debit"]) for r in rows)
        bs_c = sum(_num0(r["bs_credit"]) for r in rows)
        if round(be_d, 2) != round(be_c, 2):
            alerts.append("Période %s : balance d'entrée non équilibrée (Débit %.2f / Crédit %.2f)." % (periode, be_d, be_c))
        if round(mv_d, 2) != round(mv_c, 2):
            alerts.append("Période %s : mouvements non équilibrés (Débit %.2f / Crédit %.2f)." % (periode, mv_d, mv_c))
        if round(bs_d, 2) != round(bs_c, 2):
            alerts.append("Période %s : balance de sortie non équilibrée (Débit %.2f / Crédit %.2f)." % (periode, bs_d, bs_c))

    racines = ["106", "154", "1984", "4781", "4791", "4816", "4817", "4818"]
    for periode in ("N", "N1"):
        bal_rows = parsed["balance"].get(periode, [])
        tft_rows = parsed["tft"].get(periode, [])
        for racine in racines:
            bal_total = sum(
                _num0(r["bs_debit"]) - _num0(r["bs_credit"])
                for r in bal_rows if r["compte"].startswith(racine)
            )
            tft_total = sum(
                _num0(r["bs_debit"]) - _num0(r["bs_credit"])
                for r in tft_rows if r["compte"].startswith(racine)
            )
            if round(bal_total, 2) != round(tft_total, 2):
                alerts.append(
                    "Période %s : détail du compte %s non conforme entre la balance (%.2f) et le détail TFT (%.2f)."
                    % (periode, racine, bal_total, tft_total)
                )
    return alerts


def _rows_to_csv_bytes(rows):
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["compte", "designation", "be_debit", "be_credit", "mvt_debit", "mvt_credit", "bs_debit", "bs_credit"])
    for r in rows:
        writer.writerow([r["compte"], r["designation"], r["be_debit"], r["be_credit"],
                          r["mvt_debit"], r["mvt_credit"], r["bs_debit"], r["bs_credit"]])
    return buf.getvalue().encode("utf-8-sig")


def _extract_multipart_file(req, field_name):
    body = req._raw_body
    ctype = req.environ.get("CONTENT_TYPE", "")
    boundary = ctype.split("boundary=")[-1].encode()
    parts = body.split(b"--" + boundary)
    needle = ('name="%s"' % field_name).encode()
    for part in parts:
        if b"filename=" in part and needle in part:
            header, _, content = part.partition(b"\r\n\r\n")
            return content.rstrip(b"\r\n--")
    return None


def handle_xlsx_to_csv(req):
    """Génère, à partir du fichier xlsx officiel uploadé, un zip de 4 CSV
    (balance_N/N1, tft_N/N1) au format attendu par l'import CSV existant —
    sans rien écrire en base. Retourne None si le fichier est illisible."""
    file_bytes = _extract_multipart_file(req, "fichier_xlsx_csv")
    if not file_bytes:
        return None
    try:
        parsed = parse_xlsx_import(file_bytes)
    except Exception:
        return None
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("balance_N.csv", _rows_to_csv_bytes(parsed["balance"]["N"]))
        zf.writestr("balance_N1.csv", _rows_to_csv_bytes(parsed["balance"]["N1"]))
        zf.writestr("tft_N.csv", _rows_to_csv_bytes(parsed["tft"]["N"]))
        zf.writestr("tft_N1.csv", _rows_to_csv_bytes(parsed["tft"]["N1"]))
    return Response(
        buf.getvalue(),
        content_type="application/zip",
        headers=[("Content-Disposition", 'attachment; filename="tafiroha_csv_import.zip"')],
    )


def load_balance(conn, exercice_id, periode):
    rows = conn.execute(
        "SELECT * FROM balance_lignes WHERE exercice_id=? AND periode=? ORDER BY compte",
        (exercice_id, periode),
    ).fetchall()
    return ce.build_balance_rows([dict(r) for r in rows])


def load_tft_detail(conn, exercice_id, periode):
    rows = conn.execute(
        "SELECT * FROM tft_detail_lignes WHERE exercice_id=? AND periode=? ORDER BY compte",
        (exercice_id, periode),
    ).fetchall()
    return ce.build_balance_rows([dict(r) for r in rows])


def load_note3_manual(conn, exercice_id):
    rows = conn.execute(
        "SELECT sheet, coord, valeur FROM note3_manuel WHERE exercice_id=?", (exercice_id,)
    ).fetchall()
    out = {}
    for r in rows:
        out.setdefault(r["sheet"], {})[r["coord"]] = r["valeur"]
    return out


def save_note3_manual(conn, exercice_id, sheet, cell_values):
    for coord, raw in cell_values.items():
        val = ce._num(raw)
        conn.execute(
            "INSERT INTO note3_manuel (exercice_id, sheet, coord, valeur) VALUES (?,?,?,?) "
            "ON CONFLICT(exercice_id, sheet, coord) DO UPDATE SET valeur=excluded.valeur",
            (exercice_id, sheet, coord, val),
        )
    conn.commit()


def load_note_texte(conn, exercice_id):
    rows = conn.execute(
        "SELECT sheet, champ, texte FROM note_texte WHERE exercice_id=?", (exercice_id,)
    ).fetchall()
    out = {}
    for r in rows:
        out.setdefault(r["sheet"], {})[r["champ"]] = r["texte"]
    return out


def save_note_texte(conn, exercice_id, sheet, champ_values):
    for champ, texte in champ_values.items():
        conn.execute(
            "INSERT INTO note_texte (exercice_id, sheet, champ, texte) VALUES (?,?,?,?) "
            "ON CONFLICT(exercice_id, sheet, champ) DO UPDATE SET texte=excluded.texte",
            (exercice_id, sheet, champ, texte),
        )
    conn.commit()


# Definitions des lignes des NOTE 3A-3E (immobilisations), reprises de la
# structure du classeur Excel : (ligne, libelle, est_un_sous_total)
NOTE3A_LINES = [
    (11, "IMMOBILISATIONS INCORPORELLES", True),
    (12, "Frais de développement et de prospection", False),
    (13, "Brevets, licences, logiciels, et droits similaires", False),
    (14, "Fonds commercial et droit au bail", False),
    (15, "Autres immobilisations incorporelles", False),
    (16, "IMMOBILISATIONS CORPORELLES", True),
    (17, "Terrains hors immeuble de placement", False),
    (18, "Terrains - immeuble de placement", False),
    (19, "Bâtiments hors immeuble de placement", False),
    (20, "Bâtiments - immeuble de placement", False),
    (21, "Aménagements, agencements et installations", False),
    (22, "Matériel, mobilier et actifs biologiques", False),
    (23, "Matériel de transport", False),
    (24, "AVANCES ET ACOMPTES VERSES SUR IMMOBILISATIONS", True),
    (25, "Avances et acomptes sur immobilisations incorporelles", False),
    (26, "Avances et acomptes sur immobilisations corporelles", False),
    (27, "IMMOBILISATIONS FINANCIERES", True),
    (28, "Titres de participation", False),
    (29, "Autres immobilisations financières", False),
    (30, "TOTAL GENERAL", True),
]
NOTE3A_COLS = [
    ("D", "Montant brut ouverture", False),
    ("E", "Acquisitions/Apports/Créations", False),
    ("F", "Virements de poste à poste (entrée)", True),
    ("G", "Réévaluation", True),
    ("H", "Cessions/Mises hors service", False),
    ("I", "Virements de poste à poste (sortie)", True),
    ("J", "Montant brut clôture", False),
]

NOTE3C_LINES = [
    (11, "Frais de développement et de prospection", False),
    (12, "Brevets, licences, logiciels et droits similaires", False),
    (13, "Fonds commercial et droit au bail", False),
    (14, "Autres immobilisations incorporelles", False),
    (15, "IMMOBILISATIONS INCORPORELLES", True),
    (16, "Terrains hors immeuble de placement", False),
    (17, "Terrains - immeuble de placement", False),
    (18, "Bâtiments hors immeuble de placement", False),
    (19, "Bâtiments - immeuble de placement", False),
    (20, "Aménagements, agencements et installations", False),
    (21, "Matériel, mobilier et actifs biologiques", False),
    (22, "Matériel de transport", False),
    (23, "IMMOBILISATIONS CORPORELLES", True),
    (24, "TOTAL GENERAL", True),
]
NOTE3C_COLS = [
    ("D", "Amortissements cumulés ouverture", False),
    ("F", "Dotations de l'exercice", False),
    ("H", "Amortissements relatifs aux sorties", False),
    ("J", "Reprises d'amortissements", True),
    ("L", "Virements de poste à poste", True),
    ("N", "Cumul amortissements clôture", False),
]

NOTE3CBIS_LINES = NOTE3C_LINES
NOTE3CBIS_COLS = [
    ("D", "Provisions cumulées ouverture", False),
    ("F", "Dotations de l'exercice", False),
    ("H", "Reprises (sorties/reprises)", False),
    ("J", "Cumul provisions clôture", False),
]

NOTE3D_LINES = [
    (10, "Frais de développement et de prospection", False),
    (11, "Brevets, licences, logiciels et droits similaires", False),
    (12, "Fonds commercial et droit au bail", False),
    (13, "Autres immobilisations incorporelles", False),
    (14, "IMMOBILISATIONS INCORPORELLES", True),
    (15, "Terrains", False),
    (16, "Bâtiments", False),
    (17, "Aménagements, agencements et installations", False),
    (18, "Matériel, mobilier et actifs biologiques", False),
    (19, "Matériel de transport", False),
    (20, "IMMOBILISATIONS CORPORELLES", True),
    (21, "Titres de participations", False),
    (22, "Autres immobilisations financières", False),
    (23, "IMMOBILISATIONS FINANCIERES", True),
    (24, "TOTAL GENERAL", True),
]
NOTE3D_COLS = [
    ("D", "Montant brut", False),
    ("E", "Amortissements pratiqués", False),
    ("F", "Valeur comptable nette", False),
    ("G", "Prix de cession", True),
    ("H", "Plus-value ou moins-value", False),
]

NOTE3B_LINES = [
    (11, "Terrains", False),
    (12, "Bâtiments", False),
    (13, "Matériel, mobilier", False),
    (14, "Matériel de transport", False),
    (15, "TOTAL IMMOBILISATIONS EN LOCATION-ACQUISITION", True),
]
NOTE3B_COLS = [
    ("F", "Montant brut ouverture", True),
    ("G", "Acquisitions/Apports/Créations", True),
    ("H", "Virements de poste à poste (entrée)", True),
    ("I", "Réévaluation", True),
    ("J", "Cessions/Mises hors service", True),
    ("K", "Virements de poste à poste (sortie)", True),
    ("L", "Montant brut clôture", False),
]


# Definitions des NOTE 4 a 11 (lot A — actif circulant HAO, creances, stocks,
# clients, autres creances, titres de placement, valeurs a encaisser,
# disponibilites), structure derivee automatiquement des formules du
# classeur original (memes regles que NOTE3 : (ligne, libelle, sous-total)
# et (colonne, libelle, saisie_manuelle)).
NOTE4_LINES = [
    (9, "Titres de participation", False),
    (10, "Prêts et créances", False),
    (11, "Prêt au personnel", False),
    (12, "Créances sur l'état", False),
    (13, "Titres immobilisés", False),
    (14, "Dépôts et cautionnements", False),
    (15, "Intérêts courus", False),
    (16, "Créances rattachées à des avances et participations à des GIE", False),
    (17, "Immobilisations financières diverses", False),
    (18, "TOTAL BRUT", True),
    (19, "Dépréciations des titres de participation", False),
    (20, "Dépréciations des autres immobilisations financières", False),
    (21, "TOTAL NET DE DEPRECIATION", True),
]
NOTE4_COLS = [
    ("F", "Année N", False),
    ("G", "Année N-1", False),
    ("H", "Variation en %", False),
    ("I", "Créances à un an au plus", True),
    ("J", "Créances à plus d'un an et à deux ans au plus", True),
    ("K", "Créances à plus de deux ans", False),
]

NOTE5_LINES = [
    (10, "Créances sur cessions d'immobilisations", False),
    (11, "Autres créances hors activités ordinaires", False),
    (12, "TOTAL BRUT", True),
    (13, "Dépréciations des créances HAO", False),
    (14, "TOTAL NET DE DEPRECIATIONS", True),
]
NOTE5_COLS = [
    ("E", "Année N", False),
    ("G", "Année N-1", False),
    ("I", "Variation en %", False),
]
# Le classeur source comporte un second tableau sur la meme feuille NOTE 5,
# "DETTES CIRCULANTES HAO" (lignes 22-26), avec les memes colonnes E/G/I —
# absent jusqu'ici de l'application.
NOTE5_DETTES_LINES = [
    (22, "Fournisseurs d'investissements", False),
    (23, "Fournisseurs d'investissements effets à payer", False),
    (24, "Versements restant à effectuer sur titres de participation et titres immobilisés non libérés", False),
    (25, "Autres dettes hors activités ordinaires", False),
    (26, "TOTAL", True),
]

NOTE6_LINES = [
    (9, "Marchandises", False),
    (10, "Matières premières et fournitures liées", False),
    (11, "Autres approvisionnements", False),
    (12, "Produits en cours", False),
    (13, "Services en cours", False),
    (14, "Produits finis", False),
    (15, "Produits intermédiaires", False),
    (16, "Stocks en cours de route, en consignation ou en dépôt", False),
    (17, "TOTAL BRUT STOCKS ET EN COURS", True),
    (18, "Dépréciations des stocks", False),
    (19, "TOTAL NET DE DEPRECIATIONS", True),
]
NOTE6_COLS = [
    ("E", "Année N", False),
    ("G", "Année N-1", False),
    ("I", "Variation de stock en valeur", False),
    ("J", "Variation en %", False),
]

NOTE7_LINES = [
    (9, "Clients (hors réserves de propriété et Groupe)", False),
    (10, "Clients effets à recevoir (hors réserves de propriété et groupe)", False),
    (11, "Clients avec réserves de propriété", False),
    (12, "Clients et effets à recevoir Groupe", False),
    (13, "Créances sur cession d'immobilisations", False),
    (14, "Clients effets escomptés et non échus", False),
    (15, "Créances litigieuses ou douteuses", False),
    (16, "Clients produits à recevoir", False),
    (17, "TOTAL BRUT CLIENTS", True),
    (18, "Dépréciations des comptes clients", False),
    (19, "TOTAL NET DE DEPRECIATIONS", True),
    (20, "Clients, avances reçues hors groupe", False),
    (21, "Clients, avances reçues groupe", False),
    (22, "Autres clients créditeurs", False),
    (23, "TOTAL CLIENTS CREDITEURS", True),
]
NOTE7_COLS = [
    ("E", "Année N", False),
    ("F", "Année N-1", False),
    ("G", "Variation en %", False),
    ("H", "Créances à un an au plus", False),
    ("I", "Créances à plus d'un an et à deux ans au plus", True),
    ("J", "Créances à plus de deux ans", True),
]

NOTE8_LINES = [
    (9, "Personnel", False),
    (10, "Organismes sociaux", False),
    (11, "Etat et Collectivités publiques", False),
    (12, "Organismes internationaux", False),
    (13, "Apporteurs, associés et groupe", False),
    (14, "Compte transitoire ajustement spécial lié à la révision du SYSCOHADA (Voir Notes 8A & 8C)", False),
    (15, "Autres débiteurs divers", False),
    (16, "Comptes permanents non bloqués des établissements et des succursales", False),
    (17, "Comptes de liaison charges et produits", False),
    (18, "Comptes de liaison des sociétés en participation", False),
    (19, "TOTAL BRUT AUTRES CREANCES", True),
    (20, "Dépréciations des autres créances", False),
    (21, "TOTAL NET DE DEPRECIATIONS", True),
]
NOTE8_COLS = NOTE7_COLS

NOTE9_LINES = [
    (9, "Titres de trésor et bons de caisse à court terme", False),
    (10, "Actions", False),
    (11, "Obligations", False),
    (12, "Bons de souscription", False),
    (13, "Titres négociables hors régions", False),
    (14, "Intérêts courus", False),
    (15, "Autres valeurs assimilées", False),
    (16, "TOTAL BRUT TITRES", True),
    (17, "Dépréciations des titres", False),
    (18, "TOTAL NET DE DEPRECIATIONS", True),
]
NOTE9_COLS = NOTE5_COLS

NOTE10_LINES = [
    (9, "Effets à encaisser", False),
    (10, "Effets à l'encaissement", False),
    (11, "Chèques à encaisser", False),
    (12, "Chèques à l'encaissement", False),
    (13, "Cartes de crédit à encaisser", False),
    (14, "Autres valeurs à encaisser", False),
    (15, "TOTAL BRUT VALEURS A ENCAISSER", True),
    (16, "Dépréciations des valeurs à encaisser", False),
    (17, "TOTAL NET DE DEPRECIATIONS", True),
]
NOTE10_COLS = NOTE5_COLS

NOTE11_LINES = [
    (9, "Banques locales", False),
    (10, "Banques autres états région", False),
    (11, "Banques, dépôt à terme", False),
    (12, "Autres Banques", False),
    (13, "Banques intérêts courus", False),
    (14, "Chèques postaux", False),
    (15, "Autres établissement financiers", False),
    (16, "Etablissement financiers intérêts courus", False),
    (17, "Instruments de trésorerie", False),
    (18, "Instruments de monnaie électronique", False),
    (19, "Caisse", False),
    (20, "Régies d'avances et virements accréditifs", False),
    (21, "TOTAL BRUT DISPONIBILITES", True),
    (22, "Dépréciations", False),
    (23, "TOTAL NET DE DEPRECIATIONS", True),
]
NOTE11_COLS = NOTE5_COLS

NOTES_LOT_A = [
    ("NOTE 4", NOTE4_LINES, NOTE4_COLS, "NOTE 4 — IMMOBILISATIONS FINANCIERES (créances)"),
    ("NOTE 5", NOTE5_LINES, NOTE5_COLS, "NOTE 5 — ACTIF CIRCULANT HAO"),
    ("NOTE 5", NOTE5_DETTES_LINES, NOTE5_COLS, "NOTE 5 — DETTES CIRCULANTES HAO"),
    ("NOTE 6", NOTE6_LINES, NOTE6_COLS, "NOTE 6 — STOCKS ET EN-COURS"),
    ("NOTE 7", NOTE7_LINES, NOTE7_COLS, "NOTE 7 — CLIENTS"),
    ("NOTE 8", NOTE8_LINES, NOTE8_COLS, "NOTE 8 — AUTRES CREANCES"),
    ("NOTE 9", NOTE9_LINES, NOTE9_COLS, "NOTE 9 — TITRES DE PLACEMENT"),
    ("NOTE 10", NOTE10_LINES, NOTE10_COLS, "NOTE 10 — VALEURS A ENCAISSER"),
    ("NOTE 11", NOTE11_LINES, NOTE11_COLS, "NOTE 11 — DISPONIBILITES"),
]

# ---------------------------------------------------------------------------
# Lot B : NOTE 14, 15A, 16A, 17-20 (capitaux propres, dettes financieres,
# dettes circulantes). Comme NOTE 4-11, utilisent BalanceN ET BalanceN1.
# ---------------------------------------------------------------------------

NOTE14_LINES = [
    (9, "Primes d'émission", False),
    (10, "Prime d'apport", False),
    (11, "Prime de fusion", False),
    (12, "Prime de conversion", False),
    (13, "Autres primes", False),
    (14, "TOTAL PRIMES", True),
    (15, "Réserves légales", False),
    (16, "Réserves statutaires", False),
    (17, "Réserves de plus-values nettes à long terme", False),
    (18, "Réserves d'attribution gratuite d'actions au personnel salarié et aux dirigeants", False),
    (19, "Autres réserves réglementées", False),
    (20, "TOTAL RESERVES INDISPONIBLES", True),
    (21, "Réserves libres", False),
    (22, "Report à nouveau", False),
]
NOTE14_COLS = [
    ("F", "Année N", False),
    ("G", "Année N-1", False),
    ("H", "Variation en valeur", False),
    ("I", "Variation en %", False),
]

NOTE15A_LINES = [
    (9, "État", False),
    (10, "Régions", False),
    (11, "Départements", False),
    (12, "Communes et collectivités publiques décentralisées", False),
    (13, "Entités publiques ou mixtes", False),
    (14, "Entités et organismes privés", False),
    (15, "Organismes internationaux", False),
    (16, "Autres", False),
    (17, "TOTAL SUBVENTIONS", True),
    (18, "Amortissements dérogatoires", False),
    (19, "Plus-value de cession à réinvestir", False),
    (20, "Provisions spéciales de réévaluation", False),
    (21, "Provisions réglementées relatives aux immobilisations", False),
    (22, "Provisions réglementées relatives aux stocks", False),
    (23, "Provisions pour investissement", False),
    (24, "Autres provisions et fonds réglementés", False),
    (25, "TOTAL PROVISIONS REGLEMENTEES", True),
    (26, "TOTAL SUBVENTIONS ET PROVISIONS REGLEMENTEES", True),
]
NOTE15A_COLS = [
    ("E", "Année N", False),
    ("F", "Année N-1", False),
    ("G", "Variation en valeur", False),
    ("H", "Variation en %", False),
    ("I", "Régime fiscal", True),
    ("J", "Echéances", True),
]

NOTE16A_LINES = [
    (9, "Emprunts obligataires", False),
    (10, "Emprunts et dettes auprès des établissements de crédit", False),
    (11, "Avances reçues de l'Etat", False),
    (12, "Avances reçues et comptes courants bloqués", False),
    (13, "Dépôts et cautionnements reçus", False),
    (14, "Intérêts courus", False),
    (15, "Avances et dettes assorties de conditions particulières", False),
    (16, "Autres emprunts et dettes", False),
    (17, "Dettes liées à des participations et sociétés en participation", False),
    (18, "Comptes permanents bloqués des établissements et succursales", False),
    (19, "TOTAL EMPRUNTS ET DETTES FINANCIERES", True),
    (20, "Crédit bail immobilier", False),
    (21, "Crédit bail mobilier", False),
    (22, "Location vente", False),
    (23, "Intérêts courus", False),
    (24, "Autres dettes de location acquisition", False),
    (25, "TOTAL DETTES DE LOCATION ACQUISITION", True),
    (26, "Provisions pour litiges", False),
    (27, "Provisions pour garantie donnés aux clients", False),
    (28, "Provisions pour pertes sur marchés à achèvement futur", False),
    (29, "Provisions pour pertes de change", False),
    (30, "Provisions pour impôts", False),
    (31, "Provisions pour pensions et obligations assimilées - engagements de retraite", False),
    (32, "Actif du régime de retraite (1)", False),
    (33, "Provisions pour restructuration", False),
    (34, "Provisions pour amendes et pénalités", False),
    (35, "Provisions pour propre assureur", False),
    (36, "Provisions pour démantèlement et remise en état", False),
    (37, "Provisions de droits à réduction ou avantage en nature", False),
    (38, "Autres provisions", False),
    (39, "TOTAL PROVISIONS POUR RISQUES ET CHARGES", True),
]
NOTE16A_COLS = [
    ("E", "Année N", False),
    ("F", "Année N-1", False),
    ("G", "Variation en valeur", False),
    ("H", "Variation en %", False),
    ("I", "Dettes à un an au plus", True),
    ("J", "Dettes à plus d'un an et à deux ans au plus", True),
    ("K", "Dettes à plus de deux ans", False),
]

NOTE17_LINES = [
    (9, "Fournisseurs dettes en compte (hors groupe)", False),
    (10, "Fournisseurs, sous-traitants", False),
    (11, "Fournisseurs, réserve de propriété", False),
    (12, "Fournisseurs, retenue de garantie", False),
    (13, "Fournisseurs effets à payer (hors groupe)", False),
    (14, "Fournisseurs, dettes et effets à payer groupe", False),
    (15, "Fournisseurs, acquisitions courantes d'immobilisations", False),
    (16, "Fournisseurs factures non parvenues (hors groupe)", False),
    (17, "Fournisseurs factures non parvenues groupe", False),
    (18, "TOTAL FOURNISSEURS", True),
    (19, "Fournisseurs, avances et acomptes (hors groupe)", False),
    (20, "Fournisseurs, avances et acomptes groupe", False),
    (21, "Autres fournisseurs débiteurs", False),
    (22, "TOTAL FOURNISSEURS DEBITEURS", True),
]
NOTE17_COLS = [
    ("F", "Année N", False),
    ("G", "Année N-1", False),
    ("H", "Variation en %", False),
    ("I", "Dettes à un an au plus", False),
    ("J", "Dettes à plus d'un an et à deux ans au plus", True),
    ("K", "Dettes à plus de deux ans", True),
]

NOTE18_LINES = [
    (9, "Personnel rémunérations dues", False),
    (10, "Personnel, congés à payer", False),
    (11, "Charges sociales sur congés à payer", False),
    (12, "Autres personnel", False),
    (13, "Caisse de sécurité sociale", False),
    (14, "Caisse de retraite", False),
    (15, "Mutuelle de santé", False),
    (16, "Assurance Retraite", False),
    (17, "Autres charges sociales à payer", False),
    (18, "Autres cotisations et organismes sociaux", False),
    (19, "TOTAL DETTES SOCIALES", True),
    (20, "Etat, impôts sur les bénéfices", False),
    (21, "Etat, impôts et taxes", False),
    (22, "Etat, TVA", False),
    (23, "Etat, impôts retenus à la source", False),
    (24, "Autres dettes Etat", False),
    (25, "TOTAL DETTES FISCALES", True),
    (26, "TOTAL DETTES SOCIALES ET FISCALES", True),
]
NOTE18_COLS = [
    ("E", "Année N", False),
    ("F", "Année N-1", False),
    ("G", "Variation en valeur", False),
    ("H", "Variation en %", False),
    ("I", "Dettes à un an au plus", False),
    ("J", "Dettes à plus d'un an et à deux ans au plus", True),
    ("K", "Dettes à plus de deux ans", True),
]

NOTE19_LINES = [
    (9, "Organismes internationaux", False),
    (10, "Apporteurs, opérations sur le capital", False),
    (11, "Associés, compte courant", False),
    (12, "Associés dividendes à payer", False),
    (13, "Groupe, comptes courants", False),
    (14, "Autres dettes associés", False),
    (15, "TOTAL DETTES ASSOCIES", True),
    (16, "Créditeurs divers", False),
    (17, "Obligataires", False),
    (18, "Rémunérations d'administrateurs", False),
    (19, "Compte d'affacturage et de titrisation", False),
    (20, "Versements restant à effectuer sur titres de placement non libérés", False),
    (21, "Compte transitoire ajustement spécial lié à la révision du SYSCOHADA (Voir Note 8B)", False),
    (22, "Autres créditeurs divers", False),
    (23, "TOTAL  CREDITEURS DIVERS", True),
    (24, "Comptes permanents non bloqués des établissements et des succursales", False),
    (25, "Comptes de liaison charges et produits", False),
    (26, "Comptes de liaison des sociétés en participation", False),
    (27, "TOTAL  COMPTES DE LIAISON", True),
    (28, "TOTAL  AUTRES DETTES", True),
    (29, "Provisions pour risques et charges à court terme (voir note 28)", False),
]
NOTE19_COLS = NOTE18_COLS

NOTE20_LINES = [
    (9, "Escomptes de crédit de campagne", False),
    (10, "Escomptes de crédit ordinaires", False),
    (11, "TOTAL: BANQUES, CREDITS D'ESCOMPTE ET DE TRESORERIE", True),
    (12, "Banques locales", False),
    (13, "Banques autres états région", False),
    (14, "Autres Banques", False),
    (15, "Banques intérêts courus", False),
    (16, "Crédit de trésorerie", False),
    (17, "TOTAL: BANQUES, CREDITS DE TRESORERIE", True),
    (18, "TOTAL GENERAL", True),
]
NOTE20_COLS = [
    ("F", "Année N", False),
    ("G", "Année N-1", False),
    ("H", "Variation en %", False),
]

NOTES_LOT_B = [
    ("NOTE 14", NOTE14_LINES, NOTE14_COLS, "NOTE 14 — PRIMES ET RESERVES"),
    ("NOTE 15A", NOTE15A_LINES, NOTE15A_COLS, "NOTE 15A — SUBVENTIONS D'INVESTISSEMENT ET PROVISIONS REGLEMENTEES"),
    ("NOTE 16A", NOTE16A_LINES, NOTE16A_COLS, "NOTE 16A — DETTES FINANCIERES ET RESSOURCES ASSIMILES"),
    ("NOTE 17", NOTE17_LINES, NOTE17_COLS, "NOTE 17 — FOURNISSEURS D'EXPLOITATION"),
    ("NOTE 18", NOTE18_LINES, NOTE18_COLS, "NOTE 18 — DETTES FISCALES ET SOCIALES"),
    ("NOTE 19", NOTE19_LINES, NOTE19_COLS, "NOTE 19 — AUTRES DETTES ET PROVISIONS POUR RISQUES ET CHARGES A COURT TERME"),
    ("NOTE 20", NOTE20_LINES, NOTE20_COLS, "NOTE 20 — BANQUES, CREDIT D'ESCOMPTE ET DE TRESORERIE"),
]

# --------------------------------------------------------------- Lot C ----
# NOTE 21 a 29 (sauf 27B, a saisie manuelle, reportee au Lot D) : produits,
# achats, charges par nature, charges de personnel, provisions, charges et
# revenus financiers. Comme Lot A/B, utilisent BalanceN ET BalanceN1.
SIMPLE_VAR_COLS = [
    ("F", "Année N", False),
    ("G", "Année N-1", False),
    ("H", "Variation en %", False),
]

# NOTE 21 ligne 27 (Produits accessoires) est retiree de cette liste : dans le
# classeur Excel, F27/G27 = SUM(...) sur une ligne de detail masquee (28) via
# INDIRECT/ADDRESS, le meme principe self-referentiel que NOTE 3E/30/32/33.
# Elle est donc traitee a part comme un bloc a lignes dynamiques (voir
# compute_note21_produits_accessoires) et reinjectee dans le tableau au
# rendu (entre la ligne 26 et la ligne 29).
NOTE21_LINES_A = [
    (9, "Ventes de marchandises dans l'Etat partie", False),
    (10, "Ventes de marchandises dans les autres Etats parties de la Région (2)", False, ("F", "G")),
    (11, "Ventes de marchandises hors Région (2)", False),
    (12, "Ventes de marchandises groupe", False),
    (13, "Ventes de marchandises sur internet", False),
    (14, "TOTAL : VENTES MARCHANDISES", True),
    (15, "Ventes de produits fabriqués dans l'Etat partie", False),
    (16, "Ventes de produits fabriqués dans les autres Etats parties de la Région (2)", False, ("F", "G")),
    (17, "Ventes de produits fabriqués hors Région (2)", False),
    (18, "Ventes de produits fabriqués groupe", False),
    (19, "Ventes de produits fabriqués sur internet", False),
    (20, "TOTAL : VENTES DE PRODUITS FABRIQUES", True),
    (21, "Ventes de travaux et services dans l'Etat partie", False),
    (22, "Ventes de travaux et services dans les autres Etats parties de la Région (2)", False, ("F", "G")),
    (23, "Ventes de travaux et services hors Région (2)", False),
    (24, "Ventes de travaux et services groupe", False),
    (25, "Ventes de travaux et services sur internet", False),
    (26, "TOTAL : VENTES DE TRAVAUX ET SERVICES VENDUS", True),
]
NOTE21_LINES_B = [
    (29, "TOTAL : CHIFFRES D'AFFAIRES", True),
    (30, "Production immobilisée", False, ("F", "G")),
    (31, "Subventions d'exploitation", False, ("F", "G")),
    (32, "Autres produits (1)", False, ("F", "G")),
    (33, "TOTAL : AUTRES PRODUITS", True),
    (34, "TOTAL", True),
]
# Conservee pour compatibilite (libelles complets, non utilisee pour le rendu).
NOTE21_LINES = NOTE21_LINES_A + [
    (27, "Produits accessoires à détailler par nature d'activité économique", False),
] + NOTE21_LINES_B
NOTE21_COLS = SIMPLE_VAR_COLS

NOTE22_LINES = [
    (9, "Achats de marchandises dans l'Etat partie", False),
    (10, "Achats de marchandises dans les autres Etats parties de la Région (2)", False),
    (11, "Achats de marchandises hors Région (2)", False),
    (12, "Achats de marchandises groupe", False),
    (13, "TOTAL : ACHATS DE MARCHANDISES", True),
    (14, "Achats de matières premières et fournitures liées dans l'Etat partie", False),
    (15, "Achats de matières premières et fournitures liées dans les autres Etats parties de la Région (2)", False),
    (16, "Achats matières premières et fournitures liées hors Région (2)", False),
    (17, "Achats matières premières et fournitures liées groupe", False),
    (18, "TOTAL : ACHATS MATIERES PREMIERES ET FOURNITURES LIEES", True),
    (19, "Matières consommables", False),
    (20, "Matières combustibles", False),
    (21, "Produits d'entretien", False),
    (22, "Fournitures d'atelier, d'usine et de magasin", False),
    (24, "Electricité", False),
    (25, "Autres énergies", False),
    (26, "Fourniture d'entretien", False),
    (27, "Fourniture de bureau", False),
    (28, "Petit matériel et outillages", False),
    (29, "Achats études, prestations de services, de travaux matériels et équipements", False),
    (30, "Achats d'emballages", False),
    (31, "Frais sur achats (1)", False),
    (32, "Remises rabais, remises et ristournes (non ventilés)", False),
    (33, "TOTAL : AUTRES ACHATS", True),
]
NOTE22_COLS = SIMPLE_VAR_COLS

NOTE23_LINES = [
    (9, "Transports sur ventes", False),
    (10, "Transports pour le compte de tiers", False),
    (11, "Transport du personnel", False),
    (12, "Transports de plis", False),
    (13, "Voyage déplacement (transport)", False),
    (14, "Transport entre établissements ou chantiers", False),
    (15, "Transports administratifs", False),
    (16, "TOTAL", True),
]
NOTE23_COLS = SIMPLE_VAR_COLS

NOTE24_LINES = [
    (9, "Sous-traitance générale", False),
    (10, "Locations et charges locatives", False),
    (11, "Redevances de location acquisition", False),
    (12, "Entretien, réparations et maintenance", False),
    (13, "Primes d'assurance", False),
    (14, "Etudes, recherches et documentation", False),
    (15, "Publicité, publications, relations publiques", False),
    (16, "Frais de télécommunications", False),
    (17, "Frais bancaires", False),
    (18, "Rémunérations d'intermédiaires et de conseils", False),
    (19, "Frais de formation du personnel", False),
    (20, "Redevances pour brevets, licences, logiciels, concession et droits similaires", False),
    (21, "Cotisations", False),
    (22, "Rémunérations de personnel extérieur à l'entité", False),
    (23, "Autres charges externes", False),
    (24, "TOTAL", True),
]
NOTE24_COLS = SIMPLE_VAR_COLS

NOTE25_LINES = [
    (9, "Impôts et taxes directs", False),
    (10, "Impôts et taxes indirects", False),
    (11, "Droits d'enregistrement", False),
    (12, "Pénalités et amendes fiscales", False),
    (13, "Autres impôts et taxes", False),
    (14, "TOTAL", True),
]
NOTE25_COLS = SIMPLE_VAR_COLS

NOTE26_LINES = [
    (9, "Pertes sur créances clients", False),
    (10, "Pertes sur autres débiteurs", False),
    (11, "Quote-part de résultat sur opérations faites en commun", False),
    (12, "Valeur comptable des cessions courantes d'immobilisations", False),
    (13, "Perte de change sur créances et dettes commerciales", False),
    (14, "Pénalités et amendes pénales", False),
    (15, "Indemnités de fonction et autres rémunérations d'administrateurs", False),
    (16, "Dons et mécénat", False),
    (17, "Autres charges diverses", False),
    (18, "Charges pour dépréciations et provisions pour risques à court terme d'exploitation (voir note 28)", False),
    (19, "TOTAL", True),
]
NOTE26_COLS = SIMPLE_VAR_COLS

NOTE27A_LINES = [
    (9, "Rémunérations directes versées au personnel national", False),
    (10, "Rémunérations directes versées au personnel non national", False),
    (11, "Indemnités forfaitaires versées au personnel", False),
    (12, "Charges sociales (personnel national)", False),
    (13, "Charges sociales (personnel non national)", False),
    (14, "Rémunérations et charges sociales de l'exploitant individuel", False),
    (15, "Rémunération transférée de personnel extérieur", False),
    (16, "Autres charges sociales", False),
    (17, "TOTAL", True),
]
NOTE27A_COLS = SIMPLE_VAR_COLS

NOTE28_LINES = [
    (11, "Provisions réglementées", False),
    (12, "Provisions financières pour risques et charges", False, ("G", "H", "J", "K")),
    (13, "Dépréciations des immobilisations", False, ("G", "H", "J", "K")),
    (14, "TOTAL DOTATIONS", True),
    (15, "Dépréciations des stocks et en cours", False, ("H", "K")),
    (16, "Dépréciations des comptes fournisseurs", False, ("H", "K")),
    (17, "Dépréciations des comptes clients", False, ("H", "K")),
    (18, "Dépréciations autres créances d'exploitation", False, ("H", "K")),
    (19, "Dépréciations des comptes de créances HAO", False, ("F", "K")),
    (20, "Dépréciations des titres de placement", False, ("H", "K")),
    (21, "Dépréciations des valeurs à encaisser", False, ("H", "K")),
    (22, "Dépréciations des comptes banques", False, ("H", "K")),
    (23, "Dépréciations des comptes établissements financiers et assimilés", False, ("H", "K")),
    (24, "Dépréciations des comptes d'instruments de trésorerie", False, ("H", "K")),
    (25, "Provisions pour risques à court terme d'exploitation", False),
    (26, "Provisions pour risques à court terme HAO", False),
    (27, "Provisions pour risques à court terme à caractère financier", False),
    (28, "TOTAL CHARGES POUR DEPRECIATIONS ET PROVISIONS A COURT TERME", True),
    (29, "TOTAL", True),
]
NOTE28_COLS = [
    ("E", "Ouverture", False),
    ("F", "Dotations - exploitation", False),
    ("G", "Dotations - financières", False),
    ("H", "Dotations - HAO", False),
    ("I", "Reprises - exploitation", False),
    ("J", "Reprises - financières", False),
    ("K", "Reprises - HAO", False),
    ("L", "Clôture", False),
]

NOTE29_LINES = [
    (9, "Intérêts des emprunts", False),
    (10, "Intérêts dans loyers de locations acquisition", False),
    (11, "Escomptes accordés", False),
    (12, "Autres intérêts", False),
    (13, "Escomptes des effets de commerce", False),
    (14, "Pertes de change financières", False),
    (15, "Pertes sur cessions de titres de placement", False),
    (16, "Malis provenant d'attribution gratuite d'actions au personnel salarié et aux dirigeants", False),
    (17, "Pertes et charges sur risques financiers", False),
    (18, "Charges pour dépréciation et provisions à court terme à caractère financier (voir note 28)", False),
    (19, "SOUS TOTAL : FRAIS FINANCIERS (A)", True),
    (20, "Intérêts de prêts et créances diverses", False),
    (21, "Revenus de participations et autres titres immobilisés", False),
    (22, "Escomptes obtenus", False),
    (23, "Revenus de placement", False),
    (24, "Intérêts dans loyers de location-financement", False),
    (25, "Gains de change financiers", False),
    (26, "Gains sur cessions de titres de placement", False),
    (27, "Gains sur risques financiers", False),
    (28, "Reprises de charges pour dépréciation et provisions à court terme à caractère financier (voir note 28)", False),
    (29, "SOUS TOTAL : REVENUS FINANCIERS (B)", True),
    (30, "SOUS TOTAL (contrôle) : RESULTAT FINANCIER (B) - (A)", True),
]
NOTE29_COLS = SIMPLE_VAR_COLS

# NOTE 21 est retiree de cette liste generique : la ligne 27 (Produits
# accessoires) a besoin d'un bloc a lignes dynamiques + saisie des cellules
# en jaune (voir compute_note21, carte dediee "note-21" dans le template).
NOTES_LOT_C = [
    ("NOTE 22", NOTE22_LINES, NOTE22_COLS, "NOTE 22 — ACHATS"),
    ("NOTE 23", NOTE23_LINES, NOTE23_COLS, "NOTE 23 — TRANSPORTS"),
    ("NOTE 24", NOTE24_LINES, NOTE24_COLS, "NOTE 24 — SERVICES EXTERIEURS"),
    ("NOTE 25", NOTE25_LINES, NOTE25_COLS, "NOTE 25 — IMPOTS ET TAXES"),
    ("NOTE 26", NOTE26_LINES, NOTE26_COLS, "NOTE 26 — AUTRES CHARGES"),
    ("NOTE 27A", NOTE27A_LINES, NOTE27A_COLS, "NOTE 27A — CHARGES DE PERSONNEL"),
    ("NOTE 28", NOTE28_LINES, NOTE28_COLS, "NOTE 28 — DOTATIONS ET CHARGES POUR PROVISIONS ET DEPRECIATIONS"),
    ("NOTE 29", NOTE29_LINES, NOTE29_COLS, "NOTE 29 — CHARGES ET REVENUS FINANCIERS"),
]

# --------------------------------------------------------------- Lot D ----
# Feuilles annexes restantes (12, 13, 15B, 16B/16B BIS/16C, 27B,
# 30, 32, 33, 37) : contrairement a leur etiquette "saisie manuelle" dans le
# plan de travail, la plupart contiennent de vraies formules Excel (SUM,
# IFERROR, INDIRECT/ADDRESS dynamique comme NOTE 3E/NOTE 21) ; seules les
# cellules de detail (comptes, montants, hypotheses) sont a saisie manuelle.
# NOTE 35, 38, 39 sont purement textuelles (aucune formule ni donnee
# chiffree) et sont traitees a part (voir NOTES_TEXTE_DEFS plus bas).
# NOTE 8A/8B/8C ont ete retirees de l'application a la demande du client.

NOTE12_ECARTS_COLS = [
    ("E", "Devises", True),
    ("F", "Montant en devises", True),
    ("G", "Cours UML année acquisition", True),
    ("H", "Cours UML au 31/12", True),
    ("I", "Variation en valeur", False),
]
NOTE12_ECARTS_LINES = [
    (11, "Ecart de conversion actif (créances/dettes concernées)", False),
    (13, "Ecart de conversion passif (créances/dettes concernées)", False),
]
NOTE12_TRANSFERTS_COLS = [
    ("G", "Année N", True),
    ("H", "Année N-1", True),
    ("I", "Variation en %", False),
]
NOTE12_TRANSFERTS_LINES = [
    (18, "Transferts de charges d'exploitation", False),
    (20, "Transferts de charges financières", False),
]

NOTE13_COLS = [
    ("E", "N° compte contribuable", True),
    ("F", "Nationalité", True),
    ("G", "Autres nationalités", True),
    ("H", "Pays de résidence", True),
    ("I", "Nature des actions/parts", True),
    ("J", "Nombre", True),
    ("K", "Montant total", False),
    ("L", "Cessions/remb. en cours d'exercice", True),
]
NOTE13_LINES = [
    (8, "Valeur nominale des actions ou parts (colonne L)", False),
    (10, "Actionnaire / associé (détail)", False),
    (11, "Apporteurs, capital non appelé", False),
    (12, "TOTAL", True),
]

NOTE15B_COLS = [
    ("G", "Année N", True),
    ("H", "Année N-1", True),
    ("I", "Variation en valeur", False),
    ("J", "Variation en %", False),
    ("K", "Echéances", True),
]
NOTE15B_LINES = [
    (9, "Titres participatifs", False),
    (10, "Avances conditionnées", False),
    (11, "Titres subordonnés à durée indéterminée (T.S.D.I.)", False),
    (12, "Obligations remboursables en actions (O.R.A.)", False),
    (13, "Autres avances et dettes assorties de conditions particulières", False),
    (14, "TOTAL AUTRES FONDS PROPRES", True),
]

NOTE16B_HYP_COLS = [("H", "Année N", True), ("I", "Année N-1", True)]
NOTE16B_HYP_LINES = [
    (10, "Taux d'augmentation des salaires", False),
    (11, "Taux d'actualisation", False),
    (12, "Taux d'inflation", False),
    (13, "Probabilité d'être présent à la date de départ à la retraite", False),
    (14, "Probabilité d'être en vie à l'âge de départ à la retraite", False),
    (15, "Taux de rendement effectif des actifs du régime", False),
]
NOTE16B_VAR_COLS = [("H", "Année N", True), ("I", "Année N-1", True)]
NOTE16B_VAR_LINES = [
    (21, "Obligation au titre des engagements de retraite à l'ouverture", False),
    (22, "Coût des services rendus au cours de l'exercice", False),
    (23, "Coût financier", False),
    (24, "Pertes actuarielles / (gain)", False),
    (25, "Prestations payées au cours de l'exercice", False),
    (26, "Coût des services passés", False),
    (27, "Obligation au titre des engagements de retraite à la clôture", False),
]
NOTE16B_SENS_COLS = [
    ("F", "Année N — Augmentation", True),
    ("G", "Année N — Diminution", True),
    ("H", "Année N-1 — Augmentation", True),
    ("I", "Année N-1 — Diminution", True),
]
NOTE16B_SENS_LINES = [
    (34, "Taux d'actualisation (variation de ...%)", False),
    (35, "Taux de progression des salaires (variation de ...%)", False),
    (36, "Taux de départ du personnel (variation de ...%)", False),
]

NOTE16BBIS_ACTIFPASSIF_COLS = [("H", "Année N", True), ("I", "Année N-1", True)]
NOTE16BBIS_ACTIFPASSIF_LINES = [
    (10, "Valeur actuelle de l'obligation résultant de régimes financés", False),
    (11, "Valeur actuelle des actifs affectés aux plans de retraite", False),
    (12, "Excédent / Déficit de régime", False),
]
NOTE16BBIS_ACTIFS_COLS = [
    ("F", "Année N — Rendement attendu", False),
    ("G", "Année N — Juste valeur des actifs", True),
    ("H", "Année N-1 — Rendement attendu", False),
    ("I", "Année N-1 — Juste valeur des actifs", True),
]
NOTE16BBIS_ACTIFS_LINES = [
    (19, "Actions", False),
    (20, "Obligations", False),
    (21, "Autres", False),
    (22, "Total", True),
]

NOTE16C_COLS = [("G", "Année N", True), ("H", "Année N-1", True)]

NOTE27B_COLS = [
    # manual=False partout : l'editabilite reelle est pilotee ligne par
    # ligne via r.manual_cols (le groupe 2 "personnel exterieur" n'a pas
    # les colonnes masse salariale O/P/Q/R/S dans le classeur Excel, voir
    # compute_note27b).
    ("E", "Nationaux — M", False),
    ("F", "Nationaux — F", False),
    ("G", "Autres Etats Région — M", False),
    ("H", "Autres Etats Région — F", False),
    ("I", "Hors Région — M", False),
    ("J", "Hors Région — F", False),
    ("K", "TOTAL effectif", False),
    ("L", "Masse salariale — Nationaux M", False),
    ("M", "Masse salariale — Nationaux F", False),
    ("O", "Masse salariale — Autres Etats Région M", False),
    ("P", "Masse salariale — Autres Etats Région F", False),
    ("Q", "Masse salariale — Hors Région M", False),
    ("R", "Masse salariale — Hors Région F", False),
    ("S", "TOTAL masse salariale", False),
]
NOTE27B_LINES = [
    (12, "1. Cadres supérieurs", False),
    (13, "2. Techniciens supérieurs et cadres moyens", False),
    (14, "3. Techniciens, agents de maîtrise et ouvriers qualifiés", False),
    (15, "4. Employés, manœuvres, ouvriers et apprentis", False),
    (16, "TOTAL (1) — personnel propre (A)", True),
    (17, "Permanents", False),
    (18, "Saisonniers", False),
    (22, "Personnel extérieur — 1. Cadres supérieurs", False),
    (23, "Personnel extérieur — 2. Techniciens supérieurs et cadres moyens", False),
    (24, "Personnel extérieur — 3. Techniciens, agents de maîtrise et ouvriers qualifiés", False),
    (25, "Personnel extérieur — 4. Employés, manœuvres, ouvriers et apprentis", False),
    (26, "TOTAL (2) — personnel extérieur", True),
    (27, "Permanents", False),
    (28, "Saisonniers", False),
    (29, "TOTAL (1 + 2)", True),
]

NOTE30_COLS = [
    ("F", "Année N", True),
    ("G", "Année N-1", True),
    ("H", "Variation en %", False),
]
# NOTE30_LINES (9-29) a ete remplace par les deux listes ci-dessous : dans
# le classeur source, les lignes 9/18/19/29 sont des totaux calcules (SUM)
# et les lignes 10/20 ("Détail à préciser") ne reservent qu'une seule ligne
# mais avec une formule SUM dynamique (INDIRECT/ADDRESS) qui tolere
# l'insertion de lignes — voir compute_note30 plus bas (saisie manuelle
# avec "+ Ajouter une ligne" pour le detail, totaux 100% automatiques).
NOTE30_CHARGES_FIXED_LINES = [
    (11, "Charges liées aux opérations de restructuration"),
    (12, "Pertes sur créances HAO"),
    (13, "Dons et libéralités accordés"),
    (14, "Abandons de créances consentis"),
    (15, "Charges pour dépréciations et provisions pour risques à court terme HAO"),
    (16, "Dotations hors activités ordinaires"),
    (17, "Participation des travailleurs"),
]
NOTE30_PRODUITS_FIXED_LINES = [
    (21, "Produits liés aux opérations de restructuration"),
    (22, "Indemnités et subventions HAO"),
    (23, "Dons et libéralités obtenus"),
    (24, "Abandons de créances obtenus"),
    (25, "Transfert de charges H.A.O"),
    (26, "Reprises de charges pour dépréciations et provisions pour risques à court terme HAO"),
    (27, "Reprises des charges, provisions et dépréciations H.A.O"),
    (28, "Subventions d'équilibre"),
]

NOTE32_COLS = [
    ("E", "Vendue dans le pays — Qté", True),
    ("F", "Vendue dans le pays — Valeur", True),
    ("G", "Vendue autres pays OHADA — Qté", True),
    ("H", "Vendue autres pays OHADA — Valeur", True),
    ("I", "Vendue hors OHADA — Qté", True),
    ("J", "Vendue hors OHADA — Valeur", True),
    ("K", "Production immobilisée — Qté", True),
    ("L", "Production immobilisée — Valeur", True),
    ("M", "Stock ouverture — Qté", True),
    ("N", "Stock ouverture — Valeur", True),
]

NOTE33_COLS = [
    ("F", "Produits de l'Etat — Qté", True),
    ("G", "Produits de l'Etat — Valeur", True),
    ("H", "Importés, achetés dans l'Etat — Qté", True),
    ("I", "Importés, achetés dans l'Etat — Valeur", True),
    ("J", "Importés, achetés hors de l'Etat — Qté", True),
    ("K", "Importés, achetés hors de l'Etat — Valeur", True),
    ("L", "Variation des stocks (valeur)", True),
]

NOTE37_COLS = [
    ("H", "Montant", True),
    ("D", "Taux d'imposition (ligne 36 uniquement)", True),
]
NOTE37_LINES = [
    (9, "RESULTAT NET COMPTABLE DE L'EXERCICE", True),
    (11, "Amortissements non déductibles", False),
    (12, "Provisions non déductibles", False),
    (13, "Intérêts excédentaires des comptes courants d'associés", False),
    (14, "Rémunérations allouées à l'exploitant individuel et associés de personnes", False),
    (15, "Avantages en nature de l'exploitant individuel et associés de personnes", False),
    (16, "Impôts non déductibles (IMF, BIC, IGR... impôts non payés)", False),
    (17, "Amendes et pénalités de toute nature", False),
    (18, "Plus-values exonérées non réinvesties", False),
    (19, "Dons et subventions", False),
    (20, "Rémunérations des gérants majoritaires de SARL", False),
    (21, "Indemnités de fonction et autres rémunérations aux administrateurs", False),
    (22, "Autres charges non déductibles", False),
    (23, "TOTAL DES REINTEGRATIONS", True),
    (25, "Provisions antérieurement taxées ou définitivement exonérées réintégrées", False),
    (26, "Réfactions sur produits financiers", False),
    (27, "Réfactions sur produits d'actions", False),
    (28, "Déductions Art. 110 du CGI", False),
    (29, "Divers", False),
    (30, "TOTAL DES DEDUCTIONS", True),
    (31, "RESULTAT IMPOSABLE AVANT DEDUCTION DES DEFICITS", True),
    (32, "DEFICITS ANTERIEURS A L'EXERCICE", False),
    (33, "AMORTISSEMENTS REGULIEREMENT DIFFERES", False),
    (34, "AMORTISSEMENTS DE L'EXERCICE A DIFFERER", False),
    (35, "RESULTAT FISCAL DE L'EXERCICE", True),
    (36, "IMPOTS SUR LE RESULTAT (taux x résultat fiscal, colonne D)", True),
]

# Chaque entree : (sheet, lines, cols, titre). Le meme nom de feuille peut
# apparaitre plusieurs fois (ex. NOTE 12, NOTE 16B) quand le classeur
# original juxtapose plusieurs mini-tableaux de structure differente sur la
# meme feuille : ils partagent le meme cache de calcul (calcule une seule
# fois par feuille), seul l'affichage est scinde.
NOTES_LOT_D = [
    ("NOTE 12", NOTE12_ECARTS_LINES, NOTE12_ECARTS_COLS, "NOTE 12 — ECARTS DE CONVERSION"),
    ("NOTE 12", NOTE12_TRANSFERTS_LINES, NOTE12_TRANSFERTS_COLS, "NOTE 12 — TRANSFERTS DE CHARGES"),
    ("NOTE 15B", NOTE15B_LINES, NOTE15B_COLS, "NOTE 15B — AUTRES FONDS PROPRES"),
    ("NOTE 16B", NOTE16B_HYP_LINES, NOTE16B_HYP_COLS, "NOTE 16B — HYPOTHESES ACTUARIELLES (ENGAGEMENTS DE RETRAITE)"),
    ("NOTE 16B", NOTE16B_VAR_LINES, NOTE16B_VAR_COLS, "NOTE 16B — VARIATION DE LA VALEUR DE L'ENGAGEMENT DE RETRAITE"),
    ("NOTE 16B", NOTE16B_SENS_LINES, NOTE16B_SENS_COLS, "NOTE 16B — ANALYSE DE SENSIBILITE DES HYPOTHESES ACTUARIELLES"),
    ("NOTE 16B BIS", NOTE16BBIS_ACTIFPASSIF_LINES, NOTE16BBIS_ACTIFPASSIF_COLS, "NOTE 16B BIS — ACTIF/PASSIF NET DES REGIMES FINANCES"),
    ("NOTE 16B BIS", NOTE16BBIS_ACTIFS_LINES, NOTE16BBIS_ACTIFS_COLS, "NOTE 16B BIS — VALEUR ACTUELLE DES ACTIFS DU REGIME"),
    ("NOTE 37", NOTE37_LINES, NOTE37_COLS, "NOTE 37 — DETERMINATION DE L'IMPOT SUR LE RESULTAT"),
]
# NOTE 16C, 30, 32, 33 : retirees de la boucle generique ci-dessus et traitees
# a part (compute_note16c/30/32/33 plus bas) car elles ont besoin de lignes de
# detail dynamiques ("+ Ajouter une ligne") et d'un total calcule en Python
# (le classeur source ne reserve qu'une seule ligne de detail par section).
# NOTE 27B : retiree elle aussi de la boucle generique (compute_note27b plus
# bas) — contrairement a ce que suggerait le commentaire ci-dessus, cette
# feuille n'a PAS d'entree dans sheets_raw.json (aucune formule extraite),
# donc le mecanisme generique (cache de formules) renvoyait des totaux
# vides pour la colonne K et les lignes TOTAL (16, 26, 29) au lieu des
# vraies formules =SUM(E:J) / sommes par categorie du classeur source.
# NOTE 13 : retiree de meme (compute_note13 plus bas) — feuille absente de
# sheets_raw.json ; le classeur source ne reserve qu'une seule ligne de
# detail "Actionnaire / associe" (ligne 10, formule SUM totale dynamique
# tolerant l'insertion de lignes), avec K (Montant total) = L8 (valeur
# nominale, saisie unique) * J (Nombre), calcule et non saisi, plus une
# ligne fixe "Apporteurs, capital non appele" (ligne 11, montant K11 saisi
# directement) et un TOTAL (ligne 12) = somme des K du detail + K11.

# NOTE 35, 38, 39 : feuilles purement textuelles (aucune formule, aucune
# donnee chiffree dans le classeur d'origine) — zones de commentaire libre,
# stockees dans la table note_texte (par exercice/feuille/champ).
NOTES_TEXTE_DEFS = [
    ("NOTE 2", "NOTE 2 — INFORMATIONS OBLIGATOIRES", [
        ("conformite", "A - Déclaration de conformité au SYSCOHADA et faits marquants de l'exercice (2000 caractères max)"),
        ("regles_methodes", "B - Règles et méthodes comptables (2000 caractères max)"),
        ("derogations", "C - Dérogation aux postulats et conventions comptables (2000 caractères max)"),
        ("infos_complementaires", "D - Informations complémentaires relatives au bilan, au compte de résultat et au tableau des flux de trésorerie (2000 caractères max)"),
    ]),
    ("NOTE 35", "NOTE 35 — INFORMATIONS SOCIALES, ENVIRONNEMENTALES ET SOCIETALES", [
        ("emploi", "Emploi (effectif, embauches/licenciements, rémunérations)"),
        ("relations_sociales", "Relations sociales (dialogue social, accords collectifs)"),
        ("sante_securite", "Santé et sécurité au travail"),
        ("formation", "Formation"),
        ("egalite", "Égalité de traitement (femmes/hommes, personnes handicapées)"),
        ("env_general", "Politique générale en matière environnementale"),
        ("pollution_dechets", "Pollution et gestion des déchets"),
        ("ressources", "Utilisation durable des ressources"),
        ("climat", "Changement climatique (gaz à effet de serre)"),
        ("biodiversite", "Protection de la biodiversité"),
        ("impact_territorial", "Impact territorial, économique et social de l'activité"),
        ("relations_parties", "Relations avec les personnes/organisations intéressées par l'activité"),
    ]),
    ("NOTE 38", "NOTE 38 — EVENEMENTS POSTERIEURS A LA CLOTURE DE L'EXERCICE", [
        ("date_arrete", "Date d'arrêté des états financiers / organe ayant autorisé la publication"),
        ("avec_ajustement", "A — Evénements postérieurs donnant lieu à des ajustements"),
        ("sans_ajustement", "B — Evénements postérieurs ne donnant pas lieu à des ajustements"),
        ("continuite", "C — Evénements remettant en cause la continuité d'exploitation"),
    ]),
    ("NOTE 39", "NOTE 39 — CHANGEMENTS DE METHODES, D'ESTIMATIONS ET CORRECTIONS D'ERREURS", [
        ("methodes", "A — Changements de méthodes comptables"),
        ("estimations", "B — Changements d'estimations"),
        ("corrections", "C — Corrections d'erreurs"),
    ]),
]

# ----------------------------------------------------------------------------
# SOMMAIRE : feuille maitresse d'identification de l'entite (lignes 27-49 du
# classeur). C'est la SEULE feuille ou ces champs sont saisis ; GARDE et les
# FICHE R1/R2 ne font que les reprendre par formule (=Sommaire!D27, etc.) en
# entete ou dans une cellule. Pas de colonne dediee dans la table clients pour
# ces champs complementaires -> mecanisme generique note_texte
# (sheet="SOMMAIRE", champ=cle ci-dessous), comme pour les notes 35/38/39.
SOMMAIRE_CHAMPS = [
    ("sigle", "Sigle usuel"),
    ("caisse_sociale", "N° de caisse sociale"),
    ("repertoire_entreprises", "N° répertoire des entreprises"),
    ("agrement_prioritaire", "Agrément prioritaire"),
    ("code_importateur", "N° Code Importateur"),
    ("forme_juridique", "Forme juridique"),
    ("telephone", "N° de téléphone"),
    ("email", "Email"),
    ("forme_juridique_autre", "Si autre forme juridique, préciser"),
    ("exercice_debut", "Exercice du"),
    ("exercice_fin", "Exercice clos le (au)"),
    ("activite_principale", "Désignation de l'activité"),
    ("exercice_precedent_debut", "Exercice précédent du"),
    ("boite_postale_code", "Boîte Postale — code"),
    ("boite_postale_ville", "Boîte Postale — ville"),
    ("date_arret_comptes", "Date d'arrêt des comptes"),
    ("commune", "Commune"),
    ("quartier", "Quartier"),
    ("annee_premier_exercice", "Année du premier exercice"),
    ("greffe", "Greffe"),
    ("rccm", "N° Registre du Commerce (RCCM)"),
    ("taux_bic_bnc", "Taux d'imposition BIC / BNC"),
    ("regime_fiscal", "Régime fiscal"),
    ("centre_depot", "Centre des Impôts (centre de dépôt)"),
    ("taux_imf", "Taux IMF"),
    ("pays_siege", "Pays Siège social"),
    ("code_activite", "Code activité principale"),
]

# ----------------------------------------------------------------------------
# GARDE : page de garde proprement dite. Tous ses champs d'identite sont des
# formules pointant vers Sommaire (donc en lecture seule ici, voir SOMMAIRE_
# CHAMPS ci-dessus) ; les seuls champs propres a GARDE sont la liste des
# "Documents déposés" (avec marqueur de dépôt) et le bloc "Réservé à la
# Direction Générale des Impôts" (date de dépôt, agent, signature, nombre de
# pages/exemplaires) -> note_texte (sheet="GARDE").
GARDE_DOCUMENTS = [
    ("fiche_identification", "Fiche d'identification et renseignements divers"),
    ("bilan", "Bilan"),
    ("compte_resultat", "Compte de résultat"),
    ("tft", "Tableau des flux de trésorerie"),
    ("notes_annexes", "Notes annexes"),
    ("etat_dgi_ins", "Etats supplémentaires DGI"),
]

# Champs propres a FICHE R1 (personne a contacter, expert-comptable,
# commissaire aux comptes, signataire, domiciliations bancaires) : ne
# proviennent pas de Sommaire, donc directement en note_texte (sheet=
# "FICHE R1"), sans repli sur GARDE.
FICHE_R1_CHAMPS = [
    ("contact_nom", "Personne à contacter (nom, adresse, téléphone, e-mail, qualité)"),
    ("expert_comptable", "Expert-comptable ayant établi les états financiers"),
    ("commissaire_comptes", "Commissaire aux comptes (attestation du visa / mission)"),
    ("signataire_nom", "Nom du signataire des états financiers"),
    ("signataire_qualite", "Qualité du signataire"),
    ("signataire_date", "Date de signature"),
]

FICHE_R2_ACTIVITES_ROWS = 6  # 5 activités + ligne "Divers"
FICHE_R3_DIRIGEANTS_ROWS = 5
FICHE_R3_ADMIN_ROWS = 5

# Feuille "COMMENTAIRE" du classeur Excel : un encadre de saisie libre (2000
# caracteres max) par note (ou par sous-rubrique pour les notes 5, 12, 16B et
# 16B BIS qui comportent plusieurs zones). Reutilise le mecanisme generique
# note_texte (table note_texte, sheet=cle de note, champ=cle ci-dessous)
# deja en place pour NOTES_TEXTE_DEFS — aucune nouvelle table necessaire.
# Format : (note_key, libelle_du_groupe_ou_None, [(champ, libelle_zone), ...])
COMMENTAIRE_DEFS = [
    ("NOTE 1", None, [("commentaire_0", "Dettes garanties par des sûretés réelles")]),
    ("NOTE 3A", None, [("commentaire_0", "Immobilisations (brutes)")]),
    ("NOTE 3B", None, [("commentaire_0", "Biens pris en location-acquisition")]),
    ("NOTE 3C", None, [("commentaire_0", "Immobilisations (amortissements)")]),
    ("NOTE 3C BIS", None, [("commentaire_0", "Immobilisations (dépréciations)")]),
    ("NOTE 3D", None, [("commentaire_0", "Immobilisations (plus-values et moins-values de cession)")]),
    ("NOTE 4", None, [("commentaire_0", "Immobilisations financières")]),
    ("NOTE 5", "Actif circulant et dettes circulantes HAO", [
        ("commentaire_0", "Actif circulant HAO"),
        ("commentaire_1", "Dettes circulantes HAO"),
    ]),
    ("NOTE 6", None, [("commentaire_0", "Stocks et en cours")]),
    ("NOTE 7", None, [("commentaire_0", "Clients")]),
    ("NOTE 8", None, [("commentaire_0", "Autres créances")]),
    ("NOTE 9", None, [("commentaire_0", "Titres de placement")]),
    ("NOTE 10", None, [("commentaire_0", "Valeurs à encaisser")]),
    ("NOTE 11", None, [("commentaire_0", "Disponibilités")]),
    ("NOTE 12", "Ecarts de conversion et transferts de charges", [
        ("commentaire_0", "Ecarts de conversion"),
        ("commentaire_1", "Transferts de charges"),
    ]),
    ("NOTE 13", None, [("commentaire_0", "Capital")]),
    ("NOTE 14", None, [("commentaire_0", "Primes et réserves")]),
    ("NOTE 15A", None, [("commentaire_0", "Total subventions d'investissement et provisions réglementées")]),
    ("NOTE 15B", None, [("commentaire_0", "Autres fonds propres")]),
    ("NOTE 16A", None, [("commentaire_0", "Dettes financières et ressources assimilées")]),
    ("NOTE 16B", "Engagements de retraite et avantages assimilés : partie 1", [
        ("commentaire_0", "Hypothèses actuarielles"),
        ("commentaire_1", "Variation de la valeur de l'engagement de retraite au cours de l'exercice"),
        ("commentaire_2", "Analyse de sensibilité des hypothèses actuarielles"),
    ]),
    ("NOTE 16B BIS", "Engagements de retraite et avantages assimilés : partie 2", [
        ("commentaire_0", "Actif/passif net comptabilisé au titre des régimes financés"),
        ("commentaire_1", "Valeur actuelle des actifs du régime"),
    ]),
    ("NOTE 16C", None, [("commentaire_0", "Actifs et passifs éventuels")]),
    ("NOTE 17", None, [("commentaire_0", "Fournisseurs d'exploitation")]),
    ("NOTE 18", None, [("commentaire_0", "Dettes fiscales et sociales")]),
    ("NOTE 19", None, [("commentaire_0", "Autres dettes et provisions pour risques à court terme")]),
    ("NOTE 20", None, [("commentaire_0", "Banques, crédit d'escompte et de trésorerie")]),
    ("NOTE 21", None, [("commentaire_0", "Chiffre d'affaires et autres produits")]),
    ("NOTE 22", None, [("commentaire_0", "Achats")]),
    ("NOTE 23", None, [("commentaire_0", "Transports")]),
    ("NOTE 24", None, [("commentaire_0", "Services extérieurs")]),
    ("NOTE 25", None, [("commentaire_0", "Impôts et taxes")]),
    ("NOTE 26", None, [("commentaire_0", "Autres charges")]),
    ("NOTE 27A", None, [("commentaire_0", "Charges de personnel")]),
    ("NOTE 27B", None, [("commentaire_0", "Effectifs, masse salariale et personnel extérieur")]),
    ("NOTE 28", None, [("commentaire_0", "Dotations et charges pour provisions et dépréciations")]),
    ("NOTE 29", None, [("commentaire_0", "Charges et revenus financiers")]),
    ("NOTE 30", None, [("commentaire_0", "Autres charges et produits HAO")]),
    ("SUPPL6", None, [("commentaire_0", "Détail des avantages en nature et en espèces alloués au personnel")]),
]


# ============================================================ DGI-INS =====
# Module complementaire propre a la Cote d'Ivoire (etats destines a la DGI/
# INS, distincts des notes annexes OHADA). Feuilles du classeur source :
# "GARDE (DGI-INS)" (page de garde), "NOTES DGI - INS" (fiche recapitulative
# auto-calculee), "COMP-CHARGES", "COMP-TVA", "COMP-TVA (2)", "SUPPL1" a
# "SUPPL7". Les sheets "BIC"/"BNC"/"BA"/"301"/"302" cites dans la fiche
# recapitulative ne correspondent a aucun onglet reel du classeur (etats
# distincts, hors perimetre de ce fichier) et ne sont donc pas implementes.
#
# Comme NOTE 1/13/31, ces feuilles sont entierement (ou presque) a saisie
# manuelle dans le classeur d'origine : seules les lignes TOTAL portent une
# formule SUM. On reutilise donc le mecanisme generique note3_manuel
# (formulaire note3_sheet=<nom de la feuille>, champs cell_<coord>).

def _manual_num(raw, coord):
    val = raw.get(coord)
    if val in (None, ""):
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


COMP_TVA_COLS = [("G", "Exercice N"), ("H", "Exercice N-1")]
COMP_TVA_LINES = [
    (10, "T.V.A. facturée sur ventes de la période", "manual"),
    (11, "T.V.A. facturée sur prestations de services de la période", "manual"),
    (12, "T.V.A. facturée sur travaux de la période", "manual"),
    (13, "T.V.A. facturée sur production livrée à soi-même de la période", "manual"),
    (14, "T.V.A. sur factures à établir de la période", "manual"),
    (15, "T.V.A. facturée de la période", "total"),
    (16, "T.V.A. exigible de la période", "manual"),
    (17, "T.V.A. récupérable sur immobilisations de la période", "manual"),
    (18, "T.V.A. récupérable sur achats de la période", "manual"),
    (19, "T.V.A. récupérable sur transport de la période", "manual"),
    (20, "T.V.A. récupérable sur services extérieurs et autres charges de la période", "manual"),
    (21, "T.V.A. récupérable sur factures non parvenues de la période", "manual"),
    (22, "T.V.A. transférée par d'autres entités de la période", "manual"),
    (23, "T.V.A. Récupérable de la période", "total"),
    (24, "T.V.A. Récupérée de la période", "manual"),
    (25, "Prorata de déduction", "manual"),
    (26, "État, T.V.A. due", "manual"),
    (27, "État, crédit de T.V.A. à reporter", "manual"),
    (28, "T.V.A. Due ou crédit de T.V.A.", "total"),
]
COMP_TVA_TOTAL_RANGES = {15: (10, 14), 23: (17, 22), 28: (26, 27)}

COMP_TVA2_COLS = COMP_TVA_COLS
COMP_TVA2_LINES = [
    (10, "T.V.A. supportée non déductible sur les immobilisations", "manual"),
    (11, "T.V.A. supportée non déductible sur les achats de biens et de services", "manual"),
    (12, "Total T.V.A. supportée non déductible de la période", "total"),
]
COMP_TVA2_TOTAL_RANGES = {12: (10, 11)}


def _compute_simple_manual_sheet(manual, sheet, lines, cols, total_ranges):
    raw = (manual or {}).get(sheet, {})
    rows = []
    has_data = False
    for row, label, kind in lines:
        vals = {}
        for col, _lbl in cols:
            coord = "%s%d" % (col, row)
            if kind == "total":
                lo, hi = total_ranges[row]
                vals[col] = sum(_manual_num(raw, "%s%d" % (col, r)) for r in range(lo, hi + 1))
            else:
                v = raw.get(coord, "")
                if v not in (None, ""):
                    has_data = True
                vals[col] = v
        rows.append({"row": row, "label": label, "kind": kind, "vals": vals})
    return {"rows": rows, "cols": cols, "has_data": has_data}


def compute_comp_tva(manual):
    return _compute_simple_manual_sheet(manual, "COMP-TVA", COMP_TVA_LINES, COMP_TVA_COLS, COMP_TVA_TOTAL_RANGES)


def compute_comp_tva2(manual):
    return _compute_simple_manual_sheet(manual, "COMP-TVA (2)", COMP_TVA2_LINES, COMP_TVA2_COLS, COMP_TVA2_TOTAL_RANGES)


# SUPPL3 — COMPLEMENT INFORMATIONS ENTITES INDIVIDUELLES (entrepreneurs
# individuels) : 4 lignes, une seule colonne valeur (I, fusionnee I:J), aucun
# total dans le classeur source.
SUPPL3_COLS = [("I", "Valeur (F CFA)")]
SUPPL3_LINES = [
    (9, "Apports financiers au cours de l'exercice", "manual"),
    (10, "Prélèvements financiers au cours de l'exercice", "manual"),
    (11, "Avantages en nature valeur réelle", "manual"),
    (12, "Rémunérations conjoint exploitant", "manual"),
]


def compute_suppl3(manual):
    return _compute_simple_manual_sheet(manual, "SUPPL3", SUPPL3_LINES, SUPPL3_COLS, {})


# SUPPL1 — ELEMENTS STATISTIQUES UEMOA. Feuille la plus complexe du module
# DGI-INS : deux tableaux a lignes dynamiques juxtaposes (memes lignes 14 a
# 65 dans le classeur, mais ensembles de donnees independants), plus deux
# petits tableaux fixes complementaires en bas de page (COMPLEMENT NOTE 27B
# personnel UEMOA, COMPLEMENT NOTE 33 achats production UEMOA).
SUPPL1_ANNEXE_COLS = [
    ("D", "Montant autres États de l'UEMOA", True),
    ("F", "Montant autres États de la Région hors UEMOA", True),
]
SUPPL1_ANNEXE_DEFAULTS = {
    0: "Redevances pour brevets, concessions, licences, marques et droits similaires",
    1: "Redevances pour location de terrains agricoles",
    2: "Biens acquis d'occasion",
}

SUPPL1_PRODUIT_COLS = [
    ("K", "Qté vendue autres États UEMOA", True),
    ("L", "Valeur vendue autres États UEMOA", True),
    ("M", "Qté vendue autres États hors UEMOA", True),
    ("N", "Valeur vendue autres États hors UEMOA", True),
]


def compute_suppl1(manual, note_texte):
    raw = (manual or {}).get("SUPPL1", {})
    texte = (note_texte or {}).get("SUPPL1", {})

    annexe_rows, annexe_totals, annexe_n = _dyn_table(
        raw, texte, SUPPL1_ANNEXE_COLS, "_N_ANNEXE", "AX", "desig_ax", default_n=3)
    for r in annexe_rows:
        if not r["designation"]:
            r["designation"] = SUPPL1_ANNEXE_DEFAULTS.get(r["idx"], "")

    produit_rows, produit_totals, produit_n = _dyn_table(
        raw, texte, SUPPL1_PRODUIT_COLS, "_N_PRODUIT", "PR", "desig_pr", default_n=1)
    for r in produit_rows:
        r["unite"] = texte.get("unite_pr_%d" % r["idx"], "")
    nv_vals = {}
    for col, _label, _manual in SUPPL1_PRODUIT_COLS:
        v = raw.get("NV_PR_%s" % col, "")
        nv_vals[col] = v
        produit_totals[col] += ce._num(v)

    # COMPLEMENT NOTE 27B — personnel propre dans l'UEMOA (effectifs M/F,
    # masse salariale M/F), 4 categories + TOTAL + PERMANENTS/SAISONNIERS.
    n27b_lines = [
        (70, "Cadres supérieurs"),
        (71, "Techniciens supérieurs et cadres moyens"),
        (72, "Techniciens, agents de maîtrise et ouvriers qualifiés"),
        (73, "Employés, manœuvres, ouvriers et apprentis"),
    ]
    n27b_cols = [("B", "Effectifs M"), ("C", "Effectifs F"), ("D", "Masse salariale M"), ("F", "Masse salariale F")]
    n27b_rows = []
    n27b_totals = {c: 0.0 for c, _l in n27b_cols}
    for row, label in n27b_lines:
        vals = {}
        for col, _l in n27b_cols:
            v = raw.get("%s%d" % (col, row), "")
            vals[col] = v
            n27b_totals[col] += _manual_num(raw, "%s%d" % (col, row))
        n27b_rows.append({"row": row, "label": label, "vals": vals})
    n27b_permanents = {c: _manual_num(raw, "%s75" % c) for c, _l in n27b_cols}
    n27b_saisonniers = {c: n27b_totals[c] - n27b_permanents[c] for c, _l in n27b_cols}

    # COMPLEMENT NOTE 33 — achats destinés à la production dans l'UEMOA,
    # colonne unique M (Exercice N), totaux partiels (M71, M76).
    n33_lines = [
        (68, "Achats de marchandises dans l'UEMOA"),
        (69, "Achats de marchandises au groupe dans l'UEMOA"),
        (70, "Achats de marchandises"),
        (71, "Total achats de marchandises"),
        (72, "Achats de matières premières dans l'UEMOA"),
        (73, "Achats de matières premières au groupe dans l'UEMOA"),
        (74, "Achats de matières premières"),
        (75, "(report)"),
        (76, "Total achats de matières premières"),
    ]
    n33_rows = []
    for row, label in n33_lines:
        if row in (71, 76):
            lo = row - 3
            val = sum(_manual_num(raw, "M%d" % r) for r in range(lo, row))
            n33_rows.append({"row": row, "label": label, "kind": "total", "value": val})
        elif row == 75:
            continue
        else:
            n33_rows.append({"row": row, "label": label, "kind": "manual", "value": raw.get("M%d" % row, "")})

    return {
        "annexe_rows": annexe_rows, "annexe_cols": SUPPL1_ANNEXE_COLS, "annexe_totals": annexe_totals, "annexe_n": annexe_n,
        "produit_rows": produit_rows, "produit_cols": SUPPL1_PRODUIT_COLS, "produit_totals": produit_totals,
        "produit_n": produit_n, "produit_nv": nv_vals,
        "n27b_rows": n27b_rows, "n27b_cols": n27b_cols, "n27b_totals": n27b_totals,
        "n27b_permanents": n27b_permanents, "n27b_saisonniers": n27b_saisonniers,
        "n33_rows": n33_rows,
    }


# SUPPL6 — DETAIL DES AVANTAGES EN NATURE ET EN ESPECES ALLOUES AU PERSONNEL :
# 17 lignes a saisie manuelle (colonnes G=forfait/bareme, H=montant reel) +
# TOTAL (1) = SUM(9:25).
SUPPL6_COLS = [("G", "Forfait (Barème) F.CFA"), ("H", "Montant réel F.CFA")]
SUPPL6_LINES = [
    (9, "Loyers logement du personnel", "manual"),
    (10, "Frais d'hôtels & restaurant logement du personnel", "manual"),
    (11, "Frais de cantine", "manual"),
    (12, "Eau logement du personnel", "manual"),
    (13, "Electricité logement du personnel", "manual"),
    (14, "Télécommunication logement du personnel", "manual"),
    (15, "Frais de téléphone mobile", "manual"),
    (16, "Gardiennage logement du personnel", "manual"),
    (17, "Frais de transport en commun du personnel", "manual"),
    (18, "Voyages congés personnel", "manual"),
    (19, "Charge patronale assurance vie allouée au personnel", "manual"),
    (20, "Charge patronale retraite complémentaire", "manual"),
    (21, "Charge patronale autres caisses sociales étrangères", "manual"),
    (22, "Cotisation club golf et autres", "manual"),
    (23, "Frais de scolarité des enfants", "manual"),
    (24, "Dons au personnel", "manual"),
    (25, "Divers autres à préciser", "manual"),
    (26, "TOTAL (1)", "total"),
]
SUPPL6_TOTAL_RANGES = {26: (9, 25)}


def compute_suppl6(manual):
    return _compute_simple_manual_sheet(manual, "SUPPL6", SUPPL6_LINES, SUPPL6_COLS, SUPPL6_TOTAL_RANGES)


# SUPPL7 — CREANCES & DETTES ECHUES DE L'EXERCICE : 2 lignes (une par
# sous-tableau), colonnes H=Principal, I=Interets, sans formule (saisie
# manuelle pure dans le classeur source).
SUPPL7_COLS = [("H", "Principal"), ("I", "Intérêts")]
SUPPL7_LINES = [
    (9, "Créances échues de l'exercice", "manual"),
    (11, "Dettes échues de l'exercice", "manual"),
]


def compute_suppl7(manual):
    return _compute_simple_manual_sheet(manual, "SUPPL7", SUPPL7_LINES, SUPPL7_COLS, {})


# SUPPL5 — DETAIL DES FRAIS ACCESSOIRES SUR ACHATS : grille fixe 6 lignes x 7
# colonnes. 5 colonnes "rubriques" (D,F,H,J,L) + P (immobilisations) sont a
# saisie manuelle ; la colonne N (Total rubriques) est calculee par ligne
# (=SUM des 5 rubriques) et non par le mecanisme generique de total par
# plage de lignes. La ligne 15 (TOTAL) est calculee par colonne (somme des
# lignes 9-14, y compris colonne N).
SUPPL5_RUBRIQUE_COLS = ["D", "F", "H", "J", "L"]
SUPPL5_COLS = [
    ("D", "Achats de marchandises"),
    ("F", "Achats de matières premières et fournitures liées"),
    ("H", "Achats stockés de matières et fournitures consommables"),
    ("J", "Achats d'emballages"),
    ("L", "Autres achats"),
    ("N", "Total rubriques"),
    ("P", "Immobilisations"),
]
SUPPL5_LINES = [
    (9, "Droits de douane"),
    (10, "Frets et transports sur achats"),
    (11, "Assurances transport sur achats"),
    (12, "Commissions et courtages sur achats"),
    (13, "Rémunération du transitaire"),
    (14, "Autres frais accessoires d'achat"),
]


def compute_suppl5(manual):
    raw = (manual or {}).get("SUPPL5", {})
    cols_all = [c for c, _l in SUPPL5_COLS]
    rows = []
    for row, label in SUPPL5_LINES:
        vals = {}
        for col in SUPPL5_RUBRIQUE_COLS + ["P"]:
            vals[col] = raw.get("%s%d" % (col, row), "")
        vals["N"] = sum(_manual_num(raw, "%s%d" % (c, row)) for c in SUPPL5_RUBRIQUE_COLS)
        rows.append({"row": row, "label": label, "kind": "manual", "vals": vals, "computed_cols": ["N"]})
    total_vals = {}
    for col in cols_all:
        total_vals[col] = sum(_manual_num(raw, "%s%d" % (col, r)) if col != "N"
                               else sum(_manual_num(raw, "%s%d" % (c, r)) for c in SUPPL5_RUBRIQUE_COLS)
                               for r, _l in SUPPL5_LINES)
    rows.append({"row": 15, "label": "Total frais accessoires d'achats", "kind": "total", "vals": total_vals, "computed_cols": cols_all})
    return {"rows": rows, "cols": SUPPL5_COLS}


# SUPPL2 — REPARTITION DU RESULTAT FISCAL DES SOCIETES DE PERSONNES : tableau
# a lignes dynamiques (un associe par ligne). Reutilise _dyn_table pour les
# 3 colonnes numeriques (G/H/I) ; les champs texte (adresse, NCC, reference
# de note) sont stockes via note_texte avec des cles dediees (memes
# conventions que desig_prefix="desig" gere par _dyn_table pour le nom).
SUPPL2_COLS = [
    ("G", "Part résultats avant réintégration rémunération et part IMF", True),
    ("H", "Rémunération, avantages en nature et part IMF", True),
    ("I", "Total imposable au nom de chaque associé", True),
]


def compute_suppl2(manual, note_texte):
    raw = (manual or {}).get("SUPPL2", {})
    texte = (note_texte or {}).get("SUPPL2", {})
    rows, totals, n = _dyn_table(raw, texte, SUPPL2_COLS, "_N_DETAIL", "D", "desig", default_n=1)
    for r in rows:
        i = r["idx"]
        r["adresse"] = texte.get("adresse_%d" % i, "")
        r["ncc"] = texte.get("ncc_%d" % i, "")
        r["ref"] = texte.get("ref_%d" % i, "")
    return {"rows": rows, "n": n, "totals": totals, "cols": SUPPL2_COLS}


# SUPPL4 — TABLEAU DES AMORTISSEMENTS ET INVENTAIRE DES IMMOBILISATIONS :
# tableau a lignes dynamiques. Colonnes E/G/H/I/L a saisie manuelle (taux,
# valeur d'acquisition, amortissements anterieurs/exercice, prix de cession) ;
# A (numero de compte), B (designation) et F (date de mise en service) sont
# du texte libre (note_texte) ; J (total amortissements), K (valeur
# residuelle), M (plus-value) et N (moins-value) sont calcules par ligne,
# exactement comme dans le classeur source (J=H+I, K=G-J,
# M=SI(L>K,L-K,""), N=SI(L<K,K-L,"")).
SUPPL4_COLS = [
    ("E", "Taux amort. %", True),
    ("G", "Valeur d'acquisition", True),
    ("H", "Amortissements antérieurs", True),
    ("I", "Amortissements de l'exercice", True),
    ("L", "Prix de cession", True),
]


def compute_suppl4(manual, note_texte):
    raw = (manual or {}).get("SUPPL4", {})
    texte = (note_texte or {}).get("SUPPL4", {})
    rows, totals, n = _dyn_table(raw, texte, SUPPL4_COLS, "_N_DETAIL", "D", "desig", default_n=1)
    for col in ("J", "K", "M", "N"):
        totals[col] = 0.0
    for r in rows:
        i = r["idx"]
        r["compte"] = texte.get("compte_%d" % i, "")
        r["date_service"] = texte.get("date_%d" % i, "")
        g = ce._num(r["vals"].get("G"))
        h = ce._num(r["vals"].get("H"))
        i_amort = ce._num(r["vals"].get("I"))
        l = ce._num(r["vals"].get("L"))
        j = h + i_amort
        k = g - j
        m = (l - k) if l > k else ""
        n_ = (k - l) if (l not in (0, "", None) and l < k) else ""
        r["vals"]["J"] = j
        r["vals"]["K"] = k
        r["vals"]["M"] = m
        r["vals"]["N"] = n_
        totals["J"] += j
        totals["K"] += k
        totals["M"] += ce._num(m)
        totals["N"] += ce._num(n_)
    return {"rows": rows, "n": n, "totals": totals, "cols": SUPPL4_COLS}


# COMP-CHARGES — ETAT COMPLEMENTAIRE N°1 : DETAIL DES CHARGES (276 lignes).
# Feuille entierement calculee a partir de la balance (SUMIFS par plage de
# compte, sur le modele de NOTE 31) : aucune saisie manuelle. Chaque ligne de
# detail (kind="range") calcule Debit-Credit sur la plage [lo,hi) de comptes ;
# les lignes de sous-total (codes lettres RA-RN) sont kind="sum" (plage de
# lignes contigue), "summulti" (plusieurs plages de lignes, cas des sauts de
# page Excel) ou "ref" (simple report d'une autre ligne, cas des lignes de
# variation de stock qui dupliquent la ligne de detail correspondante). La
# ligne 275 (TOTAL DES CHARGES ORDINAIRES) est kind="sumrefs" (somme d'une
# liste explicite de lignes de sous-total) et la ligne 276 (ecart de
# verification) est kind="ecart" (SUMIFS global comptes 60000-70000 moins le
# total ligne 275 — doit etre nul si la note est complete). Extrait
# automatiquement des formules du classeur via openpyxl (cf. memoire projet).
COMP_CHARGES_LINES = [
    (11, 6011, 'Achats de marchandises dans la région', 'range', (60110, 60120)),
    (12, 6012, 'Achats de marchandises hors région', 'range', (60120, 60130)),
    (13, 6013, 'Achats de marchandises aux entités du groupe dans la région', 'range', (60130, 60140)),
    (14, 6014, 'Achats de marchandises aux entités du groupe hors région', 'range', (60140, 60150)),
    (15, 6015, 'Frais sur achats', 'range', (60150, 60160)),
    (16, 6019, 'Rabais, remises et ristournes obtenus (non ventilés)', 'range', (60190, 60200)),
    (17, 'RA', 'Achats de marchandises', 'sum', (11, 16)),
    (18, 6031, 'Variation des stocks de marchandises', 'range', (60310, 60320)),
    (19, 'RB', 'Variation de stocks', 'ref', 18),
    (20, 6021, 'Achats de matières premières et fournitures liées dans la région', 'range', (60210, 60220)),
    (21, 6022, 'Achats de matières premières et fournitures liées hors région', 'range', (60220, 60230)),
    (22, 6023, 'Achats de matières premières et fournitures liées aux entités du groupe dans la région', 'range', (60230, 60240)),
    (23, 6024, 'Achats de matières premières et fournitures liées aux entités du groupe hors région', 'range', (60240, 60250)),
    (24, 6025, 'Frais sur achats', 'range', (60250, 60260)),
    (25, 6029, 'Rabais, remises et ristournes obtenus (non ventilés)', 'range', (60290, 60300)),
    (26, 'RC', 'Achats de matières premières et fournitures liées', 'sum', (20, 25)),
    (27, 6032, 'Variation des stocks de matières premières et fournitures liées', 'range', (60320, 60330)),
    (28, 'RD', 'Variation de stocks de matières premières', 'ref', 27),
    (29, 6041, 'Matières consommables', 'range', (60410, 60420)),
    (30, 6042, 'Matières combustibles', 'range', (60420, 60430)),
    (31, 6043, "Produits d'entretien", 'range', (60430, 60440)),
    (32, 6044, "Fournitures d'atelier et d'usine", 'range', (60440, 60450)),
    (33, 6045, 'Frais sur achats', 'range', (60450, 60460)),
    (34, 6046, 'Fournitures de magasin', 'range', (60460, 60470)),
    (35, 6047, 'Fournitures de bureau', 'range', (60470, 60480)),
    (36, 6049, 'Rabais, remises et ristournes obtenus (non ventilés)', 'range', (60490, 60500)),
    (37, 6051, 'Fournitures non stockables-Eau', 'range', (60510, 60520)),
    (38, 6052, 'Fournitures non stockables-Electricité', 'range', (60520, 60530)),
    (39, 6053, 'Fournitures non stockables-Autres énergies', 'range', (60530, 60540)),
    (40, 6054, "Fournitures d'entretien non stockables", 'range', (60540, 60550)),
    (41, 6055, 'Fournitures de bureau non stockables', 'range', (60550, 60560)),
    (42, 6056, 'Achats de petit matériel et outillage', 'range', (60560, 60570)),
    (43, 6057, "Achats d'études et prestations de services", 'range', (60570, 60580)),
    (44, 6058, 'Achats de travaux, matériels et équipements', 'range', (60580, 60590)),
    (45, 6059, 'Rabais, remises et ristournes obtenus (non ventilés)', 'range', (60590, 60600)),
    (46, 6081, 'Emballages perdus', 'range', (60810, 60820)),
    (47, 6082, 'Emballages récupérables non identifiables', 'range', (60820, 60830)),
    (48, 6083, 'Emballages à usage mixte', 'range', (60830, 60840)),
    (49, 6085, 'Frais sur achats', 'range', (60850, 60860)),
    (50, 6089, 'Rabais, remises et ristournes obtenus (non ventilés)', 'range', (60890, 60900)),
    (51, 'RE', 'Autres achats', 'sum', (29, 50)),
    (52, 6033, "Variation des stocks d'autres approvisionnements", 'range', (60330, 60340)),
    (53, 'RF', "Variation de stocks d'autres approvisionnements", 'ref', 52),
    (54, 612, 'Transports sur ventes', 'range', (61200, 61300)),
    (55, 613, 'Transports pour le compte de tiers', 'range', (61300, 61400)),
    (56, 614, 'Transports du personnel', 'range', (61400, 61500)),
    (57, 616, 'Transports de plis', 'range', (61600, 61700)),
    (58, 6181, 'Voyages et déplacements', 'range', (61810, 61820)),
    (59, 6182, 'Transports entre établissements et chantiers', 'range', (61820, 61830)),
    (60, 6183, 'Transports administratifs', 'range', (61830, 61840)),
    (61, 619, 'Rabais, remises et ristournes obtenus (non ventilés)', 'range', (61900, 62000)),
    (62, 'RG', 'Transports', 'sum', (54, 60)),
    (63, 621, 'Sous-traitance générale', 'range', (62100, 62200)),
    (64, 6221, 'Location de terrains', 'range', (62210, 62220)),
    (65, 6222, 'Location de bâtiments', 'range', (62220, 62230)),
    (66, 6223, 'Location de matériels et outillages', 'range', (62230, 62240)),
    (67, 6224, 'Malis sur emballages', 'range', (62240, 62250)),
    (68, 6225, "Locations d'emballages", 'range', (62250, 62260)),
    (69, 6226, 'Fermages et loyers du foncier', 'range', (62260, 62270)),
    (70, 6228, 'Locations  et charges locatives diverses', 'range', (62280, 62290)),
    (71, 6232, 'Crédit-bail immobilier', 'range', (62320, 62330)),
    (72, 6233, 'Crédit-bail mobilier', 'range', (62330, 62340)),
    (73, 6234, 'Location-vente', 'range', (62340, 62350)),
    (85, 6238, 'Autres contrats de location-acquisition', 'range', (62380, 62390)),
    (86, 6241, 'Entretien et réparation des biens immobiliers', 'range', (62410, 62420)),
    (87, 6242, 'Entretien et réparation des biens mobiliers', 'range', (62420, 62430)),
    (88, 6243, 'Maintenance', 'range', (62430, 62440)),
    (89, 6244, 'Charges de démentellement et remise en état', 'range', (62440, 62450)),
    (90, 6248, 'Autre entretiens et réparation', 'range', (62480, 62490)),
    (91, 6251, 'Assurances multirisques ', 'range', (62510, 62520)),
    (92, 6252, 'Assurances matériels de transport', 'range', (62520, 62530)),
    (93, 6253, "Assurances risques d'exploitation", 'range', (62530, 62540)),
    (94, 6254, 'Assurances responsabilité du producteur', 'range', (62540, 62550)),
    (95, 6255, 'Assurances insolvabilité clients', 'range', (62550, 62560)),
    (96, 6257, 'Assurances transport sur ventes', 'range', (62570, 62580)),
    (97, 6258, "Autres primes d'assurances", 'range', (62580, 62590)),
    (98, 6261, 'Etudes et recherches', 'range', (62610, 62620)),
    (99, 6265, 'Documentation générale', 'range', (62650, 62660)),
    (100, 6266, 'Documentation technique', 'range', (62660, 62670)),
    (101, 6271, 'Annonces, insertions', 'range', (62710, 62720)),
    (102, 6272, 'Catalogues, imprimés publicitaires', 'range', (62720, 62730)),
    (103, 6273, 'Échantillons', 'range', (62730, 62740)),
    (104, 6274, 'Foires et expositions', 'range', (62740, 62750)),
    (105, 6275, 'Publications', 'range', (62750, 62760)),
    (106, 6276, 'Cadeaux à la clientèle', 'range', (62760, 62770)),
    (107, 6277, 'Frais de colloques, séminaires, conférences', 'range', (62770, 62780)),
    (108, 6278, 'Autres charges de publicité et relations publiques', 'range', (62780, 62790)),
    (109, 6281, 'Frais de téléphone', 'range', (62810, 62820)),
    (110, 6282, 'Frais de télex', 'range', (62820, 62830)),
    (111, 6283, 'Frais de télécopie', 'range', (62830, 62840)),
    (112, 6284, "Frais d'internet", 'range', (62840, 62850)),
    (113, 6288, 'Autres frais de télécommunications', 'range', (62880, 62890)),
    (114, 6311, 'Frais sur titres (vente, garde)', 'range', (63110, 63120)),
    (115, 6312, 'Frais sur effets', 'range', (63120, 63130)),
    (116, 6313, 'Location de coffres', 'range', (63130, 63140)),
    (117, 6314, "Commissions d'affacturage et de titrisation", 'range', (63140, 63150)),
    (118, 6315, 'Commissions sur cartes de crédit', 'range', (63150, 63160)),
    (119, 6316, "Frais d'émission d'emprunts", 'range', (63160, 63170)),
    (120, 6317, 'Frais sur instruments monnaie électronique', 'range', (63170, 63180)),
    (121, 6318, 'Autres frais bancaires', 'range', (63180, 63190)),
    (122, 6322, 'Commissions et courtages sur ventes', 'range', (63220, 63230)),
    (123, 6324, 'Honoraires des professions règlementées', 'range', (63240, 63250)),
    (124, 6325, "Frais d'actes et de contentieux", 'range', (63250, 63260)),
    (125, 6326, "Rémunérations d'affacturage et de titrisation", 'range', (63260, 63270)),
    (126, 6327, 'Rémunérations des autres prestataires de services', 'range', (63270, 63280)),
    (127, 6328, 'Divers frais', 'range', (63280, 63290)),
    (128, 633, 'Frais de formation du personnel', 'range', (63300, 63400)),
    (129, 6342, 'Redevances pour brevets, licences', 'range', (63420, 63430)),
    (130, 6343, 'Redevances pour logiciels', 'range', (63430, 63440)),
    (131, 6344, 'Redevances pour marques', 'range', (63440, 63450)),
    (132, 6345, 'Redevances pour sites internet', 'range', (63450, 63460)),
    (133, 6346, 'Redevances pour concessions, droits et valeurs similaires', 'range', (63460, 63470)),
    (134, 6351, 'Cotisations', 'range', (63510, 63520)),
    (135, 6358, 'Concours divers', 'range', (63580, 63590)),
    (136, 6371, 'Personnel intérimaire', 'range', (63710, 63720)),
    (137, 6372, "Personnel détaché ou prêté à l'entité", 'range', (63720, 63730)),
    (138, 6381, 'Frais de recrutement du personnel', 'range', (63810, 63820)),
    (139, 6382, 'Frais de déménagement', 'range', (63820, 63830)),
    (140, 6383, 'Réceptions', 'range', (63830, 63840)),
    (141, 6384, 'Missions', 'range', (63840, 63850)),
    (142, 6385, 'Charges de copropriété', 'range', (63850, 63860)),
    (143, 6388, 'Charges externes diverses', 'range', (63880, 63890)),
    (144, 'RH', 'Services extérieurs', 'summulti', [(63, 73), (85, 143)]),
    (145, 6411, 'Impôts fonciers et taxes annexes', 'range', (64110, 64120)),
    (146, 6412, 'Patentes, licences et taxes annexes', 'range', (64120, 64130)),
    (147, 6413, 'Taxes sur appointements et salaires', 'range', (64130, 64140)),
    (148, 6414, "Taxes d'apprentissage", 'range', (64140, 64150)),
    (160, 6415, 'Formation professionnelle continue', 'range', (64150, 64160)),
    (161, 6418, 'Autres impôts et taxes directs', 'range', (64180, 64190)),
    (162, 645, 'Impôts et taxes indirects', 'range', (64500, 64600)),
    (163, 6461, 'Droits de mutation', 'range', (64610, 64620)),
    (164, 6462, 'Droits de timbre', 'range', (64620, 64630)),
    (165, 6463, 'Taxes sur les véhicules de société', 'range', (64630, 64640)),
    (166, 6464, 'Vignettes', 'range', (64640, 64650)),
    (167, 6468, "Autres droits d'enregistrement", 'range', (64680, 64690)),
    (168, 6471, "Pénalités d'assiette, impôts directs", 'range', (64710, 64720)),
    (169, 6472, "Pénalités d'assiette, impôts indirects", 'range', (64720, 64730)),
    (170, 6473, 'Pénalités de recouvrement, impôts directs', 'range', (64730, 64740)),
    (171, 6474, 'Pénalités de recouvrement, impôts indirects', 'range', (64740, 64750)),
    (172, 6478, 'Autres pénalités et amendes fiscales', 'range', (64780, 64790)),
    (173, 648, 'Autres impôts et taxes', 'range', (64800, 64900)),
    (174, 'RI', 'Impôts et taxes', 'summulti', [(145, 148), (160, 173)]),
    (175, 6511, 'Pertes sur créances clients', 'range', (65110, 65120)),
    (176, 6515, 'Pertes sur autres débiteurs', 'range', (65150, 65160)),
    (177, 6521, 'Quote-part transférée de bénéfices (comptabilité du gérant)', 'range', (65210, 65220)),
    (178, 6525, 'Pertes imputées par transfert (comptabilité des associés non gérants)', 'range', (65250, 65260)),
    (179, 6541, "Valeur comptable des cessions courantes d'immobilisations incorporelles", 'range', (65410, 65420)),
    (180, 6542, "Valeur comptable des cessions courantes d'immobilisations corporelles", 'range', (65420, 65430)),
    (181, 656, 'Perte de change sur créances et dettes commerciales', 'range', (65600, 65700)),
    (182, 657, 'Pénalités et amendes pénales', 'range', (65700, 65800)),
    (183, 6581, "Indemnités de fonction et autres rémunérations d'administrateurs", 'range', (65810, 65820)),
    (184, 6582, 'Dons', 'range', (65820, 65830)),
    (185, 6583, 'Mécénat', 'range', (65830, 65840)),
    (186, 6588, 'Autres charges diverses', 'range', (65880, 65890)),
    (187, 6591, "Charges pour dépréciation et provisions d'exploitation sur risques à court terme", 'range', (65910, 65920)),
    (188, 6593, "Charges pour dépréciation et provisions pour risque à court terme d'exploitation sur stocks", 'range', (65930, 65940)),
    (189, 6594, "Charges pour dépréciation et provisions pour risque à court terme d'exploitation sur créances", 'range', (65940, 65950)),
    (190, 6598, "Autres charges pour dépréciations et provisions pour risques à court terme d'exploitation", 'range', (65980, 65990)),
    (191, 'RJ', 'Autres charges', 'sum', (175, 190)),
    (192, 6611, 'Appointements salaires et commissions versés au personnel national', 'range', (66110, 66120)),
    (193, 6612, 'Primes et gratifications versées au personnel national', 'range', (66120, 66130)),
    (194, 6613, 'Congés payés versés au personnel national', 'range', (66130, 66140)),
    (195, 6614, "Indemnités de préavis, de licenciement et de recherche d'embauche versées au personnel national", 'range', (66140, 66150)),
    (196, 6615, 'Indemnités de maladie versées aux travailleurs nationaux', 'range', (66150, 66160)),
    (197, 6616, 'Supplément familial versé au personnel national', 'range', (66160, 66170)),
    (198, 6617, 'Avantages en nature du personnel national', 'range', (66170, 66180)),
    (199, 6618, 'Autres rénumérations directes versées au personnel national', 'range', (66180, 66190)),
    (200, 6621, 'Appointements salaires et commissions versés au personnel non national', 'range', (66210, 66220)),
    (201, 6622, 'Primes et gratifications versées au personnel non national', 'range', (66220, 66230)),
    (202, 6623, 'Congés payés versés au personnel non national', 'range', (66230, 66240)),
    (203, 6624, "Indemnités de préavis, de licenciement et de recherche d'embauche versées au personnel non national", 'range', (66240, 66250)),
    (204, 6625, 'Indemnités de maladie versées aux travailleurs non nationaux', 'range', (66250, 66260)),
    (205, 6626, 'Supplément familial versé au personnel non national', 'range', (66260, 66270)),
    (206, 6627, 'Avantages en nature du personnel non national', 'range', (66270, 66280)),
    (207, 6628, 'Autres rénumérations directes versées au personnel non national', 'range', (66280, 66290)),
    (208, 6631, 'Indemnités forfaitaires de logement versées au personnel', 'range', (66310, 66320)),
    (209, 6632, 'Indemnités forfaitaires de représentation versées au personnel', 'range', (66320, 66330)),
    (210, 6633, "Indemnités forfaitaires d'expatriation versées au personnel", 'range', (66330, 66340)),
    (211, 6634, 'Indemnités forfaitaires de transport versées au personnel', 'range', (66340, 66350)),
    (212, 6638, 'Autres indemnités et avantages divers versés au personnel', 'range', (66380, 66390)),
    (213, 6641, 'Charges sociales sur rémunération du personnel national', 'range', (66410, 66420)),
    (214, 6642, 'Charges sociales sur rémunération du personnel non national', 'range', (66420, 66430)),
    (215, 6661, "Rémunérations du travail de l'exploitant individuel", 'range', (66610, 66620)),
    (216, 6662, "Charges sociales de l'exploitant individuel", 'range', (66620, 66630)),
    (217, 6671, 'Rémunérations transférée du personnel intérimaire', 'range', (66710, 66720)),
    (218, 6672, "Rémunérations transférée du personnel détaché ou prêté à l'entité", 'range', (66720, 66730)),
    (219, 6681, "Versements aux Syndicats et Comités d'entreprise, d'établissement", 'range', (66810, 66820)),
    (220, 6682, "Versements aux Comités d'hygiène et de sécurité", 'range', (66820, 66830)),
    (221, 6683, 'Versements et contributions aux autres œuvres sociales', 'range', (66830, 66840)),
    (233, 6684, 'Médecine du travail et pharmacie', 'range', (66840, 66850)),
    (234, 6685, 'Assurances et organismes de santé', 'range', (66850, 66860)),
    (235, 6686, 'Assurances retraite et fonds de pension', 'range', (66860, 66870)),
    (236, 6687, 'Majorations et pénalités sociales', 'range', (66870, 66880)),
    (237, 6688, 'Charges sociales diverses', 'range', (66880, 66890)),
    (238, 'RK', 'Charges de personnel', 'summulti', [(192, 221), (233, 237)]),
    (239, 6711, 'Intérêts des emprunts obligataires', 'range', (67110, 67120)),
    (240, 6712, 'Intérêts des emprunts auprès des établissements de crédit', 'range', (67120, 67130)),
    (241, 6713, 'Intérêts des dettes liées à des participations', 'range', (67130, 67140)),
    (242, 6714, 'Intérêts des primes de remboursement des obligations', 'range', (67140, 67150)),
    (243, 6721, 'Intérêts dans loyers de location acquisition/crédit-bail immobilier', 'range', (67210, 67220)),
    (244, 6722, 'Intérêts dans loyers de location acquisition/crédit-bail mobilier', 'range', (67220, 67230)),
    (245, 6723, 'Intérêts dans loyers de location acquisition/location-vente', 'range', (67230, 67240)),
    (246, 6728, 'Intérêts dans loyers des autres locations acquisition', 'range', (67280, 67290)),
    (247, 673, 'Escomptes accordés', 'range', (67300, 67400)),
    (248, 6741, 'Intérêts sur avances reçues et dépôts créditeurs', 'range', (67410, 67420)),
    (249, 6742, 'Intérêts sur Comptes courants bloqués', 'range', (67420, 67430)),
    (250, 6743, 'Intérêts sur obligations cautionnées', 'range', (67430, 67440)),
    (251, 6744, 'Intérêts sur dettes commerciales', 'range', (67440, 67450)),
    (252, 6745, 'Intérêts bancaires et sur opérations de financement (escompte…)', 'range', (67450, 67460)),
    (253, 6748, 'Intérêts sur dettes diverses                     ', 'range', (67480, 67490)),
    (254, 675, 'Escomptes des effets de commerce', 'range', (67500, 67600)),
    (255, 676, 'Pertes de change financières', 'range', (67600, 67700)),
    (256, 6771, 'Pertes sur cessions de titre de placement', 'range', (67710, 67720)),
    (257, 6772, 'Malis provenant d’attribution gratuite d’actions au personnel salarié et aux dirigeants', 'range', (67720, 67730)),
    (258, 6781, 'Pertes et charges sur rentes viagères', 'range', (67810, 67820)),
    (259, 6782, 'Pertes et charges sur opérations financières', 'range', (67820, 67830)),
    (260, 6784, 'Pertes et charges sur instrument de trésorerie', 'range', (67840, 67850)),
    (261, 6791, 'Charges pour dépréciations et provisions sur risques financiers à court terme', 'range', (67910, 67920)),
    (262, 6795, 'Charges pour dépréciations et provisions sur titres de placement', 'range', (67950, 67960)),
    (263, 6798, 'Autres charges pour dépréciations et provisions pour risques à court terme financières', 'range', (67980, 67990)),
    (264, 'RM', 'Frais financiers et charges assimiliés', 'sum', (239, 263)),
    (265, 6812, 'Dotations aux amortissements des immobilisations incorporelles', 'range', (68120, 68130)),
    (266, 6813, 'Dotations aux amortissements des immobilisations corporelles', 'range', (68130, 68140)),
    (267, 6911, 'Dotations aux provisions pour risques et charges', 'range', (69110, 69120)),
    (268, 6913, 'Dotations aux dépréciations des immobilisations incorporelles', 'range', (69130, 69140)),
    (269, 6914, 'Dotations aux dépréciations des immobilisations corporelles', 'range', (69140, 69150)),
    (270, 'RL', 'Dotations aux amortissements, aux provisions et dépréciations', 'sum', (265, 269)),
    (271, 6971, 'Dotations aux provisions pour risques et charges financières', 'range', (69710, 69720)),
    (272, 6972, 'Dotations aux dépréciations des immobilisations financières', 'range', (69720, 69730)),
    (273, 'RN', 'Dotations aux provisions et aux dépréciations financières', 'sum', (271, 272)),
    (275, '', 'TOTAL DES CHARGES ORDINAIRES', 'sumrefs', [17, 19, 26, 28, 51, 53, 62, 144, 174, 191, 238, 264, 270, 273]),
    (276, '', 'Ecart (charges 60-69 hors note - total)', 'ecart', (60000, 70000, 275)),
]


def compute_comp_charges(balN, balN1):
    SUMIFS_ = ce.make_sumifs(balN, balN1 or [])

    def sf(table, field, lo, hi):
        return SUMIFS_(table, field, ">=", lo, "<", hi)

    def range_val(table, lo, hi):
        return sf(table, "BS_Debit", lo, hi) - sf(table, "BS_Credit", lo, hi)

    gvals, hvals = {}, {}
    rows = []
    for row, code, label, kind, params in COMP_CHARGES_LINES:
        if kind == "range":
            lo, hi = params
            g = range_val("BalanceN", lo, hi)
            h = range_val("BalanceN1", lo, hi)
        elif kind == "sum":
            lo, hi = params
            g = sum(gvals.get(r, 0.0) for r in range(lo, hi + 1))
            h = sum(hvals.get(r, 0.0) for r in range(lo, hi + 1))
        elif kind == "summulti":
            g = sum(sum(gvals.get(r, 0.0) for r in range(lo, hi + 1)) for lo, hi in params)
            h = sum(sum(hvals.get(r, 0.0) for r in range(lo, hi + 1)) for lo, hi in params)
        elif kind == "ref":
            g = gvals.get(params, 0.0)
            h = hvals.get(params, 0.0)
        elif kind == "sumrefs":
            g = sum(gvals.get(r, 0.0) for r in params)
            h = sum(hvals.get(r, 0.0) for r in params)
        elif kind == "ecart":
            lo, hi, total_row = params
            g = range_val("BalanceN", lo, hi) - gvals.get(total_row, 0.0)
            h = range_val("BalanceN1", lo, hi) - hvals.get(total_row, 0.0)
        else:
            g = h = 0.0
        gvals[row] = g
        hvals[row] = h
        var_valeur = g - h
        if h:
            var_pct = (g - h) / h
        elif g:
            var_pct = 1 if g > 0 else -1
        else:
            var_pct = ""
        rows.append({
            "row": row,
            "code": code,
            "label": label,
            "g": g,
            "h": h,
            "var_valeur": var_valeur,
            "var_pct": var_pct,
            "is_total": not isinstance(code, int),
        })
    return {"rows": rows}


# "NOTES DGI - INS" — fiche recapitulative des etats supplementaires
# presentes. Reproduit la logique decrite dans la note de bas de page de la
# feuille source (A35) : "les pages ayant au moins un tableau renseigne
# seront considerees comme applicables, les autres comme non-applicables."
# BIC/BNC/BA/301/302 sont des formulaires externes a ce classeur (regimes
# fiscaux autres que le reel normal) : ils sont hors perimetre de cette
# application et affiches comme tels plutot que comme "non applicable" par
# defaut (cf. memoire projet [[project_tafiroha_dgi_ins_module]]).
NOTES_DGI_INS_RECAP_DEFS = [
    ("COMP-CHARGES", "ETAT COMPLEMENTAIRE : DETAIL DES CHARGES"),
    ("COMP-TVA (1)", "ETAT COMPLEMENTAIRE : TVA"),
    ("COMP-TVA (2)", "ETAT COMPLEMENTAIRE POUR l'INS : TVA SUPPORTEE NON DEDUCTIBLE"),
    ("SUPPL 1", "ELEMENTS STATISTIQUES UEMOA"),
    ("SUPPL 2", "REPARTITION DU RESULTAT FISCAL DES SOCIETES DE PERSONNES"),
    ("SUPPL 3", "COMPLEMENT INFORMATIONS ENTITES INDIVIDUELLES"),
    ("SUPPL 4", "TABLEAU DES AMORTISSEMENTS ET INVENTAIRE DES IMMOBILISATIONS"),
    ("SUPPL 5", "DETAIL DES FRAIS ACCESSOIRES SUR ACHATS"),
    ("SUPPL 6", "DETAIL DES AVANTAGES EN NATURE ET EN ESPECES IMPOSES ALLOUES AU PERSONNEL"),
    ("SUPPL 7", "CREANCES ET DETTES ECHUES DE L'EXERCICE"),
    ("BIC", "DETERMINATION DU BENEFICE INDUSTRIEL OU COMMERCIAL"),
    ("BNC", "DETERMINATION DU BENEFICE NON COMMERCIAL"),
    ("BA", "DETERMINATION DU BENEFICE AGRICOLE"),
    ("301", "DECLARATION DES REMUNERATIONS VERSEES AUX SALARIES DE L'ENTREPRISE"),
    ("302", "DECLARATION DES REMUNERATIONS VERSEES A DES CONTRIBUABLES N'AYANT PAS LA QUALITE DE SALARIES DE L'ENTREPRISE"),
]
NOTES_DGI_INS_HORS_PERIMETRE = {"BIC", "BNC", "BA", "301", "302"}


def _has_any_manual(manual, sheet):
    raw = (manual or {}).get(sheet, {})
    return any(ce._num(v) for v in raw.values())


def _has_any_texte(note_texte, sheet):
    raw = (note_texte or {}).get(sheet, {})
    return any(v not in (None, "") for v in raw.values())


def compute_notes_dgi_ins_recap(comp_charges, comp_tva, comp_tva2, suppl1, suppl2,
                                 suppl3, suppl4, suppl5, suppl6, suppl7,
                                 manual, note_texte):
    applicable = {
        "COMP-CHARGES": any(ce._num(r["g"]) for r in comp_charges["rows"]),
        "COMP-TVA (1)": bool(comp_tva.get("has_data")),
        "COMP-TVA (2)": bool(comp_tva2.get("has_data")),
        "SUPPL 1": _has_any_manual(manual, "SUPPL1") or _has_any_texte(note_texte, "SUPPL1"),
        "SUPPL 2": _has_any_manual(manual, "SUPPL2") or _has_any_texte(note_texte, "SUPPL2"),
        "SUPPL 3": bool(suppl3.get("has_data")),
        "SUPPL 4": _has_any_manual(manual, "SUPPL4") or _has_any_texte(note_texte, "SUPPL4"),
        "SUPPL 5": any(ce._num(v) for r in suppl5["rows"] if r["kind"] == "manual" for v in r["vals"].values()),
        "SUPPL 6": bool(suppl6.get("has_data")),
        "SUPPL 7": bool(suppl7.get("has_data")),
    }
    rows = []
    for code, label in NOTES_DGI_INS_RECAP_DEFS:
        hors_perimetre = code in NOTES_DGI_INS_HORS_PERIMETRE
        rows.append({
            "code": code,
            "label": label,
            "applicable": None if hors_perimetre else applicable.get(code, False),
            "hors_perimetre": hors_perimetre,
        })
    return {"rows": rows}


def compute_notes_lot_d(balN, balN1, manual):
    """NOTE 12, 15B, 16B/16B BIS, 37 :
    contrairement a leur etiquette de feuilles "a saisie manuelle", ces
    feuilles contiennent de vraies formules (SUM, IFERROR, INDIRECT/ADDRESS
    dynamique) ; seules les cellules de detail sont a saisie manuelle. Ne
    portent pas sur la balance (pas de SUMIFS), mais compute_workbook a
    quand meme besoin de balN/balN1 pour son interface standard."""
    sheets = list(dict.fromkeys(name for name, _, _, _ in NOTES_LOT_D))
    caches = ce.compute_workbook(sheets, balN or [], balN1 or [], manual=manual, row_range=(6, 60))
    groups = {}
    order = []
    for sheet, lines, cols, title in NOTES_LOT_D:
        if sheet not in groups:
            groups[sheet] = []
            order.append(sheet)
        groups[sheet].append({
            "title": title,
            "cols": cols,
            "rows": build_note3_table(caches.get(sheet, {}), lines, cols),
        })
    # Une feuille (ex: NOTE 16B, NOTE 16B BIS) peut comporter plusieurs
    # tableaux distincts dans le classeur source (hypothèses, variation,
    # sensibilité, etc.) : on les regroupe ici sous une seule carte par
    # feuille au lieu d'une carte (et d'un id HTML) dupliquée par tableau.
    return [{"sheet": sheet, "tables": groups[sheet]} for sheet in order]


def build_note3_table(cache, lines, cols):
    rows = []
    for line in lines:
        if len(line) == 4:
            row, label, is_total, manual_cols = line
        else:
            row, label, is_total = line
            manual_cols = ()
        vals = {}
        for col, _label, _manual in cols:
            vals[col] = cache.get("%s%d" % (col, row))
        rows.append({"row": row, "label": label, "total": is_total, "vals": vals,
                     "manual_cols": manual_cols})
    return rows


def _dyn_table(raw, texte, cols, count_key, coord_prefix, desig_prefix, default_n=1):
    """Construit un tableau a lignes de detail dynamiques (bouton "+ Ajouter
    une ligne"). raw = note3_manuel[sheet] (valeurs numeriques), texte =
    note_texte[sheet] (libelles/designations en texte libre). Chaque ligne i
    stocke ses valeurs sous coord_prefix+"_"+col+"_"+i et sa designation sous
    desig_prefix+"_"+i. Renvoie (rows, totals, n)."""
    n = int(ce._num(raw.get(count_key, default_n))) or default_n
    n = max(n, default_n)
    totals = {col: 0.0 for col, _label, _manual in cols}
    rows = []
    for i in range(n):
        vals = {}
        for col, _label, _manual in cols:
            coord = "%s_%s_%d" % (coord_prefix, col, i)
            v = raw.get(coord, "")
            vals[col] = v
            totals[col] += ce._num(v)
        designation = texte.get("%s_%d" % (desig_prefix, i), "")
        rows.append({"idx": i, "designation": designation, "vals": vals})
    return rows, totals, n


def compute_note32(manual, note_texte):
    """NOTE 32 — PRODUCTION DE L'EXERCICE. Feuille absente de sheets_raw.json
    (aucune formule extraite) : le classeur source ne reserve qu'une seule
    ligne de detail (ligne 11) + une ligne NON VENTILE (ligne 12) avant le
    TOTAL (ligne 13, =SUM dynamique INDIRECT/ADDRESS qui tolere l'insertion
    de lignes). On reproduit ce TOTAL en Python et on permet d'ajouter
    librement des lignes de detail supplementaires."""
    raw = (manual or {}).get("NOTE 32", {})
    texte = (note_texte or {}).get("NOTE 32", {})
    rows, totals, n = _dyn_table(raw, texte, NOTE32_COLS, "_N_DETAIL", "D", "desig", default_n=1)
    nv_vals = {}
    for col, _label, _manual in NOTE32_COLS:
        v = raw.get("NV_%s" % col, "")
        nv_vals[col] = v
        totals[col] += ce._num(v)
    return {"rows": rows, "n": n, "nv_vals": nv_vals, "totals": totals, "cols": NOTE32_COLS}


def compute_note33(manual, note_texte):
    """NOTE 33 — ACHATS DESTINES A LA PRODUCTION. Meme logique que NOTE 32 :
    ligne de detail (12) + NON VENTILE (13) + TOTAL (14) calcule en Python."""
    raw = (manual or {}).get("NOTE 33", {})
    texte = (note_texte or {}).get("NOTE 33", {})
    rows, totals, n = _dyn_table(raw, texte, NOTE33_COLS, "_N_DETAIL", "D", "desig", default_n=1)
    nv_vals = {}
    for col, _label, _manual in NOTE33_COLS:
        v = raw.get("NV_%s" % col, "")
        nv_vals[col] = v
        totals[col] += ce._num(v)
    return {"rows": rows, "n": n, "nv_vals": nv_vals, "totals": totals, "cols": NOTE33_COLS}


def compute_note16c(manual, note_texte):
    """NOTE 16C — ACTIFS ET PASSIFS EVENTUELS. Deux groupes independants
    ("Actif éventuel" / "Passif éventuel"), chacun avec sa propre ligne de
    litige(s) dynamique et son propre total (= SUM des litiges du groupe),
    calcule en Python (pas de formule extraite, classeur ne reserve qu'une
    ligne par groupe)."""
    raw = (manual or {}).get("NOTE 16C", {})
    texte = (note_texte or {}).get("NOTE 16C", {})
    actif_rows, actif_totals, n_actif = _dyn_table(raw, texte, NOTE16C_COLS, "_N_ACTIF", "ACTIF", "actif_desig", default_n=1)
    passif_rows, passif_totals, n_passif = _dyn_table(raw, texte, NOTE16C_COLS, "_N_PASSIF", "PASSIF", "passif_desig", default_n=1)
    return {
        "actif_rows": actif_rows, "actif_totals": actif_totals, "n_actif": n_actif,
        "passif_rows": passif_rows, "passif_totals": passif_totals, "n_passif": n_passif,
        "cols": NOTE16C_COLS,
    }


def _note30_group(raw, texte, fixed_lines, count_key, coord_prefix, desig_prefix,
                   total_row, total_label, sous_total_row, sous_total_label):
    val_cols = [("F", "Année N", True), ("G", "Année N-1", True)]
    detail_rows, detail_totals, n_detail = _dyn_table(
        raw, texte, val_cols, count_key, coord_prefix, desig_prefix, default_n=1)
    _inject_pct_column(detail_rows, "F", "G", "H")

    total = {"row": total_row, "label": total_label,
              "vals": {"F": detail_totals["F"], "G": detail_totals["G"]}}
    _inject_pct_column([total], "F", "G", "H")

    fixed_rows = []
    fixed_sum = {"F": 0.0, "G": 0.0}
    for row, label in fixed_lines:
        vals = {"F": raw.get("F%d" % row, ""), "G": raw.get("G%d" % row, "")}
        fixed_rows.append({"row": row, "label": label, "vals": vals})
        fixed_sum["F"] += ce._num(vals["F"])
        fixed_sum["G"] += ce._num(vals["G"])
    _inject_pct_column(fixed_rows, "F", "G", "H")

    sous_total = {"row": sous_total_row, "label": sous_total_label, "vals": {
        "F": total["vals"]["F"] + fixed_sum["F"], "G": total["vals"]["G"] + fixed_sum["G"],
    }}
    _inject_pct_column([sous_total], "F", "G", "H")

    return {"total": total, "detail_rows": detail_rows, "n_detail": n_detail,
            "fixed_rows": fixed_rows, "sous_total": sous_total}


def compute_note30(manual, note_texte):
    """NOTE 30 — AUTRES CHARGES ET PRODUITS HAO. Feuille absente de
    sheets_raw.json : dans le classeur source, la ligne "constatées (compte
    831)" [resp. "841"] est un simple renvoi (=SUM) vers l'unique ligne
    "Détail à préciser" (10/20), et le SOUS TOTAL (18/29) = ce renvoi +
    somme des rubriques nommées fixes (11-17 / 21-28). Aucune de ces lignes
    n'est liee a la balance (pas de SUMIFS). On reproduit ce calcul en
    Python et on permet d'ajouter librement des lignes de detail (la
    formule SUM dynamique du classeur, par INDIRECT/ADDRESS, tolere
    l'insertion de lignes) ; les rubriques nommees restent en saisie
    manuelle simple, et la colonne Variation en % est entierement
    calculee."""
    raw = (manual or {}).get("NOTE 30", {})
    texte = (note_texte or {}).get("NOTE 30", {})
    charges = _note30_group(
        raw, texte, NOTE30_CHARGES_FIXED_LINES, "_N_CHARGES", "CH", "desig_ch",
        9, "Charges HAO constatées (compte 831) à détailler",
        18, "SOUS TOTAL : AUTRES CHARGES HAO",
    )
    produits = _note30_group(
        raw, texte, NOTE30_PRODUITS_FIXED_LINES, "_N_PRODUITS", "PR", "desig_pr",
        19, "Produits HAO constatés (compte 841) à détailler",
        29, "SOUS TOTAL : AUTRES PRODUITS HAO",
    )
    return {"charges": charges, "produits": produits}


def compute_note27b(manual):
    """NOTE 27B — EFFECTIFS, MASSE SALARIALE ET PERSONNEL EXTERIEUR. Feuille
    absente de sheets_raw.json (aucune formule extraite) : on reproduit ici
    les vraies formules et la vraie disposition du classeur (verifiees
    cellule par cellule via openpyxl), qui n'est PAS symetrique entre les
    deux groupes :
    - Groupe 1 "personnel propre" (lignes 12-18, TOTAL (1) ligne 16) :
      colonnes E..J (effectif Nationaux/Autres Etats Région/Hors Région,
      M puis F), K = SUM(E:J), L/M (masse salariale Nationaux M/F),
      O/P/Q/R (masse salariale Autres Etats Région et Hors Région, M/F),
      S = L+M+O+P+Q+R. TOTAL (1) = SUM des lignes 12-15 sur toutes ces
      colonnes (E..S).
    - Groupe 2 "personnel extérieur" (lignes 22-28, TOTAL (2) ligne 26) :
      le classeur ne reserve QUE E..J (effectif) et L/M (masse salariale
      Nationaux M/F) à ce groupe ; K = SUM(E:J) ; il n'existe aucune
      colonne O/P/Q/R/S pour ce groupe (ces cellules sont occupees par le
      bloc annexe B/C decrit plus bas). TOTAL (2) = SUM des lignes 22-25
      sur E..J, K, L, M uniquement.
    - TOTAL (1+2) ligne 29 = ligne16 + ligne26, uniquement sur E..K (la
      masse salariale n'est pas cumulee sur cette ligne dans le classeur).
    - Bloc annexe (cellules O22/O24/O25, R22, R24, R25) : "B - Montants
      comptabilisés non rapportés dans l'état 301" (saisie manuelle, R22),
      "C - Montants des avantages en nature et en espèces comptabilisés
      en comptes de charges de personnel" (saisie manuelle, R24), "Total
      frais de personnel ligne RK (A+B+C)" = S16 + R22 + R24 (R25,
      calculé), où A = S16 = total masse salariale du personnel propre.
    """
    raw = (manual or {}).get("NOTE 27B", {})
    group1_cols = ["E", "F", "G", "H", "I", "J", "L", "M", "O", "P", "Q", "R"]
    group2_cols = ["E", "F", "G", "H", "I", "J", "L", "M"]
    sum_cols = ["E", "F", "G", "H", "I", "J"]
    masse_cols_g1 = ["L", "M", "O", "P", "Q", "R"]
    blank_cols_g2 = ["O", "P", "Q", "R", "S"]
    agg_cols_g1 = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "O", "P", "Q", "R", "S"]
    agg_cols_g2 = ["E", "F", "G", "H", "I", "J", "K", "L", "M"]

    def row_vals_g1(row):
        vals = {c: raw.get("%s%d" % (c, row), "") for c in group1_cols}
        vals["K"] = sum(ce._num(vals[c]) for c in sum_cols)
        vals["S"] = sum(ce._num(vals[c]) for c in masse_cols_g1)
        return vals

    def row_vals_g2(row):
        vals = {c: raw.get("%s%d" % (c, row), "") for c in group2_cols}
        vals["K"] = sum(ce._num(vals[c]) for c in sum_cols)
        for c in blank_cols_g2:
            vals[c] = ""
        return vals

    rows = []
    group1 = []
    for row, label in [
        (12, "1. Cadres supérieurs"),
        (13, "2. Techniciens supérieurs et cadres moyens"),
        (14, "3. Techniciens, agents de maîtrise et ouvriers qualifiés"),
        (15, "4. Employés, manœuvres, ouvriers et apprentis"),
    ]:
        r = {"row": row, "label": label, "total": False, "vals": row_vals_g1(row), "manual_cols": group1_cols}
        rows.append(r)
        group1.append(r)

    total1 = {c: sum(ce._num(r["vals"][c]) for r in group1) for c in agg_cols_g1}
    rows.append({"row": 16, "label": "TOTAL (1) — personnel propre (A)", "total": True, "vals": total1, "manual_cols": []})

    for row, label in [(17, "Permanents"), (18, "Saisonniers")]:
        rows.append({"row": row, "label": label, "total": False, "vals": row_vals_g1(row), "manual_cols": group1_cols})

    group2 = []
    for row, label in [
        (22, "Personnel extérieur — 1. Cadres supérieurs"),
        (23, "Personnel extérieur — 2. Techniciens supérieurs et cadres moyens"),
        (24, "Personnel extérieur — 3. Techniciens, agents de maîtrise et ouvriers qualifiés"),
        (25, "Personnel extérieur — 4. Employés, manœuvres, ouvriers et apprentis"),
    ]:
        r = {"row": row, "label": label, "total": False, "vals": row_vals_g2(row), "manual_cols": group2_cols}
        rows.append(r)
        group2.append(r)

    total2 = {c: sum(ce._num(r["vals"][c]) for r in group2) for c in agg_cols_g2}
    for c in blank_cols_g2:
        total2[c] = ""
    rows.append({"row": 26, "label": "TOTAL (2) — personnel extérieur", "total": True, "vals": total2, "manual_cols": []})

    for row, label in [(27, "Permanents"), (28, "Saisonniers")]:
        rows.append({"row": row, "label": label, "total": False, "vals": row_vals_g2(row), "manual_cols": group2_cols})

    total29 = {c: ce._num(total1[c]) + ce._num(total2[c]) for c in ["E", "F", "G", "H", "I", "J", "K"]}
    for c in ["L", "M", "O", "P", "Q", "R", "S"]:
        total29[c] = ""
    rows.append({"row": 29, "label": "TOTAL (1 + 2)", "total": True, "vals": total29, "manual_cols": []})

    note_rk_b = ce._num(raw.get("R22", ""))
    note_rk_c = ce._num(raw.get("R24", ""))
    side_block = {
        "b_label": "B - Montants comptabilisés non rapportés dans l'état 301",
        "b": note_rk_b,
        "c_label": "C - Montants des avantages en nature et en espèces comptabilisés en comptes de charges de personnel",
        "c": note_rk_c,
        "total_label": "Total frais de personnel ligne RK (A + B + C)",
        "total": ce._num(total1.get("S")) + note_rk_b + note_rk_c,
    }

    return rows, side_block


def compute_note13(manual, note_texte):
    """NOTE 13 — CAPITAL. Feuille absente de sheets_raw.json (aucune
    formule extraite). Structure réelle du classeur : L8 = valeur nominale
    unitaire des actions/parts (saisie unique, hors tableau) ; ligne 10 =
    détail des actionnaires/associés — une seule ligne réservée dans le
    modèle mais avec une formule SUM totale dynamique (INDIRECT/ADDRESS)
    qui tolère l'insertion de lignes, d'où le "+ Ajouter une ligne" ; pour
    chaque ligne de détail, K (Montant total) = L8 * J (Nombre), calculé
    et non saisi ; ligne 11 = "Apporteurs, capital non appelé", ligne fixe
    sans détail nominatif, K11 saisi directement ; ligne 12 = TOTAL,
    K12 = somme des K du détail (ligne 10..) + K11."""
    raw = (manual or {}).get("NOTE 13", {})
    texte = (note_texte or {}).get("NOTE 13", {})
    valeur_nominale = ce._num(raw.get("L8", ""))

    cols = [
        ("E", "N° compte contribuable", True),
        ("F", "Nationalité", True),
        ("G", "Autres nationalités à préciser", True),
        ("H", "Pays de résidence", True),
        ("I", "Nature des actions/parts", True),
        ("J", "Nombre", True),
        ("L", "Cessions/remb. en cours d'exercice", True),
    ]

    n = int(ce._num(raw.get("_N_ASSOCIES", 1))) or 1
    n = max(n, 1)
    detail_rows = []
    k_sum = 0.0
    for i in range(n):
        vals = {}
        for col, _label, _manual in cols:
            vals[col] = raw.get("ASSOC_%s_%d" % (col, i), "")
        vals["K"] = valeur_nominale * ce._num(vals["J"])
        k_sum += vals["K"]
        designation = texte.get("assoc_nom_%d" % i, "")
        detail_rows.append({"idx": i, "designation": designation, "vals": vals})

    apporteurs_k = raw.get("K11", "")
    total_k = k_sum + ce._num(apporteurs_k)

    return {
        "valeur_nominale": raw.get("L8", ""),
        "cols": cols,
        "detail_rows": detail_rows,
        "n": n,
        "apporteurs_k": apporteurs_k,
        "total_k": total_k,
    }


# ---------------------------------------------------------------- NOTE 1 ---
# DETTES GARANTIES PAR DES SURETES REELLES ET LES ENGAGEMENTS FINANCIERS.
# Feuille absente de sheets_raw.json (non extraite a l'origine, hors
# perimetre initial) : entierement a saisie manuelle, sans aucune formule
# sur les lignes de detail dans le classeur source — seules les SOUS
# TOTAL/TOTAL sont calcules (simples SUM). Reutilise le mecanisme generique
# note3_manuel (formulaire note3_sheet=NOTE 1), comme NOTE 31/34/Lot D.
NOTE1_T1_LINES = [
    (10, "Emprunts et dettes financières diverses :", 16, "header"),
    (11, "Emprunts obligataires convertibles", None, "manual"),
    (12, "Autres emprunts obligataires", None, "manual"),
    (13, "Emprunts et dettes des établissements de crédit", None, "manual"),
    (14, "Autres dettes financières", None, "manual"),
    (15, "SOUS TOTAL (1)", None, "total"),
    (16, "Dettes de location-acquisition :", 16, "header"),
    (17, "Dettes de crédit-bail immobilier", None, "manual"),
    (18, "Dettes de crédit-bail mobilier", None, "manual"),
    (19, "Dettes sur contrats de location-vente", None, "manual"),
    (20, "Autres dettes sur contrats de location-acquisition", None, "manual"),
    (21, "SOUS TOTAL (2)", None, "total"),
    (22, "Dettes du passif circulant :", None, "header"),
    (23, "Fournisseurs et comptes rattachés", 17, "manual"),
    (24, "Clients", 7, "manual"),
    (25, "Personnel", 18, "manual"),
    (26, "Sécurité sociale et organismes sociaux", 18, "manual"),
    (27, "Etat", 18, "manual"),
    (28, "Organismes internationaux", 19, "manual"),
    (29, "Associés et groupe", 19, "manual"),
    (30, "Créditeurs divers", 19, "manual"),
    (31, "SOUS TOTAL (3)", None, "total"),
    (32, "TOTAL (1) + (2) + (3)", None, "total"),
]
NOTE1_T1_TOTAL_RANGES = {15: (11, 14), 21: (17, 20), 31: (23, 30)}

NOTE1_T2_COLS = [("G", "Engagements réciproques"), ("H", "Engagements donnés"), ("I", "Engagements reçus")]
NOTE1_T2_LINES = [
    (34, "Engagements consentis à des entités liées"),
    (35, "Primes de remboursement des obligations non échues"),
    (36, "Avals, cautions, garanties"),
    (37, "Hypothèques, nantissements, gages, autres"),
    (38, "Effets escomptés non échus"),
    (39, "Créances commerciales et professionnelles cédées"),
    (40, "Abandons de créances conditionnels"),
    (41, "Achats de marchandises à terme"),
    (42, "Achats à terme de devises non couverts"),
    (43, "Commandes fermes des clients"),
    (44, "Autres engagements réciproques"),
]


def compute_note1(manual):
    raw = (manual or {}).get("NOTE 1", {})

    def num(coord):
        val = raw.get(coord)
        if val in (None, ""):
            return 0.0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    rows1 = []
    for row, label, note_ref, kind in NOTE1_T1_LINES:
        if kind == "manual":
            rows1.append({"row": row, "label": label, "note_ref": note_ref, "kind": kind,
                          "F": raw.get("F%d" % row, "")})
        elif kind == "header":
            rows1.append({"row": row, "label": label, "note_ref": note_ref, "kind": kind, "F": None})
        else:
            if row == 32:
                total = num("F15") + num("F21") + num("F31")
            else:
                lo, hi = NOTE1_T1_TOTAL_RANGES[row]
                total = sum(num("F%d" % r) for r in range(lo, hi + 1))
            rows1.append({"row": row, "label": label, "note_ref": note_ref, "kind": kind, "F": total})

    rows2 = []
    totals2 = {"G": 0.0, "H": 0.0, "I": 0.0}
    for row, label in NOTE1_T2_LINES:
        vals = {}
        for col, _label in NOTE1_T2_COLS:
            vals[col] = raw.get("%s%d" % (col, row), "")
            totals2[col] += num("%s%d" % (col, row))
        rows2.append({"row": row, "label": label, "vals": vals})

    return {"table1": rows1, "table2": rows2, "table2_cols": NOTE1_T2_COLS, "totals2": totals2}


# --------------------------------------------------------------- NOTE 31 ---
# REPARTITION DU RESULTAT ET AUTRES ELEMENTS CARACTERISTIQUES DES CINQ
# DERNIERS EXERCICES : colonnes G=N, H=N-1, I=N-2, J=N-3, K=N-4. Seules G/H
# ont des formules dans le classeur d'origine (l'application n'importe que
# les balances N et N-1) ; I/J/K sont donc toujours a saisie manuelle.
# Certaines lignes (12-16, 24-25, 27-28) sont entierement a saisie manuelle,
# meme sur G/H, dans le classeur d'origine (aucune formule, y compris pour
# l'exercice N). Plutot que de coder cette distinction en dur, le caractere
# manuel de chaque cellule est determine dynamiquement en verifiant la
# presence d'une formule dans sheets_raw.json — exactement le critere que le
# moteur de calcul utilise lui-meme pour decider si une cellule est calculee
# ou remplacee par une saisie manuelle (voir compute_workbook/GG).
NOTE31_COLS = ["G", "H", "I", "J", "K"]
NOTE31_COL_LABELS = {"G": "N", "H": "N-1", "I": "N-2", "J": "N-3", "K": "N-4"}
NOTE31_LINES = [
    (10, "STRUCTURE DU CAPITAL A LA CLOTURE DE L'EXERCICE (2)", True),
    (11, "Capital social", False),
    (12, "Actions ordinaires", False),
    (13, "Actions à dividendes prioritaires (A.D.P) sans droit de vote", False),
    (14, "Actions nouvelles à émettre :", False),
    (15, "- par conversion d'obligations", False),
    (16, "- par exercice de droits de souscription", False),
    (17, "OPERATIONS ET RESULTATS DE L'EXERCICE (3)", True),
    (18, "Chiffre d'affaires hors taxes", False),
    (19, "Résultat des activités ordinaires (R.A.O) hors dotations et reprises (exploitation et financières)", False),
    (20, "Participation des travailleurs aux bénéfices", False),
    (21, "Impôt sur le résultat", False),
    (22, "Résultat net (4)", False),
    (23, "RESULTAT ET DIVIDENDE DISTRIBUES", True),
    (24, "Résultat distribué (5)", False),
    (25, "Dividende attribué à chaque action", False),
    (26, "PERSONNEL ET POLITIQUE SALARIALE", True),
    (27, "Effectif moyen des travailleurs au cours de l'exercice (6)", False),
    (28, "Effectif moyen de personnel extérieur", False),
    (29, "Masse salariale distribuée au cours de l'exercice (7)", False),
    (30, "Avantages sociaux versés au cours de l'exercice (8) [Sécurité sociale, œuvres sociales]", False),
    (31, "Personnel extérieur facturé à l'entité (9)", False),
]


def build_note31_table(balN, balN1, manual, resultat):
    """NOTE 31 — feuille absente de sheets_raw.json (aucune formule
    extraite, comme NOTE 27B/NOTE 13) : on reproduit ici directement en
    Python les formules du classeur (SUMIFS sur la balance + references
    croisees RESULTAT!I44/J44 et RESULTAT!I52/J52) pour les colonnes G (N)
    et H (N-1), seules calculables a partir de la balance. I/J/K (N-2 a
    N-4) ainsi que les lignes 12-16/24-25/27-28 restent a saisie manuelle,
    comme dans le classeur source. Les lignes TOTAL (10, 17, 23, 26)
    sont sommees sur les 5 colonnes (G a K), manuel compris."""
    raw_manual = (manual or {}).get("NOTE 31", {})
    SUMIFS_ = ce.make_sumifs(balN, balN1 or [])

    def sf(table, field, lo, hi):
        return SUMIFS_(table, field, ">=", lo, "<", hi)

    def ca_ht(table):
        total = 0.0
        for lo, hi in [(70100, 70200), (70200, 70500), (70500, 70700), (70700, 70800)]:
            total += sf(table, "BS_Credit", lo, hi) - sf(table, "BS_Debit", lo, hi)
        return total

    def rao(table, resultat_key):
        adj = 0.0
        for lo, hi in [(79100, 79200), (79800, 80000), (68100, 68200),
                       (69100, 69200), (79700, 79800), (69700, 69800)]:
            adj += sf(table, "BS_Credit", lo, hi) - sf(table, "BS_Debit", lo, hi)
        return ce._num((resultat or {}).get(resultat_key)) - adj

    computed = {
        "G11": sf("BalanceN", "BS_Credit", 10100, 10200),
        "H11": sf("BalanceN1", "BS_Credit", 10100, 10200),
        "G18": ca_ht("BalanceN"),
        "H18": ca_ht("BalanceN1"),
        "G19": rao("BalanceN", "I44"),
        "H19": rao("BalanceN1", "J44"),
        "G20": -sf("BalanceN", "BS_Credit", 87000, 88000) + sf("BalanceN", "BS_Debit", 87000, 88000),
        "H20": -sf("BalanceN1", "BS_Credit", 87000, 88000) + sf("BalanceN1", "BS_Debit", 87000, 88000),
        "G21": -sf("BalanceN", "BS_Credit", 89000, 90000) + sf("BalanceN", "BS_Debit", 89000, 90000),
        "H21": -sf("BalanceN1", "BS_Credit", 89000, 90000) + sf("BalanceN1", "BS_Debit", 89000, 90000),
        "G22": ce._num((resultat or {}).get("I52")),
        "H22": ce._num((resultat or {}).get("J52")),
        "G29": sf("BalanceN", "BS_Debit", 66100, 66400),
        "H29": sf("BalanceN1", "BS_Debit", 66100, 66400),
        "G30": sf("BalanceN", "BS_Debit", 66400, 66500) + sf("BalanceN", "BS_Debit", 66800, 66900),
        "H30": sf("BalanceN1", "BS_Debit", 66400, 66500) + sf("BalanceN1", "BS_Debit", 66800, 66900),
        "G31": sf("BalanceN", "BS_Debit", 66700, 66800),
        "H31": sf("BalanceN1", "BS_Debit", 66700, 66800),
    }

    def cell_num(col, row):
        coord = "%s%d" % (col, row)
        if coord in computed:
            return ce._num(computed[coord])
        return ce._num(raw_manual.get(coord))

    for col in NOTE31_COLS:
        computed["%s10" % col] = sum(cell_num(col, r) for r in range(11, 17))
        computed["%s17" % col] = sum(cell_num(col, r) for r in range(18, 23))
        computed["%s23" % col] = sum(cell_num(col, r) for r in range(24, 26))
        computed["%s26" % col] = sum(cell_num(col, r) for r in range(27, 32))

    rows = []
    for row, label, is_total in NOTE31_LINES:
        vals = {}
        manual_cols = []
        for col in NOTE31_COLS:
            coord = "%s%d" % (col, row)
            if coord in computed:
                vals[col] = computed[coord]
            else:
                v = raw_manual.get(coord, "")
                vals[col] = v
                manual_cols.append(col)
        rows.append({"row": row, "label": label, "total": is_total, "vals": vals, "manual_cols": manual_cols})
    return rows


def compute_note31(balN, balN1, manual, resultat):
    """NOTE 31 reutilise le meme mecanisme de saisie manuelle (table
    note3_manuel, formulaire note3_sheet=NOTE 31) que toutes les autres
    notes ; seule la determination de quelles cellules sont manuelles
    differe (par cellule plutot que par colonne entiere, voir ci-dessus)."""
    if not balN:
        return []
    return build_note31_table(balN, balN1, manual, resultat)


# --------------------------------------------------------------- NOTE 34 ---
# FICHE DE SYNTHESE DES PRINCIPAUX INDICATEURS FINANCIERS : feuille de
# references croisees pure (BILAN/RESULTAT/TFT), sans aucune cellule a
# saisie manuelle dans le classeur d'origine — la seule note de ce type
# parmi toutes les notes annexes. Plutot que de transpiler ses formules
# (references croisees + formules matricielles vers la feuille "Table TFT"
# via des plages nommees), on les reimplemente directement en Python a
# partir des caches BILAN/RESULTAT/TFT deja calcules par l'application (la
# meme demarche que celle utilisee pour NOTE 3D/NOTE 21 — formules
# dynamiques traitees specialement) : plus simple et plus sur que d'etendre
# le moteur de calcul generique pour un cas unique (references nommees vers
# une feuille de detail TFTN/TFTN1, formules matricielles).
NOTE34_LINES = [
    (11, "Chiffre d'affaires", False, "pct"),
    (12, "Marge commerciale", False, "pct"),
    (13, "Valeur ajoutée", False, "pct"),
    (14, "Excédent brut d'exploitation (EBE)", False, "pct"),
    (15, "Résultat d'exploitation", False, "pct"),
    (16, "Résultat financier", False, "pct"),
    (17, "Résultat des activités ordinaires", False, "pct"),
    (18, "Résultat hors activités ordinaires", False, "pct"),
    (19, "Résultat net", False, "pct"),
    (21, "Excédent brut d'exploitation (EBE)", False, None),
    (22, "+ Valeurs comptables des cessions courantes d'immobilisation (compte 654)", False, None),
    (23, "- Produits des cessions courantes d'immobilisation (compte 754)", False, None),
    (24, "= CAPACITE D'AUTOFINANCEMENT D'EXPLOITATION", True, "pct"),
    (25, "+ Revenus financiers", False, None),
    (26, "+ Gains de change financiers", False, None),
    (27, "+ Transferts de charges financières", False, None),
    (28, "+ Produits HAO", False, None),
    (29, "+ Transferts de charges HAO", False, None),
    (30, "- Frais financiers", False, None),
    (31, "- Pertes de change financières", False, None),
    (32, "- Charges HAO", False, None),
    (33, "- Participations", False, None),
    (34, "- Impôts sur les résultats", False, None),
    (35, "= CAPACITE D'AUTOFINANCEMENT GLOBALE (C.A.F.G.)", True, "pct"),
    (36, "- Distributions de dividendes opérées durant l'exercice", False, "pct"),
    (37, "= AUTOFINANCEMENT", True, "pct"),
    (39, "Rentabilité économique = Résultat d'exploitation (a)/Capitaux propres + dettes financières", False, "points"),
    (40, "Rentabilité financière = Résultat net/Capitaux propres", False, "points"),
    (42, "+ Capitaux propres et ressources assimilées", False, "pct"),
    (43, "+ Dettes financières et autres ressources assimilées (b)", False, "pct"),
    (44, "= ressources stables", True, "pct"),
    (45, "- Actif immobilisé (b)", False, "pct"),
    (46, "= FONDS DE ROULEMENT (1)", True, "pct"),
    (47, "+ Actif circulant d'exploitation (b)", False, "pct"),
    (48, "- Passif circulant d'exploitation (b)", False, "pct"),
    (49, "= BESOIN DE FINANCEMENT D'EXPLOITATION (2)", True, "pct"),
    (50, "+ Actif circulant HAO (b)", False, "pct"),
    (51, "- Passif circulant HAO (b)", False, "pct"),
    (52, "= BESOIN DE FINANCEMENT HAO (3)", True, "pct"),
    (53, "BESOIN DE FINANCEMENT GLOBAL (4) = (2) + (3)", True, "pct"),
    (54, "TRESORERIE NETTE (5) = (1) - (4)", True, "pct"),
    (55, "CONTRÔLE : TRESORERIE NETTE = (TRESORERIE - ACTIF) - (TRESORERIE - PASSIF)", True, "pct"),
    (57, "+ Flux de trésorerie des activités opérationnelles", False, "pct"),
    (58, "- Flux de trésorerie des activités d'investissement", False, "pct"),
    (59, "+ Flux de trésorerie des activités de financement", False, "pct"),
    (60, "= VARIATION de LA TRESORERIE NETTE de LA PERIODE", True, "pct"),
    (62, "Endettement financier brut (Dettes financières* + Trésorerie - passif)", False, "pct"),
    (63, "- Trésorerie - actif", False, "pct"),
    (64, "= ENDETTEMENT FINANCIER NET", True, "pct"),
]


def _note34_pct(f, g):
    if g:
        return (f - g) / g
    if f:
        return 1.0 if f > 0 else -1.0
    return ""


def _note34_points(f, g):
    fv = f if isinstance(f, (int, float)) else 0
    gv = g if isinstance(g, (int, float)) else 0
    return "%d points" % round(100 * (fv - gv))


def _note34_iferror_div(num, den):
    try:
        if not den:
            return ""
        return num / den
    except Exception:
        return ""


def compute_note34(bilan, resultat, tft, balN, balN1, tftn, tftn1):
    if not balN:
        return []

    def b(coord):
        return ce._num(bilan.get(coord))

    def r(coord):
        return ce._num(resultat.get(coord))

    def t(coord):
        return ce._num(tft.get(coord))

    def s(rows, lo, hi, field):
        total = 0.0
        for row in rows or []:
            if lo <= row.get("table", 0) < hi:
                total += row.get(field, 0) or 0
        return total

    def sn(lo, hi, field):
        return s(balN, lo, hi, field)

    def sn1(lo, hi, field):
        return s(balN1, lo, hi, field)

    named = ce.tft_named_split(tftn, tftn1)
    out = {}

    for row, cf, cg, div in [
        (11, "I18", "J18", 1000), (12, "I14", "J14", 1000), (13, "I32", "J32", 1000),
        (14, "I34", "J34", 1000), (15, "I37", "J37", 1000), (16, "I43", "J43", 1000),
        (17, "I44", "J44", 1000), (18, "I49", "J49", 1000), (19, "I52", "J52", 1000),
    ]:
        out["F%d" % row] = r(cf) / div
        out["G%d" % row] = r(cg) / div

    out["F21"], out["G21"] = out["F14"], out["G14"]
    out["F22"] = sn(65400, 65500, "BS_Debit")
    out["G22"] = sn1(65400, 65500, "BS_Debit")
    out["F23"] = -sn(75400, 75500, "BS_Credit")
    out["G23"] = -sn1(75400, 75500, "BS_Credit")
    out["F24"] = out["F21"] + out["F22"] - out["F23"]
    out["G24"] = out["G21"] + out["G22"] - out["G23"]

    out["F25"] = (sn(77000, 78000, "BS_Credit") - sn(77000, 78000, "BS_Debit")
                  - (sn(77600, 77700, "BS_Credit") - sn(77600, 77800, "BS_Debit"))) / 1000
    out["G25"] = (sn1(77000, 78000, "BS_Credit") - sn1(77000, 78000, "BS_Debit")
                  - (sn1(77600, 77700, "BS_Credit") - sn1(77600, 77800, "BS_Debit"))) / 1000
    out["F26"] = (sn(77600, 77700, "BS_Credit") - sn(77600, 77800, "BS_Debit")) / 1000
    out["G26"] = (sn1(77600, 77700, "BS_Credit") - sn1(77600, 77800, "BS_Debit")) / 1000
    out["F27"] = (sn(78700, 78800, "BS_Credit") - sn(78700, 78800, "BS_Debit")) / 1000
    out["G27"] = (sn1(78700, 78800, "BS_Credit") - sn1(78700, 78800, "BS_Debit")) / 1000
    out["F28"] = (sn(84000, 85000, "BS_Credit") - sn(84000, 85000, "BS_Debit")
                  - (sn(84800, 84900, "BS_Credit") - sn(84800, 84900, "BS_Debit"))
                  + sn(88000, 89000, "BS_Credit") - sn(88000, 89000, "BS_Debit")) / 1000
    out["G28"] = (sn1(84000, 85000, "BS_Credit") - sn1(84000, 85000, "BS_Debit")
                  - (sn1(84800, 84900, "BS_Credit") - sn1(84800, 84900, "BS_Debit"))
                  + sn1(88000, 89000, "BS_Credit") - sn1(88000, 89000, "BS_Debit")) / 1000
    out["F29"] = sn(84800, 84900, "BS_Credit") - sn(84800, 84900, "BS_Debit")
    out["G29"] = sn1(84800, 84900, "BS_Credit") - sn1(84800, 84900, "BS_Debit")
    out["F30"] = (sn(67000, 68000, "BS_Credit") - sn(67000, 68000, "BS_Debit")
                  - (sn(67600, 67700, "BS_Credit") - sn(67600, 67700, "BS_Debit"))) / 1000
    out["G30"] = (sn1(67000, 68000, "BS_Credit") - sn1(67000, 68000, "BS_Debit")
                  - (sn1(67600, 67700, "BS_Credit") - sn1(67600, 67700, "BS_Debit"))) / 1000
    out["F31"] = (sn(67600, 67700, "BS_Credit") - sn(67600, 67700, "BS_Debit")) / 1000
    out["G31"] = (sn1(67600, 67700, "BS_Credit") - sn1(67600, 67700, "BS_Debit")) / 1000
    out["F32"] = (sn(83000, 84000, "BS_Debit") - sn(83000, 84000, "BS_Credit")) / 1000
    out["G32"] = (sn1(83000, 84000, "BS_Debit") - sn1(83000, 84000, "BS_Credit")) / 1000
    out["F33"] = (sn(87000, 88000, "BS_Credit") - sn(87000, 88000, "BS_Debit")) / 1000
    out["G33"] = (sn1(87000, 88000, "BS_Credit") - sn1(87000, 88000, "BS_Debit")) / 1000
    out["F34"] = (sn(89000, 90000, "BS_Credit") - sn(89000, 90000, "BS_Debit")) / 1000
    out["G34"] = (sn1(89000, 90000, "BS_Credit") - sn1(89000, 90000, "BS_Debit")) / 1000

    out["F35"] = sum(out["F%d" % i] for i in range(24, 35))
    out["G35"] = sum(out["G%d" % i] for i in range(24, 35))

    out["F36"] = t("I31") / 1000
    out["G36"] = t("J31") / 1000
    out["F37"] = out["F35"] + out["F36"]
    out["G37"] = out["G35"] + out["G36"]

    TAUX_IS = 0.25  # Sommaire!M43 dans le classeur d'origine (valeur fixe, non éditable)
    out["F39"] = _note34_iferror_div(r("I37") * (1 - TAUX_IS), b("M21") + b("M25"))
    out["G39"] = _note34_iferror_div(r("J37") * (1 - TAUX_IS), b("N21") + b("N25"))
    out["F40"] = _note34_iferror_div(out["F19"], b("M21"))
    out["G40"] = _note34_iferror_div(out["G19"], b("N21"))

    out["F42"] = b("M21") / 1000
    out["G42"] = b("N21") / 1000
    out["F43"] = (b("M25") + sn(47940, 47950, "BS_Credit") - sn(47840, 47850, "BS_Debit")) / 1000
    out["G43"] = (b("N25") + sn1(47940, 47950, "BS_Credit") - sn1(47840, 47850, "BS_Debit")) / 1000
    out["F44"] = out["F42"] + out["F43"]
    out["G44"] = out["G42"] + out["G43"]
    out["F45"] = -b("H26") / 1000
    out["G45"] = -b("I26") / 1000
    out["F46"] = out["F44"] + out["F45"]
    out["G46"] = out["G44"] + out["G45"]

    out["F47"] = (b("H28") + b("H29") + named["ECEN"]) / 1000
    out["G47"] = (b("I28") + b("I29") + named["ECN"]) / 1000
    out["F48"] = (-b("M28") - b("M29") - b("M30") - b("M31") - b("M32") - named["EPN"]) / 1000
    out["G48"] = (-b("N28") - b("N29") - b("N30") - b("N31") - b("N32") - named["EPA"]) / 1000
    out["F49"] = out["F47"] + out["F48"]
    out["G49"] = out["G47"] + out["G48"]
    out["F50"] = (b("H27") + named["ECHN"]) / 1000
    out["G50"] = (b("I27") + named["ECH"]) / 1000
    out["F51"] = (-b("M27") - named["EPH"]) / 1000
    out["G51"] = (-b("N27") - named["EPB"]) / 1000
    out["F52"] = out["F50"] + out["F51"]
    out["G52"] = out["G50"] + out["G51"]
    out["F53"] = out["F49"] + out["F52"]
    out["G53"] = out["G49"] + out["G52"]
    out["F54"] = out["F46"] - out["F53"]
    out["G54"] = out["G46"] - out["G53"]
    out["F55"] = (b("H37") - b("M37")) / 1000
    out["G55"] = (b("I37") - b("N37")) / 1000

    out["F57"] = t("I19") / 1000
    out["G57"] = t("J19") / 1000
    out["F58"] = t("I26") / 1000
    out["G58"] = t("J26") / 1000
    out["F59"] = t("I38") / 1000
    out["G59"] = t("J38") / 1000
    out["F60"] = out["F57"] + out["F58"] + out["F59"]
    out["G60"] = out["G57"] + out["G58"] + out["G59"]

    out["F62"] = (b("M22") + b("M23") + b("M37")) / 1000
    out["G62"] = (b("N22") + b("N23") + b("N37")) / 1000
    out["F63"] = -b("H37") / 1000
    out["G63"] = -b("I37") / 1000
    out["F64"] = out["F62"] + out["F63"]
    out["G64"] = out["G62"] + out["G63"]

    for row, _label, _total, kind in NOTE34_LINES:
        f, g = out.get("F%d" % row), out.get("G%d" % row)
        if kind == "pct":
            out["H%d" % row] = _note34_pct(f, g)
        elif kind == "points":
            out["H%d" % row] = _note34_points(f, g)

    return build_note34_rows(out)


def build_note34_rows(out):
    rows = []
    for row, label, is_total, kind in NOTE34_LINES:
        rows.append({
            "row": row, "label": label, "total": is_total, "kind": kind,
            "F": out.get("F%d" % row), "G": out.get("G%d" % row), "H": out.get("H%d" % row),
        })
    return rows


def compute_note3(balN, manual):
    if not balN:
        return {}, {}, {}, {}, {}
    caches = ce.compute_workbook(
        ["NOTE 3A", "NOTE 3C", "NOTE 3C BIS", "NOTE 3D", "NOTE 3B"],
        balN, balN, manual=manual,
    )
    note3a = build_note3_table(caches.get("NOTE 3A", {}), NOTE3A_LINES, NOTE3A_COLS)
    note3c = build_note3_table(caches.get("NOTE 3C", {}), NOTE3C_LINES, NOTE3C_COLS)
    note3cbis = build_note3_table(caches.get("NOTE 3C BIS", {}), NOTE3CBIS_LINES, NOTE3CBIS_COLS)
    note3d = build_note3_table(caches.get("NOTE 3D", {}), NOTE3D_LINES, NOTE3D_COLS)
    note3b = build_note3_table(caches.get("NOTE 3B", {}), NOTE3B_LINES, NOTE3B_COLS)
    return note3a, note3c, note3cbis, note3d, note3b


NOTE3E_COLS = [
    ("F", "Montants en coûts historiques"),
    ("G", "Montant réévalués"),
    ("H", "Ecarts de réévaluation"),
    ("I", "Amortissements supplémentaires"),
]


def compute_note3e(manual, note_texte):
    """NOTE 3E — INFORMATIONS SUR LES REEVALUATIONS EFFECTUEES PAR L'ENTITE.

    Dans le classeur Excel : ligne 11 = UNE ligne de detail masquee
    ("Eléments réévalués par poste du bilan"), ligne 12 = TOTAL GENERAL,
    self-referentiel via INDIRECT/ADDRESS (meme principe que NOTE 21/30/32/33).
    Ici, transformee en bloc a lignes dynamiques ("+ Ajouter une ligne"), une
    ligne par poste du bilan reevalue. S'ajoutent les champs hors-tableau du
    classeur : 3 zones de texte libre (nature et date des reevaluations,
    methode de reevaluation utilisee, traitement fiscal de l'ecart) et un
    montant manuel (ecart de reevaluation incorpore au capital, A17/H17)."""
    raw = (manual or {}).get("NOTE 3E", {})
    texte = (note_texte or {}).get("NOTE 3E", {})
    n = int(ce._num(raw.get("_N_REEVAL", 1))) or 1
    n = max(n, 1)
    cols = NOTE3E_COLS
    rows = []
    totals = {col: 0.0 for col, _label in cols}
    for i in range(n):
        vals = {col: raw.get("REEVAL_%s_%d" % (col, i), "") for col, _label in cols}
        for col, _label in cols:
            totals[col] += ce._num(vals[col])
        rows.append({
            "idx": i,
            "designation": texte.get("reeval_poste_%d" % i, ""),
            "vals": vals,
        })
    return {
        "cols": cols, "rows": rows, "totals": totals, "n": n,
        "nature_date": texte.get("nature_date", ""),
        "methode": texte.get("methode", ""),
        "traitement_fiscal": texte.get("traitement_fiscal", ""),
        "montant_capital": raw.get("MONTANT_CAPITAL", ""),
    }


def compute_notes_lot_a(balN, balN1, manual):
    """NOTE 4 a 11 : utilisent BalanceN ET BalanceN1 (colonne Année N-1), a
    la difference de la famille NOTE 3 (mouvements N uniquement). NOTE 5
    comporte deux tableaux distincts dans le classeur source (ACTIF
    CIRCULANT HAO et DETTES CIRCULANTES HAO) : on les regroupe sous une
    seule carte par feuille, comme pour le Lot D (voir compute_notes_lot_d)."""
    sheets = list(dict.fromkeys(name for name, _, _, _ in NOTES_LOT_A))
    caches = ce.compute_workbook(sheets, balN, balN1, manual=manual, row_range=(1, 40)) if balN else {}
    groups = {}
    order = []
    for name, lines, cols, title in NOTES_LOT_A:
        if name not in groups:
            groups[name] = []
            order.append(name)
        groups[name].append({
            "title": title,
            "cols": cols,
            "rows": build_note3_table(caches.get(name, {}), lines, cols) if balN else [],
        })
    return [{"sheet": s, "tables": groups[s]} for s in order]


def _inject_pct_column(rows, n_col, n1_col, pct_col):
    """Ajoute/recalcule une colonne de variation en % a partir de deux
    colonnes deja presentes (memes formules IFERROR que le classeur :
    =(F-G)/G, ou +-1 si G=0 et F<>0). Utilise pour les feuilles ou cette
    colonne n'existe pas dans le classeur source (ex. NOTE 14)."""
    for row in rows:
        f, g = row["vals"].get(n_col), row["vals"].get(n1_col)
        try:
            f = float(f) if f not in (None, "") else 0.0
        except (TypeError, ValueError):
            f = 0.0
        try:
            g = float(g) if g not in (None, "") else 0.0
        except (TypeError, ValueError):
            g = 0.0
        if g:
            row["vals"][pct_col] = (f - g) / g
        elif f:
            row["vals"][pct_col] = 1.0 if f > 0 else -1.0
        else:
            row["vals"][pct_col] = ""
    return rows


def compute_notes_lot_b(balN, balN1, manual):
    """NOTE 14, 15A, 16A, 17-20 : capitaux propres / dettes financieres et
    circulantes. Comme NOTE 4-11, utilisent BalanceN ET BalanceN1."""
    if not balN:
        return {name: [] for name, _, _, _ in NOTES_LOT_B}
    sheets = [name for name, _, _, _ in NOTES_LOT_B]
    caches = ce.compute_workbook(sheets, balN, balN1, manual=manual, row_range=(1, 45))
    out = {}
    for name, lines, cols, _title in NOTES_LOT_B:
        out[name] = build_note3_table(caches.get(name, {}), lines, cols)
    _inject_pct_column(out["NOTE 14"], "F", "G", "I")
    return out


def compute_notes_lot_c(balN, balN1, manual):
    """NOTE 22-26, 27A, 28, 29 : achats, charges par nature, charges de
    personnel, provisions, charges et revenus financiers. Comme Lot A/B,
    utilisent BalanceN ET BalanceN1. (NOTE 21 a sa propre fonction, voir
    compute_note21, en raison de sa ligne 27 a detail dynamique.)"""
    if not balN:
        return {name: [] for name, _, _, _ in NOTES_LOT_C}
    sheets = [name for name, _, _, _ in NOTES_LOT_C]
    caches = ce.compute_workbook(sheets, balN, balN1, manual=manual, row_range=(1, 50))
    out = {}
    for name, lines, cols, _title in NOTES_LOT_C:
        out[name] = build_note3_table(caches.get(name, {}), lines, cols)
    return out


def compute_note21(balN, balN1, manual, note_texte):
    """NOTE 21 — CHIFFRE D'AFFAIRES ET AUTRES PRODUITS.

    Cellules en jaune dans le classeur Excel (saisie manuelle car non
    deductibles de la balance) : F10/G10, F16/G16, F22/G22 (ventes hors
    Etat partie/Region par categorie) et F30/G30, F31/G31, F32/G32
    (Production immobilisée, Subventions d'exploitation, Autres produits —
    aucune formule dans le classeur source).

    Ligne 27 (Produits accessoires à détailler par nature d'activité
    économique) : dans Excel, F27/G27 = SUM(...) sur UNE ligne de detail
    masquee (28) via INDIRECT/ADDRESS/SUBSTITUTE — meme principe self-
    referentiel que NOTE 3E/30/32/33. Ici, transformee en bloc a lignes
    dynamiques ("+ Ajouter une ligne") ; le total (F27/G27) est reinjecte
    comme saisie manuelle synthetique avant le calcul du reste du tableau
    (F29 = F14+F20+F26+F27 etc. suit alors la formule Excel normalement)."""
    raw = (manual or {}).get("NOTE 21", {}) if manual else {}
    texte = (note_texte or {}).get("NOTE 21", {}) if note_texte else {}

    n = int(ce._num(raw.get("_N_PA", 1))) or 1
    n = max(n, 1)
    pa_rows = []
    f_sum = 0.0
    g_sum = 0.0
    for i in range(n):
        f_val = raw.get("PA_F_%d" % i, "")
        g_val = raw.get("PA_G_%d" % i, "")
        f_sum += ce._num(f_val)
        g_sum += ce._num(g_val)
        pa_rows.append({
            "idx": i,
            "designation": texte.get("pa_designation_%d" % i, ""),
            "F": f_val, "G": g_val,
        })

    if not balN:
        return {
            "rows_a": [], "rows_b": [], "cols": NOTE21_COLS,
            "pa_rows": pa_rows, "pa_n": n, "pa_f": f_sum, "pa_g": g_sum,
            "pa_h": "",
        }

    manual_aug = dict(raw)
    manual_aug["F27"] = f_sum
    manual_aug["G27"] = g_sum
    manual_full = dict(manual or {})
    manual_full["NOTE 21"] = manual_aug

    caches = ce.compute_workbook(["NOTE 21"], balN, balN1, manual=manual_full, row_range=(1, 40))
    c21 = caches.get("NOTE 21", {})
    rows_a = build_note3_table(c21, NOTE21_LINES_A, NOTE21_COLS)
    rows_b = build_note3_table(c21, NOTE21_LINES_B, NOTE21_COLS)
    pa_h = c21.get("H27", "")
    return {
        "rows_a": rows_a, "rows_b": rows_b, "cols": NOTE21_COLS,
        "pa_rows": pa_rows, "pa_n": n, "pa_f": f_sum, "pa_g": g_sum,
        "pa_h": pa_h,
    }


def extract_lines(cache, sheet, line_defs):
    """line_defs: list of (label, row, col_N, col_N1)"""
    out = []
    for label, row, col_n, col_n1 in line_defs:
        out.append({
            "label": label,
            "n": cache.get("%s%d" % (col_n, row)),
            "n1": cache.get("%s%d" % (col_n1, row)) if col_n1 else None,
        })
    return out


def _is_total_label(label):
    """Heuristique reprenant la mise en forme du classeur Excel : les
    libellés des lignes de sous-total/total (rubriques de regroupement et
    soldes intermédiaires de gestion XA..XI) sont intégralement en
    MAJUSCULES dans le classeur source ; les lignes de détail sont en
    casse normale. Permet de reproduire le style gras des totaux sans
    dupliquer une liste séparée."""
    return label.strip().isupper()


# BILAN — ACTIF : reproduit fidèlement la disposition de la feuille BILAN
# (colonnes REF=A, ACTIF=B, NOTE=E, BRUT=F, AMORT. ET DEPREC.=G, NET N=H,
# NET N-1=I), lignes 11 à 39, vérifiées cellule par cellule via openpyxl.
BILAN_ACTIF_LINES = [
    (11, "AD", "IMMOBILISATIONS INCORPORELLES", "3"),
    (12, "AE", "Frais de développement et de prospection", ""),
    (13, "AF", "Brevets, licences, logiciels et droits similaires", ""),
    (14, "AG", "Fonds commercial et droit au bail", ""),
    (15, "AH", "Autres immobilisations incorporelles", ""),
    (16, "AI", "IMMOBILISATIONS CORPORELLES", "3"),
    (17, "AJ", "Terrains", ""),
    (18, "AK", "Bâtiments", ""),
    (19, "AL", "Aménagements, agencements et installations", ""),
    (20, "AM", "Matériel, mobilier et actifs biologiques", ""),
    (21, "AN", "Matériel de transport", ""),
    (22, "AP", "AVANCES ET ACOMPTES VERSES SUR IMMOBILISATIONS", "3"),
    (23, "AQ", "IMMOBILISATIONS FINANCIERES", "4"),
    (24, "AR", "Titres de participation", ""),
    (25, "AS", "Autres immobilisations financières", ""),
    (26, "AZ", "TOTAL ACTIF IMMOBILISE", ""),
    (27, "BA", "ACTIF CIRCULANT HAO", "5"),
    (28, "BB", "STOCKS ET ENCOURS", "6"),
    (29, "BG", "CREANCES ET EMPLOIS ASSIMILES", ""),
    (30, "BH", "Fournisseurs avances versées", "17"),
    (31, "BI", "Clients", "7"),
    (32, "BJ", "Autres créances", "8"),
    (33, "BK", "TOTAL ACTIF CIRCULANT", ""),
    (34, "BQ", "Titres de placement", "9"),
    (35, "BR", "Valeurs à encaisser", "10"),
    (36, "BS", "Banques, chèques postaux, caisse et assimilés", "11"),
    (37, "BT", "TOTAL TRESORERIE-ACTIF", ""),
    (38, "BU", "Ecart de conversion-Actif", "12"),
    (39, "BZ", "TOTAL GENERAL", ""),
]

# BILAN — PASSIF : colonnes REF=J, PASSIF=K, NOTE=L, NET N=M, NET N-1=N.
BILAN_PASSIF_LINES = [
    (11, "CA", "Capital", "13"),
    (12, "CB", "Apporteurs capital non appelé (-)", "13"),
    (13, "CD", "Primes liées au capital social", "14"),
    (14, "CE", "Ecarts de réévaluation", "3e"),
    (15, "CF", "Réserves indisponibles", "14"),
    (16, "CG", "Réserves libres", "14"),
    (17, "CH", "Report à nouveau (+ ou -)", "14"),
    (18, "CJ", "Résultat net de l'exercice (bénéfice + ou perte -)", ""),
    (19, "CL", "Subventions d'investissement", "15"),
    (20, "CM", "Provisions réglementées", "15"),
    (21, "CP", "TOTAL CAPITAUX PROPRES ET RESSOURCES ASSIMILEES", ""),
    (22, "DA", "Emprunts et dettes financières diverses", "16"),
    (23, "DB", "Dettes de location-acquisition", "16"),
    (24, "DC", "Provisions pour risques et charges", "16"),
    (25, "DD", "TOTAL DETTES FINANCIERES ET RESSOURCES ASSIMILEES", ""),
    (26, "DF", "TOTAL RESSOURCES STABLES", ""),
    (27, "DH", "Dettes circulantes HAO", "5"),
    (28, "DI", "Clients, avances reçues", "7"),
    (29, "DJ", "Fournisseurs d'exploitation", "17"),
    (30, "DK", "Dettes fiscales et sociales", "18"),
    (31, "DM", "Autres dettes", "19"),
    (32, "DN", "Provisions pour risques et charges à court terme", "19"),
    (33, "DP", "TOTAL PASSIF CIRCULANT", ""),
    (35, "DQ", "Banques, crédits d'escompte", "20"),
    (36, "DR", "Banques, établissements financiers et crédits de trésorerie", "20"),
    (37, "DT", "TOTAL TRESORERIE-PASSIF", ""),
    (38, "DV", "Ecart de conversion-Passif", "12"),
    (39, "DZ", "TOTAL GENERAL", ""),
]

# COMPTE DE RESULTAT : colonnes REF=A, LIBELLES=B, NOTE=H, EXERCICE N=I,
# EXERCICE N-1=J, lignes 11 à 52 (soldes intermédiaires de gestion XA..XI
# inclus).
RESULTAT_LINES = [
    (11, "TA", "Ventes de marchandises", "21"),
    (12, "RA", "Achats de marchandises", "22"),
    (13, "RB", "Variation de stocks de marchandises", "6"),
    (14, "XA", "MARGE COMMERCIALE (Somme TA à RB)", ""),
    (15, "TB", "Ventes de produits fabriqués", "21"),
    (16, "TC", "Travaux, services vendus", "21"),
    (17, "TD", "Produits accessoires", "21"),
    (18, "XB", "CHIFFRE D'AFFAIRES (A + B + C + D)", ""),
    (19, "TE", "Production stockée (ou déstockage)", "6"),
    (20, "TF", "Production immobilisée", "21"),
    (21, "TG", "Subventions d'exploitation", "21"),
    (22, "TH", "Autres produits", "21"),
    (23, "TI", "Transferts de charges d'exploitation", "12"),
    (24, "RC", "Achats de matières premières et fournitures liées", "22"),
    (25, "RD", "Variation de stocks de matières premières et fournitures liées", "6"),
    (26, "RE", "Autres achats", "22"),
    (27, "RF", "Variation de stocks d'autres approvisionnements", "6"),
    (28, "RG", "Transports", "23"),
    (29, "RH", "Services extérieurs", "24"),
    (30, "RI", "Impôts et taxes", "25"),
    (31, "RJ", "Autres charges", "26"),
    (32, "XC", "VALEUR AJOUTEE (XB +RA+RB) + (somme TE à RJ)", ""),
    (33, "RK", "Charges de personnel", "27"),
    (34, "XD", "EXCEDENT BRUT D'EXPLOITATION (XC+RK)", ""),
    (35, "TJ", "Reprises d'amortissements, provisions et dépréciations", "28"),
    (36, "RL", "Dotations aux amortissements, aux provisions et dépréciations", "3C&28"),
    (37, "XE", "RESULTAT D'EXPLOITATION (XD+TJ+RL)", ""),
    (38, "TK", "Revenus financiers et assimilés", "29"),
    (39, "TL", "Reprises de provisions et dépréciations financières", "28"),
    (40, "TM", "Transferts de charges financières", "12"),
    (41, "RM", "Frais financiers et charges assimilées", "29"),
    (42, "RN", "Dotations aux provisions et aux dépréciations financières", "3C&28"),
    (43, "XF", "RESULTAT FINANCIER (somme TK à RN)", ""),
    (44, "XG", "RESULTAT DES ACTIVITES ORDINAIRES (XE+XF)", ""),
    (45, "TN", "Produits des cessions d'immobilisations", "3D"),
    (46, "TO", "Autres Produits HAO", "30"),
    (47, "RO", "Valeurs comptables des cessions d'immobilisations", "3D"),
    (48, "RP", "Autres Charges HAO", "30"),
    (49, "XH", "RESULTAT HORS ACTIVITES ORDINAIRES (somme TN à RP)", ""),
    (50, "RQ", "Participation des travailleurs", "30"),
    (51, "RS", "Impôts sur le résultat", "37"),
    (52, "XI", "RESULTAT NET (XG+XH+RQ+RS)", ""),
]


def extract_bilan_actif(cache):
    out = []
    for row, ref, label, note in BILAN_ACTIF_LINES:
        out.append({
            "ref": ref, "label": label, "note": note,
            "total": _is_total_label(label),
            "brut": cache.get("F%d" % row),
            "amort": cache.get("G%d" % row),
            "net": cache.get("H%d" % row),
            "net1": cache.get("I%d" % row),
        })
    return out


def extract_bilan_passif(cache):
    out = []
    for row, ref, label, note in BILAN_PASSIF_LINES:
        out.append({
            "ref": ref, "label": label, "note": note,
            "total": _is_total_label(label),
            "net": cache.get("M%d" % row),
            "net1": cache.get("N%d" % row),
        })
    return out


def extract_resultat_full(cache):
    out = []
    for row, ref, label, note in RESULTAT_LINES:
        out.append({
            "ref": ref, "label": label, "note": note,
            "total": _is_total_label(label),
            "n": cache.get("I%d" % row),
            "n1": cache.get("J%d" % row),
        })
    return out


# ----------------------------------------------------------------- views --

def view_login(req, conn):
    error = None
    if req.method == "POST":
        email = req.form.get("email", "").strip().lower()
        password = req.form.get("password", "")
        row = conn.execute("SELECT * FROM users WHERE lower(email)=?", (email,)).fetchone()
        if row and db.verify_password(password, row["password_hash"]):
            token = db.create_session(conn, row["id"])
            dest = "/admin" if row["role"] == "admin" else "/client/%d" % row["client_id"]
            return redirect(dest, set_cookie=token)
        error = "Identifiants invalides."
    return Response(render("login.html", error=error))


def view_logout(req, conn):
    token = req.cookies.get("session")
    db.delete_session(conn, token)
    return redirect("/login", delete_cookie=True)


def view_admin_dashboard(req, conn, user):
    if user["role"] != "admin":
        return Response("Accès refusé", status="403 Forbidden")
    clients = conn.execute("SELECT * FROM clients ORDER BY raison_sociale").fetchall()
    counts = {}
    for c in clients:
        n = conn.execute("SELECT COUNT(*) c FROM exercices WHERE client_id=?", (c["id"],)).fetchone()["c"]
        counts[c["id"]] = n
    return Response(render("admin_dashboard.html", user=user, clients=clients, counts=counts))


def view_client_new(req, conn, user):
    if user["role"] != "admin":
        return Response("Accès refusé", status="403 Forbidden")
    error = None
    if req.method == "POST":
        raison = req.form.get("raison_sociale", "").strip()
        ncc = req.form.get("ncc", "").strip()
        ntd = req.form.get("ntd", "").strip()
        adresse = req.form.get("adresse", "").strip()
        email = req.form.get("email", "").strip().lower()
        password = req.form.get("password", "").strip()
        if not raison or not email or not password:
            error = "Raison sociale, email et mot de passe sont obligatoires."
        else:
            existing = conn.execute("SELECT id FROM users WHERE lower(email)=?", (email,)).fetchone()
            if existing:
                error = "Cet email est déjà utilisé."
            else:
                cur = conn.execute(
                    "INSERT INTO clients (raison_sociale, ncc, ntd, adresse) VALUES (?,?,?,?)",
                    (raison, ncc, ntd, adresse),
                )
                client_id = cur.lastrowid
                db.create_user(conn, email, password, "client", client_id)
                conn.commit()
                return redirect("/client/%d" % client_id)
    return Response(render("client_new.html", user=user, error=error))


def view_client_dashboard(req, conn, user, client_id):
    if user["role"] != "admin" and user["client_id"] != client_id:
        return Response("Accès refusé", status="403 Forbidden")
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    if not client:
        return Response("Client introuvable", status="404 Not Found")
    exercices = conn.execute(
        "SELECT * FROM exercices WHERE client_id=? ORDER BY annee DESC", (client_id,)
    ).fetchall()
    return Response(render("client_dashboard.html", user=user, client=client, exercices=exercices))


def view_exercice_new(req, conn, user, client_id):
    if user["role"] != "admin" and user["client_id"] != client_id:
        return Response("Accès refusé", status="403 Forbidden")
    error = None
    if req.method == "POST":
        try:
            annee = int(req.form.get("annee", "").strip())
        except ValueError:
            annee = None
        libelle = req.form.get("libelle", "").strip()
        date_debut = req.form.get("date_debut", "").strip() or None
        date_fin = req.form.get("date_fin", "").strip() or None
        if not annee:
            error = "Année invalide."
        else:
            try:
                cur = conn.execute(
                    "INSERT INTO exercices (client_id, annee, libelle, date_debut, date_fin) VALUES (?,?,?,?,?)",
                    (client_id, annee, libelle or ("Exercice %d" % annee), date_debut, date_fin),
                )
                conn.commit()
                return redirect("/exercice/%d" % cur.lastrowid)
            except Exception:
                error = "Un exercice pour cette année existe déjà."
    return Response(render("exercice_new.html", user=user, client_id=client_id, error=error))


def view_exercice(req, conn, user, exercice_id):
    exo = conn.execute("SELECT * FROM exercices WHERE id=?", (exercice_id,)).fetchone()
    if not exo:
        return Response("Exercice introuvable", status="404 Not Found")
    client = conn.execute("SELECT * FROM clients WHERE id=?", (exo["client_id"],)).fetchone()
    if user["role"] != "admin" and user["client_id"] != client["id"]:
        return Response("Accès refusé", status="403 Forbidden")

    upload_msg = None
    if req.method == "POST" and req.environ.get("CONTENT_TYPE", "").startswith("multipart/form-data"):
        if b'name="fichier_xlsx_csv"' in req._raw_body:
            zip_resp = handle_xlsx_to_csv(req)
            if zip_resp is not None:
                return zip_resp
            upload_msg = "Fichier xlsx illisible : impossible de générer les CSV."
        else:
            upload_msg = handle_upload(req, conn, exercice_id)
    elif req.method == "POST" and req.form.get("exercice_dates_save"):
        # Dates de debut/fin d'exercice : ne sont pas saisies a la creation de
        # l'exercice (cf. exercice_new.html, qui ne demande que l'annee) -> on
        # permet de les renseigner/corriger ici. Elles alimentent le champ
        # d'entete "Exercice clos le :" / "Durée (en mois) :" sur chaque
        # feuille (cf. classeur Excel, Sommaire!M35/Q35).
        date_debut_in = req.form.get("exercice_date_debut", "").strip()
        date_fin_in = req.form.get("exercice_date_fin", "").strip()
        conn.execute(
            "UPDATE exercices SET date_debut=?, date_fin=? WHERE id=?",
            (date_debut_in or None, date_fin_in or None, exercice_id),
        )
        conn.commit()
        exo = conn.execute("SELECT * FROM exercices WHERE id=?", (exercice_id,)).fetchone()
        upload_msg = "Dates de l'exercice enregistrées."
    elif req.method == "POST" and req.form.get("sommaire_save"):
        checked_keys = [
            key for i, (key, _label, _cid) in enumerate(SOMMAIRE_SHEETS)
            if req.form.get("appl_%d" % i)
        ]
        save_sommaire_selection(conn, exercice_id, checked_keys)
        upload_msg = "Sélection des notes à imprimer (Sommaire) enregistrée."
    elif req.method == "POST" and (req.form.get("note3_sheet") or req.form.get("note_texte_sheet")):
        sheet = req.form.get("note3_sheet") or req.form.get("note_texte_sheet")
        cell_values = {k[5:]: v for k, v in req.form.items() if k.startswith("cell_")}
        champ_values = {k[6:]: v for k, v in req.form.items() if k.startswith("texte_")}
        # Boutons "+ Ajouter une ligne" : name="ajouter_<CLE_COMPTEUR>", incremente
        # le compteur de lignes dynamiques correspondant (voir NOTE 32/33/16C).
        for k, v in req.form.items():
            if k.startswith("ajouter_") and v:
                compteur = k[len("ajouter_"):]
                cell_values[compteur] = ce._num(cell_values.get(compteur, 0)) + 1
        if cell_values:
            save_note3_manual(conn, exercice_id, sheet, cell_values)
        if champ_values:
            save_note_texte(conn, exercice_id, sheet, champ_values)
        upload_msg = "Saisie enregistrée (%s)." % sheet

    # Duree de l'exercice en mois (champ d'entete "Durée (en mois) :" présent
    # sur chaque feuille du classeur Excel, lignes 1-6) : calculee a partir de
    # date_debut/date_fin de l'exercice quand elles sont renseignees.
    exo_duree_mois = ""
    try:
        if exo["date_debut"] and exo["date_fin"]:
            d1 = datetime.strptime(str(exo["date_debut"])[:10], "%Y-%m-%d")
            d2 = datetime.strptime(str(exo["date_fin"])[:10], "%Y-%m-%d")
            mois = (d2.year - d1.year) * 12 + (d2.month - d1.month) + 1
            exo_duree_mois = mois
    except Exception:
        exo_duree_mois = ""

    balN = load_balance(conn, exercice_id, "N")
    balN1 = load_balance(conn, exercice_id, "N1")
    tftn = load_tft_detail(conn, exercice_id, "N")
    tftn1 = load_tft_detail(conn, exercice_id, "N1")

    bilan = ce.compute_sheet("BILAN", balN, balN1) if balN else {}
    resultat = ce.compute_sheet("RESULTAT", balN, balN1) if balN else {}
    tft = ce.compute_sheet("TFT", balN, balN1, tftn, tftn1) if balN else {}

    note3_manual = load_note3_manual(conn, exercice_id)
    note_texte = load_note_texte(conn, exercice_id)
    note3a, note3c, note3cbis, note3d, note3b = compute_note3(balN, note3_manual)
    for r in note3b:
        r["nature_contrat"] = note_texte.get("NOTE 3B", {}).get("nature_contrat_%d" % r["row"], "")
    note3e = compute_note3e(note3_manual, note_texte)
    notes_lot_a = compute_notes_lot_a(balN, balN1, note3_manual)
    notes_lot_b = compute_notes_lot_b(balN, balN1, note3_manual)
    notes_lot_c = compute_notes_lot_c(balN, balN1, note3_manual)
    tables_lot_d = compute_notes_lot_d(balN, balN1, note3_manual)
    note31 = compute_note31(balN, balN1, note3_manual, resultat)
    note34 = compute_note34(bilan, resultat, tft, balN, balN1, tftn, tftn1)
    note1 = compute_note1(note3_manual)
    note32 = compute_note32(note3_manual, note_texte)
    note33 = compute_note33(note3_manual, note_texte)
    note16c = compute_note16c(note3_manual, note_texte)
    note30 = compute_note30(note3_manual, note_texte)
    note27b_rows, note27b_side = compute_note27b(note3_manual)
    note13 = compute_note13(note3_manual, note_texte)
    note21 = compute_note21(balN, balN1, note3_manual, note_texte)

    bilan_actif = extract_bilan_actif(bilan)
    bilan_passif = extract_bilan_passif(bilan)
    resultat_lines = extract_resultat_full(resultat)
    tft_lines = extract_lines(tft, "TFT", [
        ("Trésorerie nette au 1er janvier (A)", 10, "I", "J"),
        ("Capacité d'Autofinancement Globale (CAFG)", 12, "I", "J"),
        ("Variation Actif circulant HAO", 13, "I", "J"),
        ("Variation des stocks", 14, "I", "J"),
        ("Variation des créances", 15, "I", "J"),
        ("Variation du passif circulant", 16, "I", "J"),
        ("Flux de trésorerie des activités opérationnelles (B)", 19, "I", "J"),
        ("Décaissements liés aux immobilisations", 21, "I", "J"),
        ("Encaissements liés aux cessions d'immobilisations", 22, "I", "J"),
        ("Décaissements liés aux acquisitions financières", 23, "I", "J"),
        ("Encaissements liés aux cessions financières", 24, "I", "J"),
        ("Variation des autres actifs immobilisés", 25, "I", "J"),
        ("Flux de trésorerie des activités d'investissement (C)", 26, "I", "J"),
        ("Augmentations de capital par apports nouveaux", 28, "I", "J"),
        ("Subventions d'investissement reçues", 29, "I", "J"),
        ("Prélèvements sur le capital", 30, "I", "J"),
        ("Distributions de dividendes versées", 31, "I", "J"),
        ("Flux activités de financement par capitaux propres (D)", 32, "I", "J"),
        ("Emprunts et autres dettes financières", 34, "I", "J"),
        ("Remboursements des emprunts et autres dettes financières", 35, "I", "J"),
        ("Variation des autres dettes financières", 36, "I", "J"),
        ("Flux activités de financement par capitaux étrangers (E)", 37, "I", "J"),
        ("Flux de trésorerie provenant des activités de financement (D+E=F)", 38, "I", "J"),
        ("Variation de la trésorerie nette de la période (B+C+F)", 39, "I", "J"),
        ("Trésorerie nette au 31 décembre", 40, "I", "J"),
    ])

    comp_tva = compute_comp_tva(note3_manual)
    comp_tva2 = compute_comp_tva2(note3_manual)
    suppl1 = compute_suppl1(note3_manual, note_texte)
    suppl2 = compute_suppl2(note3_manual, note_texte)
    suppl3 = compute_suppl3(note3_manual)
    suppl4 = compute_suppl4(note3_manual, note_texte)
    suppl5 = compute_suppl5(note3_manual)
    suppl6 = compute_suppl6(note3_manual)
    suppl7 = compute_suppl7(note3_manual)
    comp_charges = compute_comp_charges(balN, balN1)
    notes_dgi_ins_recap = compute_notes_dgi_ins_recap(
        comp_charges, comp_tva, comp_tva2, suppl1, suppl2, suppl3, suppl4,
        suppl5, suppl6, suppl7, note3_manual, note_texte,
    )

    # SOMMAIRE — selection manuelle des feuilles a imprimer (cf. classeur
    # Excel, feuille Sommaire, cellule O17 : "seules les feuilles cochees ici
    # seront imprimees"). FICHE R4 reprend exactement ces memes cases sous
    # forme de colonnes A / N-A (cf. formules =IF(Sommaire!xx=TRUE,"x","")).
    sommaire_sel = load_sommaire_selection(conn, exercice_id)
    hidden_card_ids = [cid for key, _label, cid in SOMMAIRE_SHEETS if not sommaire_sel.get(key, True)]
    fiche_r4_rows = [
        {"key": key, "label": label, "applicable": sommaire_sel.get(key, True)}
        for key, label, _cid in SOMMAIRE_SHEETS
    ]
    garde = note_texte.get("GARDE", {})
    sommaire_info = note_texte.get("SOMMAIRE", {})
    fiche_r1 = note_texte.get("FICHE R1", {})
    fiche_r2 = note_texte.get("FICHE R2", {})
    fiche_r3 = note_texte.get("FICHE R3", {})

    # FICHE R1 — "Domiciliations bancaires" : tableau a lignes dynamiques
    # (Banque / N° de compte), meme mecanisme "+ Ajouter une ligne" que les
    # autres notes (compteur "_N_BANQUE" stocke en note3_manual, valeurs en
    # texte libre stockees en note_texte).
    n_banques = int(ce._num(note3_manual.get("FICHE R1", {}).get("_N_BANQUE", 1))) or 1
    fiche_r1_banques = [
        {
            "idx": i,
            "banque": fiche_r1.get("banque_nom_%d" % i, ""),
            "compte": fiche_r1.get("banque_compte_%d" % i, ""),
        }
        for i in range(n_banques)
    ]

    # FICHE R3 — "Dirigeants" et "Membres du Conseil d'Administration" :
    # mêmes tableaux à lignes dynamiques que la Fiche R1 (Banque/N° compte),
    # un compteur "+ Ajouter une ligne" indépendant par sous-groupe.
    n_dirigeants = int(ce._num(note3_manual.get("FICHE R3", {}).get(
        "_N_DIRIGEANT", FICHE_R3_DIRIGEANTS_ROWS))) or FICHE_R3_DIRIGEANTS_ROWS
    n_admin = int(ce._num(note3_manual.get("FICHE R3", {}).get(
        "_N_ADMIN", FICHE_R3_ADMIN_ROWS))) or FICHE_R3_ADMIN_ROWS
    fiche_r3_dirigeants_rows = range(1, n_dirigeants + 1)
    fiche_r3_admin_rows = range(1, n_admin + 1)

    return Response(render(
        "exercice_view.html", user=user, client=client, exo=exo, exo_duree_mois=exo_duree_mois,
        balN=balN, balN1=balN1, tftn=tftn, tftn1=tftn1, upload_msg=upload_msg,
        bilan_actif=bilan_actif, bilan_passif=bilan_passif, resultat_lines=resultat_lines,
        tft_lines=tft_lines,
        note3a=note3a, note3a_cols=NOTE3A_COLS,
        note3c=note3c, note3c_cols=NOTE3C_COLS,
        note3cbis=note3cbis, note3cbis_cols=NOTE3CBIS_COLS,
        note3d=note3d, note3d_cols=NOTE3D_COLS,
        note3b=note3b, note3b_cols=NOTE3B_COLS,
        note3e=note3e,
        notes_lot_a=notes_lot_a, notes_lot_a_defs=NOTES_LOT_A,
        notes_lot_b=notes_lot_b, notes_lot_b_defs=NOTES_LOT_B,
        notes_lot_c=notes_lot_c, notes_lot_c_defs=NOTES_LOT_C,
        tables_lot_d=tables_lot_d,
        note31=note31, note31_cols=NOTE31_COLS, note31_col_labels=NOTE31_COL_LABELS,
        note34=note34,
        note1=note1,
        note32=note32, note33=note33, note16c=note16c, note30=note30,
        note27b_rows=note27b_rows, note27b_cols=NOTE27B_COLS, note27b_side=note27b_side,
        note13=note13,
        note21=note21,
        note_texte=note_texte, notes_texte_defs=NOTES_TEXTE_DEFS,
        commentaire_defs=COMMENTAIRE_DEFS,
        commentaire_map={k: fields for k, _label, fields in COMMENTAIRE_DEFS},
        comp_tva=comp_tva,
        comp_tva2=comp_tva2,
        suppl1=suppl1,
        suppl2=suppl2,
        suppl3=suppl3,
        suppl4=suppl4,
        suppl5=suppl5,
        suppl6=suppl6,
        suppl7=suppl7,
        comp_charges=comp_charges,
        notes_dgi_ins_recap=notes_dgi_ins_recap,
        sommaire_sel=sommaire_sel, sommaire_sheets=SOMMAIRE_SHEETS,
        hidden_card_ids=hidden_card_ids, fiche_r4_rows=fiche_r4_rows,
        garde=garde, garde_documents=GARDE_DOCUMENTS,
        sommaire_info=sommaire_info, sommaire_champs=SOMMAIRE_CHAMPS,
        fiche_r1=fiche_r1, fiche_r1_champs=FICHE_R1_CHAMPS, fiche_r1_banques=fiche_r1_banques,
        fiche_r2=fiche_r2, fiche_r3=fiche_r3,
        fiche_r2_activites_rows=range(1, FICHE_R2_ACTIVITES_ROWS + 1),
        fiche_r3_dirigeants_rows=fiche_r3_dirigeants_rows,
        fiche_r3_admin_rows=fiche_r3_admin_rows,
    ))


def handle_upload(req, conn, exercice_id):
    body = req._raw_body
    ctype = req.environ.get("CONTENT_TYPE", "")
    boundary = ctype.split("boundary=")[-1].encode()
    parts = body.split(b"--" + boundary)
    periode = "N"
    upload_type = "balance"
    file_bytes = None
    for part in parts:
        if b'name="periode"' in part:
            periode = part.split(b"\r\n\r\n", 1)[-1].strip(b"\r\n -").decode("utf-8", "replace") or "N"
        if b'name="type"' in part:
            upload_type = part.split(b"\r\n\r\n", 1)[-1].strip(b"\r\n -").decode("utf-8", "replace") or "balance"
        if b"filename=" in part and (
            b'name="fichier"' in part or b'name="fichier_tft"' in part or b'name="fichier_xlsx"' in part
        ):
            header, _, content = part.partition(b"\r\n\r\n")
            file_bytes = content.rstrip(b"\r\n--")
            if b'name="fichier_tft"' in part:
                upload_type = "tft_detail"
            elif b'name="fichier_xlsx"' in part:
                upload_type = "xlsx_import"
    if not file_bytes:
        return "Aucun fichier reçu."

    if upload_type == "xlsx_import":
        try:
            parsed = parse_xlsx_import(file_bytes)
        except Exception as exc:
            return "Fichier xlsx illisible : %s" % exc
        nb_bal, nb_tft = 0, 0
        for p, rows in parsed["balance"].items():
            conn.execute("DELETE FROM balance_lignes WHERE exercice_id=? AND periode=?", (exercice_id, p))
            for r in rows:
                conn.execute(
                    "INSERT INTO balance_lignes (exercice_id, periode, compte, designation, be_debit, be_credit, "
                    "mvt_debit, mvt_credit, bs_debit, bs_credit) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (exercice_id, p, r["compte"], r["designation"],
                     ce._num(r["be_debit"]), ce._num(r["be_credit"]),
                     ce._num(r["mvt_debit"]), ce._num(r["mvt_credit"]),
                     ce._num(r["bs_debit"]), ce._num(r["bs_credit"])),
                )
            nb_bal += len(rows)
        for p, rows in parsed["tft"].items():
            conn.execute("DELETE FROM tft_detail_lignes WHERE exercice_id=? AND periode=?", (exercice_id, p))
            for r in rows:
                conn.execute(
                    "INSERT INTO tft_detail_lignes (exercice_id, periode, compte, designation, be_debit, be_credit, "
                    "mvt_debit, mvt_credit, bs_debit, bs_credit) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (exercice_id, p, r["compte"], r["designation"],
                     ce._num(r["be_debit"]), ce._num(r["be_credit"]),
                     ce._num(r["mvt_debit"]), ce._num(r["mvt_credit"]),
                     ce._num(r["bs_debit"]), ce._num(r["bs_credit"])),
                )
            nb_tft += len(rows)
        conn.commit()
        alerts = compute_controles_xlsx(parsed)
        msg = ("Fichier xlsx importé : %d ligne(s) de balance (N+N-1), %d ligne(s) de détail TFT (N+N-1)."
               % (nb_bal, nb_tft))
        if alerts:
            msg += " ⚠️ Anomalies détectées — " + " | ".join(alerts)
        else:
            msg += " Aucune anomalie détectée par les contrôles intégrés (équilibre, comptes à éclater)."
        return msg

    periode = "N1" if "N1" in periode else "N"
    rows = parse_csv_balance(file_bytes)
    if not rows:
        return "Fichier CSV vide ou illisible."

    table = "tft_detail_lignes" if upload_type == "tft_detail" else "balance_lignes"
    conn.execute("DELETE FROM %s WHERE exercice_id=? AND periode=?" % table, (exercice_id, periode))
    for r in rows:
        conn.execute(
            "INSERT INTO %s (exercice_id, periode, compte, designation, be_debit, be_credit, "
            "mvt_debit, mvt_credit, bs_debit, bs_credit) VALUES (?,?,?,?,?,?,?,?,?,?)" % table,
            (
                exercice_id, periode, r["compte"], r["designation"],
                ce._num(r["be_debit"]), ce._num(r["be_credit"]),
                ce._num(r["mvt_debit"]), ce._num(r["mvt_credit"]),
                ce._num(r["bs_debit"]), ce._num(r["bs_credit"]),
            ),
        )
    conn.commit()
    label = "détail TFT" if upload_type == "tft_detail" else "balance"
    return "%d lignes (%s) importées pour la période %s." % (len(rows), label, periode)


NOT_FOUND = Response("Page introuvable", status="404 Not Found")

# Fichiers statiques téléchargeables depuis l'application (liste blanche par
# sécurité : on ne sert que ces fichiers connus, jamais un chemin arbitraire).
STATIC_FILES = {
    "TAFIROHA-DGI_Import.xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def view_static(req, conn, filename):
    ctype = STATIC_FILES.get(filename)
    if not ctype:
        return NOT_FOUND
    path = os.path.join(STATIC_DIR, filename)
    if not os.path.isfile(path):
        return NOT_FOUND
    with open(path, "rb") as f:
        data = f.read()
    return Response(
        data, content_type=ctype,
        headers=[("Content-Disposition", 'attachment; filename="%s"' % filename)],
    )


def dispatch(req, conn):
    path = req.path.rstrip("/") or "/"

    if path == "/login":
        return view_login(req, conn)

    user = require_login(req, conn)

    if path == "/logout":
        return view_logout(req, conn)

    if not user:
        return redirect("/login")

    if path.startswith("/static/"):
        return view_static(req, conn, path[len("/static/"):])

    if path == "/" or path == "":
        return redirect("/admin" if user["role"] == "admin" else "/client/%d" % user["client_id"])

    if path == "/admin":
        return view_admin_dashboard(req, conn, user)

    if path == "/admin/clients/new":
        return view_client_new(req, conn, user)

    if path.startswith("/client/") and path.endswith("/exercices/new"):
        client_id = int(path.split("/")[2])
        return view_exercice_new(req, conn, user, client_id)

    if path.startswith("/client/"):
        client_id = int(path.split("/")[2])
        return view_client_dashboard(req, conn, user, client_id)

    if path.startswith("/exercice/"):
        exercice_id = int(path.split("/")[2])
        return view_exercice(req, conn, user, exercice_id)

    return NOT_FOUND


def application(environ, start_response):
    req = Request(environ)
    conn = db.get_conn()
    try:
        resp = dispatch(req, conn)
    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        resp = Response("<pre>%s</pre>" % tb, status="500 Internal Server Error")
    finally:
        conn.close()
    start_response(resp.status, resp.headers)
    return [resp.body]


if __name__ == "__main__":
    db.init_db()
    port = int(os.environ.get("PORT", "8000"))
    print("TAFIROHA en ligne — http://127.0.0.1:%d  (login: admin@tafiroha.local / admin1234)" % port)
    make_server("0.0.0.0", port, application).serve_forever()
